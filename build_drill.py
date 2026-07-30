# -*- coding: utf-8 -*-
"""
Расчётная модель TCO буровых станков (методика аннуитета руб/пог.м и руб/м³).
Полный погодовой дисконтированный жизненный цикл (пост-налоговый), построенный по
аналогии с моделью экскаваторов, но с расчётом ПРОИЗВОДИТЕЛЬНОСТИ по методике из
корпоративного файла буровых станков (000_БСЛ-2027):

  ПРОИЗВОДИТЕЛЬНОСТЬ (физика бурения):
    - теоретическая скорость бурения (п.м/час) = (0,1·N_вращ)/(2,73^(0,017·σ)·D_заряда²);
      D_заряда = D_долота·k_разбура; N_вращ — мощность вращателя;
    - время цикла скважины = бурение + продувка + наращивание + подъём + разбор + переезд;
    - скважин/смену → часовая производительность (п.м/час);
    - КТГ → КИО = КТГ·(1 − простои_смены/(60·смена)) → эффективное время = КФВ·КИО;
    - годовой объём бурения (тыс.пог.м) и взрывания (тыс.м³ = тыс.пог.м · выход м³/пог.м).

  ЭКОНОМИКА (погодовой движок, 10 лет, пост-налоговый):
    топливо/энергия, ТОиР (ТО+ТР+капремонты+сервис — по годам из листов поставщиков),
    буровой инструмент (долота, штанги), расходники, персонал, транспортный налог,
    амортизация (НУ), налог на прибыль (щит), остаточная стоимость, дисконтирование.

  РЕЗУЛЬТАТ: удельный аннуитет руб/пог.м (главный критерий) и руб/м³ (по взрыванию),
    операционный + инвестиционный, NPV владения, дашборд, сравнение, чувствительность,
    аналитика. Заложен пример: 5 буровых станков класса 250–270 мм (дизель).

Файл: Модель_TCO_буровых_станков_аннуитет.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef
from openpyxl.worksheet.datavalidation import DataValidation

S_MET="Методика"; S_DASH="Дашборд"; S_PAR="Параметры"; S_IN="Ввод данных"
S_PYR="Данные по годам"
S_PROD="Производительность"; S_CF="Денежный поток"; S_ANN="Расчёт аннуитета"
S_CMP="Сравнение"; S_ANL="Аналитика"; S_SENS="Чувствительность"; S_REF="Справочник"

MACH=["D","E","F","G","H"]; NM=5
LC=MACH[-1]                                  # последняя колонка машины (H для 5)
LCN=3+NM                                     # номер последней колонки (8)
H=10                                         # горизонт, лет (периоды 0..10)
PCOLS=[get_column_letter(4+p) for p in range(H+1)]   # D..N (период 0..10)

# ---- палитра ----
DARK="1F3864"; MID="2E5496"; ACC="C55A11"; INP="FFF2CC"; RES="E2EFDA"; HEAD="D6DCE5"
GREY="F2F2F2"; BEST="C6EFCE"; TILE="1F3864"
Fd=PatternFill("solid",fgColor=DARK); Fm=PatternFill("solid",fgColor=MID); Fi=PatternFill("solid",fgColor=INP)
Fr=PatternFill("solid",fgColor=RES); Fh=PatternFill("solid",fgColor=HEAD); Fg=PatternFill("solid",fgColor=GREY)
Fb=PatternFill("solid",fgColor=BEST); Ftile=PatternFill("solid",fgColor=TILE); Fa=PatternFill("solid",fgColor="FCE4D6")
FT=Font(size=15,bold=True,color="FFFFFF"); FSub=Font(size=10,italic=True,color="FFFFFF")
FSec=Font(size=11,bold=True,color="FFFFFF"); FHd=Font(size=9,bold=True,color=DARK)
FB=Font(size=10,bold=True); FN=Font(size=10); FRz=Font(size=10,bold=True,color=DARK); FU=Font(size=8,italic=True,color="808080")
thin=Side(style="thin",color="BFBFBF"); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
L=Alignment("left",vertical="center",wrap_text=True); C=Alignment("center",vertical="center",wrap_text=True); Rr=Alignment("right",vertical="center")
M="#,##0"; M1="#,##0.0"; M2="#,##0.00"; P1="0.0%"; NUM="#,##0.000"

wb=Workbook()

def title(ws,t,s,last="K"):
    ws.merge_cells(f"B1:{last}1"); ws["B1"]=t; ws["B1"].font=FT; ws["B1"].alignment=L
    ws.merge_cells(f"B2:{last}2"); ws["B2"]=s; ws["B2"].font=FSub; ws["B2"].alignment=L
    a1=ws[f"{last}1"].column
    for r in (1,2):
        for col in range(2,a1+1): ws.cell(row=r,column=col).fill=Fd
def section(ws,row,text,c0="B",c1="K"):
    ws.merge_cells(f"{c0}{row}:{c1}{row}"); cc=ws[f"{c0}{row}"]; cc.value=text; cc.font=FSec; cc.alignment=L
    a0=ws[f"{c0}1"].column; a1=ws[f"{c1}1"].column
    for col in range(a0,a1+1): ws.cell(row=row,column=col).fill=Fm
def PVAF(r,n): return f"IF({r}=0,{n},(1-(1+{r})^-{n})/{r})"

# ================================================================= #
# ПАРАМЕТРЫ
# ================================================================= #
wp=wb.active; wp.title=S_PAR; wp.sheet_view.showGridLines=False
title(wp,"ГЛОБАЛЬНЫЕ ПАРАМЕТРЫ МОДЕЛИ","Единые допущения для всех вариантов. Жёлтые ячейки — ввод.",last="F")
wp.column_dimensions["A"].width=2; wp.column_dimensions["B"].width=48
wp.column_dimensions["C"].width=13; wp.column_dimensions["D"].width=12; wp.column_dimensions["E"].width=54
def phdr(row):
    for col,t in (("B","Параметр"),("C","Значение"),("D","Ед."),("E","Комментарий")):
        c=wp[f"{col}{row}"]; c.value=t; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
prow={}
def padd(row,key,label,val,unit,fmt,comment):
    prow[key]=row
    wp[f"B{row}"]=label; wp[f"B{row}"].font=FN; wp[f"B{row}"].border=BD; wp[f"B{row}"].alignment=L
    c=wp[f"C{row}"]; c.value=val; c.fill=Fi; c.border=BD; c.alignment=C; c.font=FB
    if fmt: c.number_format=fmt
    wp[f"D{row}"]=unit; wp[f"D{row}"].font=FU; wp[f"D{row}"].alignment=C; wp[f"D{row}"].border=BD
    wp[f"E{row}"]=comment; wp[f"E{row}"].font=FU; wp[f"E{row}"].alignment=L; wp[f"E{row}"].border=BD
def pcalc(row,key,label,formula,unit,fmt,comment):
    prow[key]=row
    wp[f"B{row}"]=label; wp[f"B{row}"].font=FN; wp[f"B{row}"].border=BD; wp[f"B{row}"].alignment=L
    c=wp[f"C{row}"]; c.value="="+formula; c.fill=Fr; c.border=BD; c.alignment=C; c.font=FRz
    if fmt: c.number_format=fmt
    wp[f"D{row}"]=unit; wp[f"D{row}"].font=FU; wp[f"D{row}"].alignment=C; wp[f"D{row}"].border=BD
    wp[f"E{row}"]=comment; wp[f"E{row}"].font=FU; wp[f"E{row}"].alignment=L; wp[f"E{row}"].border=BD
def pc(key): return f"'{S_PAR}'!C{prow[key]}"

section(wp,4,"Финансовые допущения","B","E"); phdr(5)
padd(6,"disc","Ставка дисконтирования (WACC)",0.12,"%/год",P1,"Стоимость капитала для приведения к текущей стоимости")
padd(7,"tax","Налог на прибыль",0.25,"%",P1,"Формирует налоговый щит на затраты и амортизацию")
padd(8,"horizon","Срок эксплуатации (горизонт)",10,"лет",M,"Фиксирован = 10 лет в погодовом движке")
padd(9,"dep_years","Срок амортизации (налоговый учёт)",5.083,"лет",M2,"Линейная; 61 мес ≈ 5,08 года (как в исходной модели)")
padd(10,"fuel_p","Цена дизельного топлива",66.019,"тыс.руб/т","#,##0.000","СЛ 2025: 1 274 749,9 тыс.руб / 19 308,86 т")
padd(11,"el_p","Цена электроэнергии",5.133,"руб/кВт·ч","#,##0.000","Для электрических станков (СЛ 2025)")

section(wp,13,"Эскалация цен (реальный/номинальный рост)","B","E"); phdr(14)
padd(15,"esc_fuel","Эскалация цены топлива/энергии",0.0,"%/год",P1,"0 = данные уже с учётом инфляции")
padd(16,"esc_parts","Эскалация ЗЧ/инструмента/расходников",0.0,"%/год",P1,"ТОиР задаётся по годам напрямую")
padd(17,"esc_lab","Эскалация оплаты труда",0.0,"%/год",P1,"Рост ФОТ экипажей")

section(wp,19,"Курсы валют","B","E"); phdr(20)
padd(21,"cny","Курс юаня (CNY)",11.0,"руб",NUM,"Для оборудования из Китая")
padd(22,"usd","Курс доллара (USD)",80.0,"руб",NUM,"")
padd(23,"eur","Курс евро (EUR)",90.0,"руб",NUM,"")

section(wp,25,"Параметры производительности (карьер / буровзрывные работы)","B","E"); phdr(26)
padd(27,"kfv","Календарный фонд времени",8760,"час/год",M,"24 ч × 365 дн")
padd(28,"shift","Длительность смены",12,"час",M,"Для пересчёта простоев и производительности")
padd(29,"strength","Предел прочности породы на сжатие",130,"МПа",M,"Влияет на скорость бурения (крепость f)")
padd(30,"razbur","Коэффициент разбура",1.05,"коэф.",M2,"Диаметр заряда = диаметр долота × k_разбура")
padd(31,"depth","Средняя глубина скважин",16.16,"м",M2,"(гл.руда×доля + гл.вскрыша×доля); из «W час»")
padd(32,"hoist_eff","КПД скорости подъёма става",0.8,"коэф.",M2,"Фактическая скорость = паспортная × КПД")
padd(33,"blow","Время продувки скважины",2.0,"мин",M1,"Часть цикла бурения одной скважины")
padd(34,"tram","Время переезда между скважинами",2.5,"мин",M1,"Часть цикла бурения одной скважины")
padd(35,"yield","Выход горной массы с 1 пог.м скважины",49.465,"м³/пог.м","#,##0.000","Для пересчёта руб/пог.м → руб/м³ (сетка БВР)")

section(wp,37,"Расчётные финансовые коэффициенты","B","E"); phdr(38)
pcalc(39,"S","Коэф. аннуитета Σдиск.факторов (r, N)",PVAF(pc('disc'),pc('horizon')),"коэф.",NUM,"Сумма дисконт-факторов за срок службы")
_d=pc('dep_years'); _r=pc('disc')
_dsf=f"(1/{_d})*((1-(1+{_r})^-INT({_d}))/{_r}+({_d}-INT({_d}))*(1+{_r})^-(INT({_d})+1))"
pcalc(40,"dsf","Коэф. приведённой амортизации",_dsf,"коэф.",NUM,"Приведённая амортизация на 1 руб. цены (налог. щит)")

lg=42
wp[f"B{lg}"]="Легенда:"; wp[f"B{lg}"].font=FB
wp[f"C{lg}"].fill=Fi; wp[f"C{lg}"].border=BD; wp[f"D{lg}"]="— ввод данных"; wp[f"D{lg}"].font=FN; wp.merge_cells(f"D{lg}:E{lg}")
wp[f"C{lg+1}"].fill=Fr; wp[f"C{lg+1}"].border=BD; wp[f"D{lg+1}"]="— расчётные показатели"; wp[f"D{lg+1}"].font=FN; wp.merge_cells(f"D{lg+1}:E{lg+1}")
print("Параметры:",len(prow))

# ================================================================= #
# ВВОД ДАННЫХ
# ================================================================= #
names=["YANGZI DR75\n(аналог Epiroc DM75)","Sunward\nSWDRT250SM","Sunward\nSWDRT270SM",
       "Sunward\nSWDRT270B","HIFO 75D\n(аналог DM75)"]
supplier=["Огунвей","Стройкомпозит-Н","Стройкомпозит-Н","Огунвей","Майнинг Солюшнс"]
engines=["Cummins QSK19","CAT C27","CAT C27","CAT C27","CAT C27"]
# цена — В ФАЙЛЕ НЕ УКАЗАНА (строка «Буровой станок» инвестиций пуста): оценка, УТОЧНИТЬ
price_val=[1900,1400,1700,1700,12000]
price_cur=["USD","USD","USD","USD","CNY"]
install=[0]*NM; resid=[0.0]*NM
# физика бурения (лист «W час» исходного файла)
rot_pow=[136,148.3,173.4,173.4,136]          # мощность вращателя, кВт (в расчёте)
bit_dia=[0.267,0.267,0.267,0.267,0.267]      # диаметр долота (в расчёте), м
hoist_v=[51.5,60,45,45,51.5]                 # скорость подъёма става, м/мин
rod_add=[1.5,1.5,1.5,1.5,1.5]                # время на наращивание/разбор става, мин
spins=[1,1,1,1,1]                            # количество операций свинчивания-развинчивания
# персонал (ФОТ+отчисления+прочее), тыс.руб/год — дизельный экипаж (3 машиниста)
personnel=[15710,15710,15710,15710,15710]
eng_hp=[800,800,800,800,811.69]              # мощность двигателя, л.с. (транспортный налог)
tax_rate_hp=15                               # ставка транспортного налога, руб/л.с.

wi=wb.create_sheet(S_IN); wi.sheet_view.showGridLines=False
title(wi,"ВВОД ИСХОДНЫХ ДАННЫХ ПО БУРОВЫМ СТАНКАМ","Сравнение до 5 машин одного класса. Пример: станки класса 250–270 мм (дизель), карьер «Сухой Лог».")
wi.column_dimensions["A"].width=2; wi.column_dimensions["B"].width=44; wi.column_dimensions["C"].width=12
for cc in MACH: wi.column_dimensions[cc].width=16
wi.column_dimensions["K"].width=34
wi.freeze_panes="D5"
HR=4
wi[f"B{HR}"]="Показатель"; wi[f"B{HR}"].font=FHd; wi[f"B{HR}"].fill=Fh; wi[f"B{HR}"].border=BD; wi[f"B{HR}"].alignment=C
wi[f"C{HR}"]="Ед."; wi[f"C{HR}"].font=FHd; wi[f"C{HR}"].fill=Fh; wi[f"C{HR}"].border=BD; wi[f"C{HR}"].alignment=C
for i,cc in enumerate(MACH):
    c=wi[f"{cc}{HR}"]; c.value=f"Станок {i+1}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
wi[f"K{HR}"]="Пояснение"; wi[f"K{HR}"].font=FHd; wi[f"K{HR}"].fill=Fh; wi[f"K{HR}"].border=BD; wi[f"K{HR}"].alignment=C

LAY=[
 ("SEC","A. ИДЕНТИФИКАЦИЯ"),
 ("R","name","Наименование / модель","—",names,"Название варианта",True,None),
 ("R","supplier","Поставщик","—",supplier,"",True,None),
 ("R","engine","Двигатель / привод","—",engines,"",True,None),
 ("SEC","B. СТОИМОСТЬ И ОСТАТОЧНАЯ ЦЕНА"),
 ("R","price_val","Цена за единицу","тыс.вал.",price_val,"⚠ В файле поставщика цена не указана — ОЦЕНКА, уточнить",False,M),
 ("R","price_cur","Валюта цены","—",price_cur,"CNY / USD / EUR / RUB",True,None),
 ("R","install","Монтаж, доставка, ПНР","тыс.руб",install,"Дополнительно к цене",False,M),
 ("R","resid","Остаточная стоимость","% от цены",resid,"Реализация в конце срока (0 = не учитывать)",False,P1),
 ("SEC","C. ПРОИЗВОДИТЕЛЬНОСТЬ (физика бурения — лист «W час»)"),
 ("R","rot_pow","Мощность вращателя (в расчёте)","кВт",rot_pow,"Ключевой драйвер скорости бурения",False,M1),
 ("R","bit_dia","Диаметр долота (в расчёте)","м",bit_dia,"Диаметр заряда = диаметр × k_разбура",False,NUM),
 ("R","hoist_v","Скорость подъёма става","м/мин",hoist_v,"Влияет на время СПО",False,M1),
 ("R","rod_add","Время наращивания/разбора става","мин",rod_add,"На одну операцию",False,M1),
 ("R","spins","Кол-во операций свинчивания","ед.",spins,"Штанг на скважину − 1",False,M),
 ("SEC","D. ПРОЧИЕ ГОДОВЫЕ ЗАТРАТЫ И ПАРАМЕТРЫ"),
 ("R","personnel","Расходы на персонал (экипаж)","тыс.руб/год",personnel,"ФОТ+отчисления+прочее на 1 машину",False,M),
 ("R","eng_hp","Мощность двигателя (транспортный налог)","л.с.",eng_hp,"× ставку налога (руб/л.с.)",False,M1),
 ("NOTE","КТГ, простои, затраты на ТОиР (ТО/ТР/капремонты/сервис), удельный расход ДТ, буровой инструмент и расходники "
         "задаются ПО ГОДАМ на листах «Поставщик 1…5». Сводка — на листе «Данные по годам»."),
]
irow={}; r=HR+2
for it in LAY:
    if it[0]=="SEC": section(wi,r,it[1],"B","K"); r+=1; continue
    if it[0]=="NOTE":
        wi.merge_cells(f"B{r}:K{r}"); nc=wi[f"B{r}"]; nc.value="ℹ  "+it[1]; nc.font=Font(size=9,bold=True,color=ACC); nc.alignment=L
        nc.fill=Fa; r+=1; continue
    _,key,label,unit,vals,comment,istext,fmt=it
    irow[key]=r
    wi[f"B{r}"]=label; wi[f"B{r}"].font=FN; wi[f"B{r}"].border=BD; wi[f"B{r}"].alignment=L
    wi[f"C{r}"]=unit; wi[f"C{r}"].font=FU; wi[f"C{r}"].border=BD; wi[f"C{r}"].alignment=C
    for i,cc in enumerate(MACH):
        c=wi[f"{cc}{r}"]; c.value=vals[i]; c.fill=Fi; c.border=BD; c.font=FN
        c.alignment=L if istext else C
        if fmt: c.number_format=fmt
    wi[f"K{r}"]=comment; wi[f"K{r}"].font=FU; wi[f"K{r}"].border=BD; wi[f"K{r}"].alignment=L
    r+=1
dv=DataValidation(type="list",formula1='"CNY,USD,EUR,RUB"',allow_blank=False); wi.add_data_validation(dv)
dv.add(f"D{irow['price_cur']}:{LC}{irow['price_cur']}")
def inp(key,col): return f"'{S_IN}'!{col}{irow[key]}"
print("Ввод:",len(irow))

# ================================================================= #
# ЛИСТЫ-ШАБЛОНЫ ПОСТАВЩИКОВ (по одному на машину) — ТО и Ремонты ПО ГОДАМ
# ================================================================= #
# Реальные КТГ "в расчёте" (прямой ввод — управляет производительностью), 10 лет.
KTG_Y={
 0:[0.9309,0.9172,0.9309,0.9172,0.8865,0.9172,0.8506,0.9198,0.9283,0.8532],
 1:[0.8948,0.8654,0.8454,0.8748,0.8654,0.8432,0.8621,0.8654,0.8498,0.8543],
 2:[0.8948,0.8654,0.8454,0.8748,0.8654,0.8432,0.8621,0.8654,0.8498,0.8543],
 3:[0.93,0.93,0.92,0.92,0.87,0.90,0.85,0.90,0.90,0.85],
 4:[0.8932,0.867,0.8536,0.8651,0.8702,0.8443,0.8584,0.8706,0.8529,0.8527]}
DT_REM={  # простои в ремонтах, ч/год (справочно; КТГ вводится напрямую)
 0:[200,300,200,300,600,300,900,300,200,900],
 1:[425.645,690.645,870.645,605.645,690.645,890.645,720.645,690.645,830.645,790.645],
 2:[425.645,690.645,870.645,605.645,690.645,890.645,720.645,690.645,830.645,790.645],
 3:[109.67,109.67,109.67,109.67,109.67,109.67,109.67,109.67,109.67,109.67],
 4:[85.0,315.66,423.65,333.06,283.55,510.36,398.41,274.9,438.21,436.2]}
FUEL_Y={  # удельный расход ДТ, кг/час (данные поставщика, ТКП)
 0:[66.3]*H, 1:[38.7]*H, 2:[50.1]*H, 3:[74.8]*H, 4:[93.5]*H}
# буровой инструмент и расходники (удельные), по данным листов станков:
BIT_RATE=[1.07,0.0,0.0,0.0,0.0]           # долота, шт./тыс.пог.м (только DR75 — данные ТКП; прочие уточнить)
BIT_PR  =[1.6,0.0,0.0,0.0,0.0]            # цена долота, тыс.вал/шт (DR75: 1600 USD = 1,6 тыс.USD)
ROD_RATE=[1.0,0.0,0.0,0.0,0.0]            # штанги, шт./тыс.пог.м
ROD_PR  =[4.0,0.0,0.0,0.0,0.0]            # цена штанги, тыс.вал/шт (DR75: 4000 USD = 4,0 тыс.USD)
CONS_PM =[0.0,0.0,0.0,0.0,0.0]            # прочие расходники, руб/пог.м
# --- Каталоги ТО и Ремонтов в формате модели экскаваторов ---------------------
# Детальный каталог узлов с ресурсом (м/ч) в ТКП дан у HIFO 75D (аналог DM75) — он
# принят КЛАССОВЫМ ШАБЛОНОМ. «Коэффициент калибровки» приводит ИТОГО ТОиР каждой
# машины к уровню её ТКП (данные drill_catalog.py, извлечены из 000_БСЛ-2027).
import drill_catalog as _dc
NARA=_dc.NARA; RATE=_dc.RATE
TO_INT=_dc.TO_INT; TO_PARTS=_dc.TO_PARTS
CUR_COST=_dc.CUR_COST; CUR_LABOR=_dc.CUR_LABOR; CAP_COST=_dc.CAP_COST; CAP_LABOR=_dc.CAP_LABOR
INT2NO={iv:k+1 for k,(_,iv,_,_) in enumerate(TO_INT)}
COEF=[0.4664,0.8225,0.8225,1.0586,0.8166]   # калибровка ИТОГО ТОиР под ТКП поставщика
TOIR_CUR=["EUR"]*NM     # каталог-шаблон в EUR; коэффициент калибровки поглощает валюту
TPL_NAMES=[f"Поставщик {i+1}" for i in range(NM)]
NP=len(TO_PARTS); NCU=len(CUR_COST); NCC=len(CAP_COST); NI=len(TO_INT)

tpl={}
for i in range(NM):
    ws=wb.create_sheet(TPL_NAMES[i]); ws.sheet_view.showGridLines=False
    title(ws,f"ДАННЫЕ ПОСТАВЩИКА — ВАРИАНТ {i+1} (ТО и Ремонты — модель по узлам)",
          "Жёлтое — ввод. Зелёное — авто (снизу вверх). ИТОГО ТОиР = (ЗЧ ТО + ЗЧ тек.+кап.ремонтов + сервис) × коэф.калибровки; тянется в «Данные по годам».",last="P")
    ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=42; ws.column_dimensions["C"].width=12; ws.column_dimensions["D"].width=11
    for p in range(1,H+1): ws.column_dimensions[PCOLS[p]].width=10
    ws.column_dimensions["P"].width=7
    ws.freeze_panes="E8"
    ws["B4"]="Буровой станок:"; ws["B4"].font=FB
    ws.merge_cells("E4:N4"); ws["E4"]=f"='{S_IN}'!{MACH[i]}{irow['name']}"; ws["E4"].font=Font(bold=True,color=DARK); ws["E4"].alignment=L
    ws["B5"]="Валюта затрат ТОиР:"; ws["B5"].font=FN
    cu=ws["D5"]; cu.value=TOIR_CUR[i]; cu.fill=Fi; cu.border=BD; cu.alignment=C; cu.font=FB
    dvt=DataValidation(type="list",formula1='"CNY,USD,EUR,RUB"'); ws.add_data_validation(dvt); dvt.add("D5")
    ws["B6"]="Показатель"; ws["C6"]="Ед."
    for cc in ("B","C","D"):
        ws[f"{cc}6"].font=FHd; ws[f"{cc}6"].fill=Fh; ws[f"{cc}6"].border=BD; ws[f"{cc}6"].alignment=C
    for p in range(1,H+1):
        c=ws[f"{PCOLS[p]}6"]; c.value=f"год {p}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C

    def yline(r,label,unit,fmt,green=False,formula=None,values=None,inpc=False):
        b=ws[f"B{r}"]; b.value=label; b.border=BD; b.alignment=L; b.font=(FRz if green else FN)
        ws[f"C{r}"]=unit; ws[f"C{r}"].font=FU; ws[f"C{r}"].border=BD; ws[f"C{r}"].alignment=C
        for p in range(1,H+1):
            col=PCOLS[p]; c=ws[f"{col}{r}"]; c.border=BD; c.alignment=C; c.number_format=fmt; c.font=(FRz if green else FN)
            if formula: c.value="="+formula(col,p); c.fill=Fr
            else:
                if values is not None: c.value=values[p-1]
                if inpc: c.fill=Fi

    # --- фиксированная разметка (каталог одинакового размера у всех машин) ---
    R_RATE=23; R_TOSEC=24; R_NARA=25; R_INTHDR=26; R_INT0=27; R_INTN=R_INT0+NI-1
    R_TODT0=R_INTN+2; R_TODT_TOT=R_TODT0+NI
    R_TOCH0=R_TODT_TOT+1; R_TOCH_TOT=R_TOCH0+NI
    R_CATSEC=R_TOCH_TOT+2; R_CATHDR=R_CATSEC+1; R_CAT0=R_CATHDR+1; R_CATN=R_CAT0+NP-1; R_CATCORR=R_CATN+1; R_CATTOT=R_CATCORR+1
    R_CUCSEC=R_CATTOT+2; R_CUCHDR=R_CUCSEC+1; R_CUC0=R_CUCHDR+1; R_CUCN=R_CUC0+NCU-1; R_CUCTOT=R_CUCN+1
    R_CULSEC=R_CUCTOT+2; R_CULHDR=R_CULSEC+1; R_CUL0=R_CULHDR+1; R_CULN=R_CUL0+NCU-1; R_CULTOT=R_CULN+1
    R_CACSEC=R_CULTOT+2; R_CACHDR=R_CACSEC+1; R_CAC0=R_CACHDR+1; R_CACN=R_CAC0+NCC-1; R_CACTOT=R_CACN+1
    R_CALSEC=R_CACTOT+2; R_CALHDR=R_CALSEC+1; R_CAL0=R_CALHDR+1; R_CALN=R_CAL0+NCC-1; R_CALTOT=R_CALN+1
    R_TOOLSEC=R_CALTOT+2; R_TOOLHDR=R_TOOLSEC+1; R_BIT=R_TOOLHDR+1; R_ROD=R_BIT+1; R_CONS=R_ROD+1
    R_COEF=19

    # ---- СВОДКА ПО ГОДАМ ----
    section(ws,7,"СВОДКА ПО ГОДАМ (зелёное — авто снизу вверх; тянется в расчёт)","B","N")
    yline(8,"Простои на ТО (авто из интервалов)","ч/год",M1,green=True,formula=lambda col,p:f"{col}{R_TODT_TOT}")
    yline(9,"Простои в ремонтах (справочно)","ч/год",M,values=DT_REM[i],inpc=True)
    yline(10,"Ежесменный осмотр (20 мин/смену), справочно","ч/год",M,green=True,
          formula=lambda col,p:f"({pc('kfv')}-{col}8-{col}9)/{pc('shift')}*20/60")
    yline(11,"КТГ (техготовность, ВВОД) → «Данные по годам»","коэф.",P1,values=KTG_Y[i],inpc=True)
    yline(12,"Затраты на ЗЧ ТО (авто из каталога)","тыс.вал/г",M2,green=True,formula=lambda col,p:f"{col}{R_CATTOT}")
    yline(13,"Затраты на ЗЧ текущих ремонтов (по ресурсу)","тыс.вал/г",M2,green=True,formula=lambda col,p:f"{col}{R_CUCTOT}")
    yline(14,"Затраты на ЗЧ капремонтов (по ресурсу)","тыс.вал/г",M2,green=True,formula=lambda col,p:f"{col}{R_CACTOT}")
    yline(15,"Трудозатраты ТО (авто)","ч-час/г",M1,green=True,formula=lambda col,p:f"{col}{R_TOCH_TOT}")
    yline(16,"Трудозатраты текущих ремонтов (авто)","ч-час/г",M1,green=True,formula=lambda col,p:f"{col}{R_CULTOT}")
    yline(17,"Трудозатраты капремонтов (авто)","ч-час/г",M1,green=True,formula=lambda col,p:f"{col}{R_CALTOT}")
    yline(18,"Сервис (трудозатраты ТО+ТР+КР)","тыс.вал/г",M2,green=True,
          formula=lambda col,p:f"({col}15+{col}16+{col}17)*$E${R_RATE}/1000")
    ws[f"B{R_COEF}"]="Коэффициент калибровки ИТОГО под ТКП"; ws[f"B{R_COEF}"].font=FN; ws[f"B{R_COEF}"].border=BD; ws[f"B{R_COEF}"].alignment=L
    ws[f"C{R_COEF}"]="коэф."; ws[f"C{R_COEF}"].font=FU; ws[f"C{R_COEF}"].border=BD; ws[f"C{R_COEF}"].alignment=C
    ce=ws[f"E{R_COEF}"]; ce.value=COEF[i]; ce.fill=Fi; ce.border=BD; ce.alignment=C; ce.number_format=M2; ce.font=FB
    yline(20,"ИТОГО ТОиР → «Данные по годам»","тыс.вал/г",M2,green=True,
          formula=lambda col,p:f"({col}12+{col}13+{col}14+{col}18)*$E${R_COEF}")
    yline(21,"Удельный расход ДТ / энергии","кг(кВт)/час",M1,values=FUEL_Y[i],inpc=True)

    # ставка сервиса
    ws[f"B{R_RATE}"]="Ставка сервиса"; ws[f"B{R_RATE}"].font=FN; ws[f"B{R_RATE}"].border=BD; ws[f"B{R_RATE}"].alignment=L
    ws[f"C{R_RATE}"]="вал/ч·час"; ws[f"C{R_RATE}"].font=FU; ws[f"C{R_RATE}"].border=BD; ws[f"C{R_RATE}"].alignment=C
    sr=ws[f"E{R_RATE}"]; sr.value=RATE; sr.fill=Fi; sr.border=BD; sr.alignment=C; sr.number_format=M2; sr.font=FB

    # ---- РАСЧЁТ ТО (по интервалам обслуживания) ----
    section(ws,R_TOSEC,"РАСЧЁТ ТО (по интервалам обслуживания), ПО ГОДАМ","B","N")
    yline(R_NARA,"Наработка по годам","м/час",M,values=[NARA]*H,inpc=True)
    for k,h in enumerate(["Вид ТО","интервал, м/ч","простой/ТО, ч","персонал"]):
        c=ws.cell(row=R_INTHDR,column=2+k,value=h); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
    for k,(nm,iv,dt,crew) in enumerate(TO_INT):
        r=R_INT0+k
        ws.cell(row=r,column=2,value=nm).font=FN; ws.cell(row=r,column=2).border=BD; ws.cell(row=r,column=2).alignment=L
        for cc2,val,fmt in ((3,iv,M),(4,dt,M2),(5,crew,M)):
            c=ws.cell(row=r,column=cc2,value=val); c.fill=Fi; c.border=BD; c.alignment=C; c.number_format=fmt
    for k in range(NI):
        rr=R_TODT0+k; ref=R_INT0+k
        yline(rr,f"{TO_INT[k][0]}: простои, ч/год","ч",M2,green=True,
              formula=lambda col,p,ref=ref:f"{col}${R_NARA}/$C{ref}*$D{ref}")
    ws.cell(row=R_TODT_TOT,column=2,value="Итого простои ТО").font=FRz; ws.cell(row=R_TODT_TOT,column=2).border=BD; ws.cell(row=R_TODT_TOT,column=2).alignment=L
    for p in range(1,H+1):
        col=PCOLS[p]; c=ws[f"{col}{R_TODT_TOT}"]; c.value=f"=SUM({col}{R_TODT0}:{col}{R_TODT0+NI-1})"
        c.fill=Fr; c.border=BD; c.alignment=C; c.number_format=M2; c.font=FRz
    for k in range(NI):
        rr=R_TOCH0+k; dtr=R_TODT0+k; ref=R_INT0+k
        yline(rr,f"{TO_INT[k][0]}: чел-часы/год","ч-час",M2,green=True,
              formula=lambda col,p,dtr=dtr,ref=ref:f"{col}{dtr}*$E{ref}")
    ws.cell(row=R_TOCH_TOT,column=2,value="Итого чел-часы ТО").font=FRz; ws.cell(row=R_TOCH_TOT,column=2).border=BD; ws.cell(row=R_TOCH_TOT,column=2).alignment=L
    for p in range(1,H+1):
        col=PCOLS[p]; c=ws[f"{col}{R_TOCH_TOT}"]; c.value=f"=SUM({col}{R_TOCH0}:{col}{R_TOCH0+NI-1})"
        c.fill=Fr; c.border=BD; c.alignment=C; c.number_format=M2; c.font=FRz

    # ---- КАТАЛОГ ЗЧ ТО ----
    section(ws,R_CATSEC,"КАТАЛОГ ЗАПЧАСТЕЙ ТО (заполняет подрядчик), ПО ГОДАМ","B","N")
    for k,h in enumerate(["Наименование","цена за ед","кол-во/ТО"]):
        c=ws.cell(row=R_CATHDR,column=2+k,value=h); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
    ws.cell(row=R_CATHDR,column=16,value="№ ТО").font=FHd; ws.cell(row=R_CATHDR,column=16).fill=Fh; ws.cell(row=R_CATHDR,column=16).border=BD; ws.cell(row=R_CATHDR,column=16).alignment=C
    for k,(nm,price,qty,iv) in enumerate(TO_PARTS):
        r=R_CAT0+k
        ws.cell(row=r,column=2,value=nm).font=FN; ws.cell(row=r,column=2).border=BD; ws.cell(row=r,column=2).alignment=L
        ws.cell(row=r,column=3,value=price); ws.cell(row=r,column=4,value=qty); ws.cell(row=r,column=16,value=INT2NO[iv])
        for cc2 in (3,4,16): ws.cell(row=r,column=cc2).fill=Fi; ws.cell(row=r,column=cc2).border=BD; ws.cell(row=r,column=cc2).alignment=C; ws.cell(row=r,column=cc2).number_format=(M2 if cc2!=16 else M)
        for p in range(1,H+1):
            col=PCOLS[p]
            f=ws.cell(row=r,column=4+p,value=f"={col}${R_NARA}/INDEX($C${R_INT0}:$C${R_INTN},$P{r})*D{r}*C{r}/1000")
            f.fill=Fr; f.border=BD; f.alignment=C; f.number_format=M2; f.font=FRz
    ws.cell(row=R_CATCORR,column=2,value="Корректировка (прочее)").font=FN; ws.cell(row=R_CATCORR,column=2).border=BD; ws.cell(row=R_CATCORR,column=2).alignment=L
    for p in range(1,H+1):
        col=PCOLS[p]; c=ws.cell(row=R_CATCORR,column=4+p,value=0); c.fill=Fi; c.border=BD; c.alignment=C; c.number_format=M2
    ws.cell(row=R_CATTOT,column=2,value="ИТОГО ЗЧ ТО, тыс.вал/год").font=FRz; ws.cell(row=R_CATTOT,column=2).border=BD; ws.cell(row=R_CATTOT,column=2).alignment=L
    for p in range(1,H+1):
        col=PCOLS[p]; c=ws.cell(row=R_CATTOT,column=4+p,value=f"=SUM({col}{R_CAT0}:{col}{R_CATCORR})")
        c.fill=Fr; c.border=BD; c.alignment=C; c.number_format=M2; c.font=FRz

    # ---- РЕМОНТЫ по ресурсу (4 блока: текущие/капитальные × затраты/трудозатраты) ----
    def rem_block(sec,hdr,r0,items,title_txt,valhdr,to_thousands):
        section(ws,sec,title_txt,"B","N")
        for cc2,h in ((2,"Наименование"),(3,valhdr),(4,"ресурс, м/ч")):
            c=ws.cell(row=hdr,column=cc2,value=h); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
        for p in range(1,H+1):
            c=ws.cell(row=hdr,column=4+p,value=f"год {p}"); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
        for k,(nm,val,res) in enumerate(items):
            r=r0+k
            ws.cell(row=r,column=2,value=nm).font=FN; ws.cell(row=r,column=2).border=BD; ws.cell(row=r,column=2).alignment=L
            ws.cell(row=r,column=3,value=val); ws.cell(row=r,column=4,value=res)
            for cc2,fmt in ((3,M2),(4,M)): ws.cell(row=r,column=cc2).fill=Fi; ws.cell(row=r,column=cc2).border=BD; ws.cell(row=r,column=cc2).alignment=C; ws.cell(row=r,column=cc2).number_format=fmt
            div="/1000" if to_thousands else ""
            for p in range(1,H+1):
                col=PCOLS[p]
                f=ws.cell(row=r,column=4+p,value=f"=(INT($E${R_NARA}*{p}/$D{r})-INT($E${R_NARA}*({p}-1)/$D{r}))*$C{r}{div}")
                f.fill=Fr; f.border=BD; f.alignment=C; f.number_format=M2; f.font=FRz
        tot=r0+len(items)
        ws.cell(row=tot,column=2,value="ИТОГО").font=FRz; ws.cell(row=tot,column=2).border=BD
        for p in range(1,H+1):
            col=PCOLS[p]; c=ws.cell(row=tot,column=4+p,value=f"=SUM({col}{r0}:{col}{tot-1})")
            c.fill=Fr; c.border=BD; c.alignment=C; c.number_format=M2; c.font=FRz
    rem_block(R_CUCSEC,R_CUCHDR,R_CUC0,CUR_COST,"ТЕКУЩИЕ РЕМОНТЫ — ЗАТРАТЫ НА ЗЧ (по ресурсу)","стоим.замены,вал",True)
    rem_block(R_CULSEC,R_CULHDR,R_CUL0,CUR_LABOR,"ТЕКУЩИЕ РЕМОНТЫ — ТРУДОЗАТРАТЫ (по ресурсу)","нормо-час/событие",False)
    rem_block(R_CACSEC,R_CACHDR,R_CAC0,CAP_COST,"КАПИТАЛЬНЫЕ РЕМОНТЫ — ЗАТРАТЫ НА ЗЧ (по ресурсу до замены узла)","стоим.замены,вал",True)
    rem_block(R_CALSEC,R_CALHDR,R_CAL0,CAP_LABOR,"КАПИТАЛЬНЫЕ РЕМОНТЫ — ТРУДОЗАТРАТЫ (по ресурсу до замены узла)","нормо-час/событие",False)

    # --- буровой инструмент и расходники (удельные, ввод) ---
    section(ws,R_TOOLSEC,"БУРОВОЙ ИНСТРУМЕНТ И РАСХОДНИКИ (удельные нормы)","B","N")
    for k,h in enumerate(["Позиция","ед.","норма","цена, тыс.вал/шт"]):
        c=ws.cell(row=R_TOOLHDR,column=2+k,value=h); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
    def toolrow(r,label,unit,rate,price):
        ws.cell(row=r,column=2,value=label).font=FN; ws.cell(row=r,column=2).border=BD; ws.cell(row=r,column=2).alignment=L
        ws.cell(row=r,column=3,value=unit).font=FU; ws.cell(row=r,column=3).border=BD; ws.cell(row=r,column=3).alignment=C
        cr=ws.cell(row=r,column=4,value=rate); cr.fill=Fi; cr.border=BD; cr.alignment=C; cr.number_format=NUM
        cp=ws.cell(row=r,column=5,value=price); cp.fill=Fi; cp.border=BD; cp.alignment=C; cp.number_format=M2
    toolrow(R_BIT,"Долота","шт./тыс.пог.м",BIT_RATE[i],BIT_PR[i])
    toolrow(R_ROD,"Штанги буровые / адаптеры","шт./тыс.пог.м",ROD_RATE[i],ROD_PR[i])
    ws.cell(row=R_CONS,column=2,value="Прочие расходники").font=FN; ws.cell(row=R_CONS,column=2).border=BD; ws.cell(row=R_CONS,column=2).alignment=L
    ws.cell(row=R_CONS,column=3,value="руб/пог.м").font=FU; ws.cell(row=R_CONS,column=3).border=BD; ws.cell(row=R_CONS,column=3).alignment=C
    ccx=ws.cell(row=R_CONS,column=4,value=CONS_PM[i]); ccx.fill=Fi; ccx.border=BD; ccx.alignment=C; ccx.number_format=M2

    tpl[i]={"sheet":TPL_NAMES[i],"ktg":11,"itogo":20,"fuel":21,"cur":"$D$5",
            "bit_rate":f"$D${R_BIT}","bit_pr":f"$E${R_BIT}","rod_rate":f"$D${R_ROD}",
            "rod_pr":f"$E${R_ROD}","cons_pm":f"$D${R_CONS}"}
    NOTEROW=R_CONS+2
    ws.merge_cells(f"B{NOTEROW}:P{NOTEROW}")
    nt=ws[f"B{NOTEROW}"]; nt.value=("Модель ТО и Ремонтов — как у экскаваторов: ТО по интервалам обслуживания (простои, чел-часы, каталог ЗЧ по годам) и "
     "ремонты ПО ПЕРИОДИЧНОСТИ — раздельно текущие и капитальные, затраты и трудозатраты (у каждого узла — стоимость/трудоёмкость замены и "
     "ресурс, м/ч; затрата возникает в годы, когда наработка пересекает кратное ресурсу значение). Каталог узлов детально заполнен по ТКП "
     "HIFO 75D (аналог DM75) и принят ШАБЛОНОМ класса; «Коэффициент калибровки» приводит ИТОГО ТОиР к уровню конкретного поставщика. "
     "Замените каталог реальными позициями из ТКП и установите коэффициент = 1. КТГ вводится напрямую (строка 11) и управляет "
     "производительностью. Буровой инструмент и расходники — удельными нормами (для инструмента у ряда станков данные ТКП отсутствуют — уточните).")
    nt.font=Font(size=9,italic=True,color="808080"); nt.alignment=L
print("Листы-шаблоны поставщиков (модель ТО/Ремонтов по узлам):",NM)

# --- СВОДКА «Данные по годам» (тянет из листов-шаблонов) ---
wy=wb.create_sheet(S_PYR); wy.sheet_view.showGridLines=False
title(wy,"ПОГОДОВЫЕ ДАННЫЕ ЭКСПЛУАТАЦИИ (СВОДКА)","Автоматически из листов «Поставщик 1…5». Правьте данные там, здесь — только просмотр.",last="N")
wy.column_dimensions["A"].width=2; wy.column_dimensions["B"].width=34; wy.column_dimensions["C"].width=11; wy.column_dimensions["D"].width=9
for p in range(1,H+1): wy.column_dimensions[PCOLS[p]].width=10
wy.freeze_panes="E6"
wy.merge_cells("B4:N4")
ins=wy["B4"]; ins.value=("Значения ТЯНУТСЯ из листов «Поставщик 1…5» (КТГ и ИТОГО ТОиР — авто). "
 "Чтобы изменить данные по машине — откройте её лист и правьте жёлтые ячейки; сводка и расчёт обновятся сами.")
ins.font=Font(size=9,italic=True,color=ACC); ins.alignment=L; ins.fill=Fa
YHR=5
wy[f"B{YHR}"]="Показатель"; wy[f"C{YHR}"]="Ед."; wy[f"D{YHR}"]="Валюта"
for cc in ("B","C","D"):
    wy[f"{cc}{YHR}"].font=FHd; wy[f"{cc}{YHR}"].fill=Fh; wy[f"{cc}{YHR}"].border=BD; wy[f"{cc}{YHR}"].alignment=C
for p in range(1,H+1):
    c=wy[f"{PCOLS[p]}{YHR}"]; c.value=f"год {p}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
pyr_ktg={}; pyr_toir={}; pyr_fuel={}; pyr_cur={}
yr=YHR+1
def yrow_pull(store,i,label,unit,tkey,fmt,cur=False):
    global yr
    store[i]=yr
    wy[f"B{yr}"]=label; wy[f"B{yr}"].font=FN; wy[f"B{yr}"].border=BD; wy[f"B{yr}"].alignment=L
    wy[f"C{yr}"]=unit; wy[f"C{yr}"].font=FU; wy[f"C{yr}"].border=BD; wy[f"C{yr}"].alignment=C
    if cur:
        pyr_cur[i]=yr; cu=wy[f"D{yr}"]; cu.value=f"='{tpl[i]['sheet']}'!{tpl[i]['cur']}"
        cu.fill=Fr; cu.border=BD; cu.alignment=C; cu.font=FB
    for p in range(1,H+1):
        c=wy[f"{PCOLS[p]}{yr}"]; c.value=f"='{tpl[i]['sheet']}'!{PCOLS[p]}{tpl[i][tkey]}"
        c.fill=Fr; c.border=BD; c.number_format=fmt; c.alignment=C; c.font=FN
    yr+=1
for i in range(NM):
    wy.merge_cells(f"B{yr}:N{yr}")
    hc=wy[f"B{yr}"]; hc.value=f"=\"Вариант {i+1}:  \"&'{S_IN}'!{MACH[i]}{irow['name']}"; hc.font=FSec; hc.alignment=L
    for col in range(2,15): wy.cell(row=yr,column=col).fill=Fm
    yr+=1
    yrow_pull(pyr_ktg,i,"КТГ (техготовность)","коэф.","ktg",P1)
    yrow_pull(pyr_toir,i,"Затраты на ТОиР и сервис","тыс.вал/г","itogo",M2,cur=True)
    yrow_pull(pyr_fuel,i,"Удельный расход ДТ/энергии","кг/час","fuel",M1)
_PYR_ROW={"ktg":pyr_ktg,"toir":pyr_toir,"fuel":pyr_fuel}
def pyr(kind,i,p): return f"'{S_PYR}'!{PCOLS[p]}{_PYR_ROW[kind][i]}"
def pyr_cur_ref(i): return f"'{S_PYR}'!$D${pyr_cur[i]}"
def toir_rate(i):
    cur=pyr_cur_ref(i)
    return f'IF({cur}="CNY",{pc("cny")},IF({cur}="USD",{pc("usd")},IF({cur}="EUR",{pc("eur")},1)))'
print("Данные по годам (сводка) готова")

# ================================================================= #
# ПРОИЗВОДИТЕЛЬНОСТЬ (физика бурения — методика из файла «W час»)
# ================================================================= #
wpr=wb.create_sheet(S_PROD); wpr.sheet_view.showGridLines=False
title(wpr,"РАСЧЁТ ПРОИЗВОДИТЕЛЬНОСТИ (методика бурения)","Часовая производительность (п.м/час) — физика бурения (геометрия неизменна по годам); полная погодовая таблица — с КТГ по годам.",last="N")
wpr.column_dimensions["A"].width=2; wpr.column_dimensions["B"].width=46; wpr.column_dimensions["C"].width=12
for cc in MACH: wpr.column_dimensions[cc].width=15
for cc in ("K","L","M","N"): wpr.column_dimensions[cc].width=10
PHR=4
wpr[f"B{PHR}"]="Показатель"; wpr[f"B{PHR}"].font=FHd; wpr[f"B{PHR}"].fill=Fh; wpr[f"B{PHR}"].border=BD; wpr[f"B{PHR}"].alignment=C
wpr[f"C{PHR}"]="Ед."; wpr[f"C{PHR}"].font=FHd; wpr[f"C{PHR}"].fill=Fh; wpr[f"C{PHR}"].border=BD; wpr[f"C{PHR}"].alignment=C
for i,cc in enumerate(MACH):
    c=wpr[f"{cc}{PHR}"]; c.value=f"='{S_IN}'!{cc}{irow['name']}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
prd={}; pr=PHR+1
def padd2(key,label,unit,fn,fmt=M1,res=False,bold=False):
    global pr
    prd[key]=pr
    b=wpr[f"B{pr}"]; b.value=label; b.border=BD; b.alignment=L; b.font=(FRz if res else (FB if bold else FN))
    wpr[f"C{pr}"]=unit; wpr[f"C{pr}"].font=FU; wpr[f"C{pr}"].border=BD; wpr[f"C{pr}"].alignment=C
    for cc in MACH:
        c=wpr[f"{cc}{pr}"]; c.value="="+fn(cc); c.number_format=fmt; c.border=BD; c.alignment=Rr
        c.font=(FRz if res else FN)
        if res: c.fill=Fr
    pr+=1
def pref(key,col): return f"{col}{prd[key]}"
def cur_rate(col):
    cur=inp('price_cur',col)
    return f'IF({cur}="CNY",{pc("cny")},IF({cur}="USD",{pc("usd")},IF({cur}="EUR",{pc("eur")},1)))'
padd2("price_rub","Стоимость в рублях","тыс.руб",lambda c:f"{inp('price_val',c)}*{cur_rate(c)}+{inp('install',c)}",M,bold=True)
padd2("dcharge","Диаметр заряда","м",lambda c:f"{inp('bit_dia',c)}*{pc('razbur')}",NUM)
padd2("teor","Теор. скорость бурения (КЗД, КИО=1)","п.м/час",
      lambda c:f"(0.1*{inp('rot_pow',c)})/(POWER(2.73,0.017*{pc('strength')})*{pref('dcharge',c)}^2)",M2)
padd2("t_drill","Время бурения скважины","мин",lambda c:f"{pc('depth')}/{pref('teor',c)}*60",M2)
padd2("t_hoist","Время подъёма става","мин",lambda c:f"{pc('depth')}/({inp('hoist_v',c)}*{pc('hoist_eff')})",M2)
padd2("t_cycle","Время цикла скважины","мин",
      lambda c:f"{pref('t_drill',c)}+{pc('blow')}+{inp('spins',c)}*{inp('rod_add',c)}*2+{pref('t_hoist',c)}+{pc('tram')}",M2)
padd2("holes","Скважин за смену","ед.",lambda c:f"{pc('shift')}/({pref('t_cycle',c)}/60)",M2)
padd2("pm_h","Часовая производительность (по КИО)","п.м/час",
      lambda c:f"{pref('holes',c)}*{pc('depth')}/{pc('shift')}",M2,res=True)
padd2("kio1","КИО (1-й год)","коэф.",
      lambda c:f"{pyr('ktg',MACH.index(c),1)}*(1-137.92/(60*{pc('shift')}))",P1)
def prodref(key,col): return f"'{S_PROD}'!{col}{prd[key]}"

# ---- ПОГОДОВАЯ ПРОИЗВОДИТЕЛЬНОСТЬ ----
pr+=1
section(wpr,pr,"ПОГОДОВАЯ ПРОИЗВОДИТЕЛЬНОСТЬ (КТГ по годам из «Данные по годам»; часовая производительность — из физики выше)","B","N"); pr+=1
wpr[f"B{pr}"]="Показатель"; wpr[f"B{pr}"].font=FHd; wpr[f"B{pr}"].fill=Fh; wpr[f"B{pr}"].border=BD; wpr[f"B{pr}"].alignment=C
wpr[f"C{pr}"]="Ед."; wpr[f"C{pr}"].font=FHd; wpr[f"C{pr}"].fill=Fh; wpr[f"C{pr}"].border=BD; wpr[f"C{pr}"].alignment=C
for p in range(1,H+1):
    c=wpr[f"{PCOLS[p]}{pr}"]; c.value=f"год {p}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
pr+=1
DT_TOTAL=137.92   # простои смены (регламентные+сменные), мин/см, для КИО (лист «Простои»)
def pyline(row,label,unit,fmt,formula_fn):
    b=wpr[f"B{row}"]; b.value=label; b.border=BD; b.alignment=L; b.font=FRz
    wpr[f"C{row}"]=unit; wpr[f"C{row}"].font=FU; wpr[f"C{row}"].border=BD; wpr[f"C{row}"].alignment=C
    for p in range(1,H+1):
        col=PCOLS[p]; c=wpr[f"{col}{row}"]; c.value="="+formula_fn(col,p); c.number_format=fmt
        c.border=BD; c.alignment=C; c.font=FRz; c.fill=Fr
    return row+1
prdy={}
for i in range(NM):
    ec=MACH[i]
    wpr.merge_cells(f"B{pr}:N{pr}")
    hc=wpr[f"B{pr}"]; hc.value=f"=\"Вариант {i+1}:  \"&'{S_IN}'!{ec}{irow['name']}"; hc.font=FSec; hc.alignment=L
    for col in range(2,15): wpr.cell(row=pr,column=col).fill=Fm
    pr+=1
    r_ktg=pr; pr=pyline(pr,"КТГ (по годам)","коэф.",P1,lambda col,p:pyr('ktg',i,p))
    r_kio=pr; pr=pyline(pr,"КИО","коэф.",P1,lambda col,p,r=r_ktg:f"{col}{r}*(1-{DT_TOTAL}/(60*{pc('shift')}))")
    r_eff=pr; pr=pyline(pr,"Эффективное время","час/год",M,lambda col,p,r=r_kio:f"{pc('kfv')}*{col}{r}")
    r_dm=pr; pr=pyline(pr,"Объём бурения","тыс.пог.м",M1,lambda col,p,r=r_eff,e=ec:f"{prodref('pm_h',e)}*{col}{r}/1000")
    r_v=pr; pr=pyline(pr,"Объём взрывания","тыс.м³",M,lambda col,p,r=r_dm:f"{col}{r}*{pc('yield')}")
    prdy[i]={"ktg":r_ktg,"kio":r_kio,"eff":r_eff,"dm":r_dm,"vol":r_v}
    pr+=1
def prodyref(key,i,p): return f"'{S_PROD}'!{PCOLS[p]}{prdy[i][key]}"
print("Производительность:",len(prd),"+ погодовая таблица,",NM,"машин")

# ================================================================= #
# ДЕНЕЖНЫЙ ПОТОК (погодовой движок)
# ================================================================= #
wf=wb.create_sheet(S_CF); wf.sheet_view.showGridLines=False
title(wf,"ПОГОДОВОЙ ДЕНЕЖНЫЙ ПОТОК И АННУИТЕТ","Пост-налоговый жизненный цикл по годам: период 0 — инвестиции, 1–10 — эксплуатация.",last="N")
wf.column_dimensions["A"].width=2; wf.column_dimensions["B"].width=42; wf.column_dimensions["C"].width=11
for p,col in enumerate(PCOLS): wf.column_dimensions[col].width=11
wf.freeze_panes="D7"
wf["B4"]="Период"; wf["B4"].font=FHd; wf["B4"].alignment=L
wf["B5"]="Календарный год"; wf["B5"].font=FU; wf["B5"].alignment=L
wf["B6"]="Дисконт-фактор"; wf["B6"].font=FHd; wf["B6"].alignment=L
for p,col in enumerate(PCOLS):
    a=wf[f"{col}4"]; a.value=p; a.font=FB; a.alignment=C; a.fill=Fh; a.border=BD
    y=wf[f"{col}5"]; y.value=2027+p; y.font=FU; y.alignment=C; y.border=BD
    d=wf[f"{col}6"]; d.value=f"=1/(1+{pc('disc')})^{col}$4"; d.number_format=NUM; d.alignment=C; d.border=BD; d.font=FN
DFROW=6; PERROW=4

CF_ROWS=[
 ("ktg","КТГ","коэф.",P1,True),
 ("eff","Эффективное время","час/год",M,True),
 ("dm","Объём бурения","тыс.пог.м",M1,True),
 ("prod","Объём взрывания","тыс.м³",M,True),
 ("fuel","Топливо/энергия","тыс.руб",M,True),
 ("maint","ТОиР и сервис (по годам)","тыс.руб",M,True),
 ("bit","Долота","тыс.руб",M,True),
 ("rod","Штанги/адаптеры","тыс.руб",M,True),
 ("cons","Расходники","тыс.руб",M,True),
 ("pers","Персонал","тыс.руб",M,True),
 ("ttax","Транспортный налог","тыс.руб",M,True),
 ("cash","Итого денежные затраты","тыс.руб",M,True),
 ("dep","Амортизация (НУ)","тыс.руб",M,True),
 ("base","Налогооблагаемая база","тыс.руб",M,True),
 ("tax","Налог на прибыль (щит)","тыс.руб",M,True),
 ("ocf","Операц. денежный поток","тыс.руб",M,True),
 ("invr","Инвестиции / остаточная ст-ть","тыс.руб",M,False),
 ("cf","Чистый денежный поток","тыс.руб",M,False),
 ("dcf","Дисконтированный поток","тыс.руб",M,False),
 ("cdcf","Кумул. дисконт. поток","тыс.руб",M,False),
 ("ann_y","Аннуитет по году","руб/пог.м",M2,True),
]
BLOCK_H=len(CF_ROWS)+2
cf={}
first_block=8
for i in range(NM):
    bs=first_block+i*BLOCK_H
    ec=MACH[i]
    wf.merge_cells(f"B{bs}:N{bs}")
    hc=wf[f"B{bs}"]; hc.value=f"=\"Вариант {i+1}:  \"&'{S_IN}'!{ec}{irow['name']}"; hc.font=FSec; hc.alignment=L
    for col in range(2,15): wf.cell(row=bs,column=col).fill=Fm
    rowmap={}
    for j,(key,label,unit,fmt,op) in enumerate(CF_ROWS):
        rr=bs+1+j; rowmap[key]=rr
        b=wf[f"B{rr}"]; b.value=label; b.font=(FRz if key=="ann_y" else FN); b.border=BD; b.alignment=L
        wf[f"C{rr}"]=unit; wf[f"C{rr}"].font=FU; wf[f"C{rr}"].border=BD; wf[f"C{rr}"].alignment=C
    cf[i]=rowmap
    def I(k): return f"'{S_IN}'!{ec}{irow[k]}"
    PR=prodref('price_rub',ec)
    for p,col in enumerate(PCOLS):
        per=f"{col}${PERROW}"
        def put(key,formula,fmt=M,op_only=True):
            rr=rowmap[key]; cell=wf[f"{col}{rr}"]
            if op_only and p==0: cell.value=None
            else: cell.value="="+formula
            cell.number_format=fmt; cell.border=BD; cell.alignment=Rr; cell.font=(FRz if key=="ann_y" else FN)
            if key in ("cash","ocf","cf","ann_y") and cell.value: cell.font=(FRz if key=="ann_y" else FB)
        put("ktg", (f"{pyr('ktg',i,p)}" if p>0 else "0"), P1)
        put("eff", f"{pc('kfv')}*{col}{rowmap['ktg']}*(1-{DT_TOTAL}/(60*{pc('shift')}))", M)
        put("dm", f"{prodref('pm_h',ec)}*{col}{rowmap['eff']}/1000", M1)
        put("prod", f"{col}{rowmap['dm']}*{pc('yield')}", M)
        put("fuel", f"{col}{rowmap['eff']}*{pyr('fuel',i,p)}/1000*{pc('fuel_p')}*(1+{pc('esc_fuel')})^({per}-1)", M)
        put("maint", (f"{pyr('toir',i,p)}*{toir_rate(i)}" if p>0 else "0"), M)
        sh_i=tpl[i]['sheet']
        put("bit", f"{col}{rowmap['dm']}*'{sh_i}'!{tpl[i]['bit_rate']}*'{sh_i}'!{tpl[i]['bit_pr']}*{toir_rate(i)}*(1+{pc('esc_parts')})^({per}-1)", M)
        put("rod", f"{col}{rowmap['dm']}*'{tpl[i]['sheet']}'!{tpl[i]['rod_rate']}*'{tpl[i]['sheet']}'!{tpl[i]['rod_pr']}*{toir_rate(i)}*(1+{pc('esc_parts')})^({per}-1)", M)
        put("cons", f"{col}{rowmap['dm']}*'{tpl[i]['sheet']}'!{tpl[i]['cons_pm']}*(1+{pc('esc_parts')})^({per}-1)", M)
        put("pers", f"{I('personnel')}*(1+{pc('esc_lab')})^({per}-1)", M)
        put("ttax", f"{I('eng_hp')}*{tax_rate_hp}/1000", M)
        put("cash", f"{col}{rowmap['fuel']}+{col}{rowmap['maint']}+{col}{rowmap['bit']}+{col}{rowmap['rod']}+{col}{rowmap['cons']}+{col}{rowmap['pers']}+{col}{rowmap['ttax']}", M)
        put("dep", f"IF({per}<=INT({pc('dep_years')}),{PR}/{pc('dep_years')},IF({per}=INT({pc('dep_years')})+1,{PR}*({pc('dep_years')}-INT({pc('dep_years')}))/{pc('dep_years')},0))", M)
        put("base", f"{col}{rowmap['cash']}+{col}{rowmap['dep']}", M)
        put("tax", f"-{col}{rowmap['base']}*{pc('tax')}", M)
        put("ocf", f"{col}{rowmap['base']}+{col}{rowmap['tax']}-{col}{rowmap['dep']}", M)
        cell=wf[f"{col}{rowmap['invr']}"]
        cell.value=f"=IF({per}=0,{PR},IF({per}={pc('horizon')},-{I('resid')}*{PR},0))"
        cell.number_format=M; cell.border=BD; cell.alignment=Rr; cell.font=FN
        ocf_ref=f"{col}{rowmap['ocf']}" if p>0 else "0"
        cfc=wf[f"{col}{rowmap['cf']}"]; cfc.value=f"={ocf_ref}+{col}{rowmap['invr']}"
        cfc.number_format=M; cfc.border=BD; cfc.alignment=Rr; cfc.font=FB
        dcfc=wf[f"{col}{rowmap['dcf']}"]; dcfc.value=f"={col}{rowmap['cf']}*{col}${DFROW}"
        dcfc.number_format=M; dcfc.border=BD; dcfc.alignment=Rr; dcfc.font=FN
        cdc=wf[f"{col}{rowmap['cdcf']}"]; cdc.value=f"=SUM($D{rowmap['dcf']}:{col}{rowmap['dcf']})"
        cdc.number_format=M; cdc.border=BD; cdc.alignment=Rr; cdc.font=FN
        ayc=wf[f"{col}{rowmap['ann_y']}"]
        if p==0: ayc.value=None
        else:
            ayc.value=(f"=SUM($D{rowmap['dcf']}:{col}{rowmap['dcf']})/SUM($E${DFROW}:{col}${DFROW})"
                       f"/(SUM($E{rowmap['dm']}:{col}{rowmap['dm']})/{per})")
        ayc.number_format=M2; ayc.border=BD; ayc.alignment=Rr; ayc.font=FRz
print("Денежный поток: блоков",NM,"высота блока",BLOCK_H)

def cfrng(i,key,p0=1,p1=H):
    c0=PCOLS[p0]; c1=PCOLS[p1]; return f"'{S_CF}'!{c0}{cf[i][key]}:{c1}{cf[i][key]}"
DF_RNG=f"'{S_CF}'!$E${DFROW}:$N${DFROW}"

# ================================================================= #
# РАСЧЁТ АННУИТЕТА
# ================================================================= #
wa=wb.create_sheet(S_ANN); wa.sheet_view.showGridLines=False
title(wa,"РАСЧЁТ АННУИТЕТА (СВОД)","Приведённые NPV затрат из погодового движка → аннуитет руб/пог.м и руб/м³.")
wa.column_dimensions["A"].width=2; wa.column_dimensions["B"].width=46; wa.column_dimensions["C"].width=12
for cc in MACH: wa.column_dimensions[cc].width=15
AHR=4
wa[f"B{AHR}"]="Статья"; wa[f"B{AHR}"].font=FHd; wa[f"B{AHR}"].fill=Fh; wa[f"B{AHR}"].border=BD; wa[f"B{AHR}"].alignment=C
wa[f"C{AHR}"]="Ед."; wa[f"C{AHR}"].font=FHd; wa[f"C{AHR}"].fill=Fh; wa[f"C{AHR}"].border=BD; wa[f"C{AHR}"].alignment=C
for i,cc in enumerate(MACH):
    c=wa[f"{cc}{AHR}"]; c.value=f"='{S_IN}'!{cc}{irow['name']}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
ar={}; a=AHR+1
def asec(t):
    global a; section(wa,a,t,"B","H"); a+=1
def aadd(key,label,unit,fni,fmt=M,res=False,bold=False):
    global a; ar[key]=a
    b=wa[f"B{a}"]; b.value=label; b.border=BD; b.alignment=L; b.font=(FRz if res else (FB if bold else FN))
    wa[f"C{a}"]=unit; wa[f"C{a}"].font=FU; wa[f"C{a}"].border=BD; wa[f"C{a}"].alignment=C
    for i,cc in enumerate(MACH):
        c=wa[f"{cc}{a}"]; c.value="="+fni(i); c.number_format=fmt; c.border=BD; c.alignment=Rr
        c.font=(FRz if res else (FB if bold else FN))
        if res: c.fill=Fr
    a+=1
def aref(key,col): return f"{col}{ar[key]}"
def acol(i): return MACH[i]
def avgdm(i): return f"AVERAGE({cfrng(i,'dm')})"          # тыс.пог.м/год
def npv(i,key): return f"SUMPRODUCT({cfrng(i,key)},{DF_RNG})"
DF10=f"'{S_CF}'!$N${DFROW}"

asec("ПРИВЕДЁННЫЕ ЗАТРАТЫ ЗА ЖИЗНЕННЫЙ ЦИКЛ (NPV, тыс.руб)")
aadd("n_fuel","NPV топлива/энергии",   "тыс.руб",lambda i:npv(i,'fuel'))
aadd("n_maint","NPV ТОиР (по годам, вкл. капремонты)","тыс.руб",lambda i:npv(i,'maint'))
aadd("n_bit","NPV долот",              "тыс.руб",lambda i:npv(i,'bit'))
aadd("n_rod","NPV штанг/адаптеров",    "тыс.руб",lambda i:npv(i,'rod'))
aadd("n_cons","NPV расходников",       "тыс.руб",lambda i:npv(i,'cons'))
aadd("n_pers","NPV персонала",         "тыс.руб",lambda i:npv(i,'pers'))
aadd("n_ttax","NPV транспортного налога","тыс.руб",lambda i:npv(i,'ttax'))
aadd("n_tax","NPV налога (щит)",       "тыс.руб",lambda i:npv(i,'tax'))
aadd("n_inv","NPV инвестиций (за вычетом остат.)","тыс.руб",
     lambda i:f"{prodref('price_rub',acol(i))}*(1-{inp('resid',acol(i))}*{DF10})")
aadd("n_op","NPV операционных затрат","тыс.руб",
     lambda i:f"{aref('n_fuel',acol(i))}+{aref('n_maint',acol(i))}+{aref('n_bit',acol(i))}+{aref('n_rod',acol(i))}+{aref('n_cons',acol(i))}+{aref('n_pers',acol(i))}+{aref('n_ttax',acol(i))}+{aref('n_tax',acol(i))}",bold=True)
aadd("n_tot","NPV полной стоимости владения","тыс.руб",
     lambda i:f"{aref('n_inv',acol(i))}+{aref('n_op',acol(i))}",res=True)
aadd("avgdm","Средний годовой объём бурения","тыс.пог.м/год",lambda i:avgdm(i))
aadd("avgvol","Средний годовой объём взрывания","тыс.м³/год",lambda i:f"{aref('avgdm',acol(i))}*{pc('yield')}")

asec("УДЕЛЬНЫЙ АННУИТЕТ (руб/пог.м) — ГЛАВНЫЙ КРИТЕРИЙ")
# NPV [тыс.руб] / S / avgdm [тыс.пог.м] = руб/пог.м (множители 1000 сокращаются)
def annpm(nkey): return lambda i:f"{aref(nkey,acol(i))}/{pc('S')}/{aref('avgdm',acol(i))}"
aadd("u_inv","Инвестиционный","руб/пог.м",annpm('n_inv'),M2)
aadd("u_fuel","Топливо/энергия","руб/пог.м",annpm('n_fuel'),M2)
aadd("u_maint","ТОиР + капремонты","руб/пог.м",annpm('n_maint'),M2)
aadd("u_bit","Долота","руб/пог.м",annpm('n_bit'),M2)
aadd("u_rod","Штанги/адаптеры","руб/пог.м",annpm('n_rod'),M2)
aadd("u_cons","Расходники","руб/пог.м",annpm('n_cons'),M2)
aadd("u_pers","Персонал","руб/пог.м",annpm('n_pers'),M2)
aadd("u_ttax","Транспортный налог","руб/пог.м",annpm('n_ttax'),M2)
aadd("u_tax","Налог на прибыль (щит)","руб/пог.м",annpm('n_tax'),M2)
aadd("u_op","Операционный аннуитет","руб/пог.м",annpm('n_op'),M2,bold=True)
aadd("u_tot","ОБЩИЙ АННУИТЕТ","руб/пог.м",annpm('n_tot'),M2,res=True)

asec("ДОПОЛНИТЕЛЬНЫЕ ПОКАЗАТЕЛИ")
aadd("v_tot","Удельный аннуитет на м³ (взрывание)","руб/м³",
     lambda i:f"{aref('u_tot',acol(i))}/{pc('yield')}",M2,bold=True)
aadd("sebest","Себестоимость (без дисконта)","руб/пог.м",
     lambda i:f"AVERAGE({cfrng(i,'cash')})/{aref('avgdm',acol(i))}",M2)
aadd("cap_share","Доля инвестиций в аннуитете","%",
     lambda i:f"{aref('u_inv',acol(i))}/{aref('u_tot',acol(i))}",P1)
def annref(key,col): return f"'{S_ANN}'!{col}{ar[key]}"
print("Аннуитет:",len(ar))

# ================================================================= #
# СРАВНЕНИЕ
# ================================================================= #
wm=wb.create_sheet(S_CMP); wm.sheet_view.showGridLines=False
title(wm,"СРАВНИТЕЛЬНЫЙ АНАЛИЗ БУРОВЫХ СТАНКОВ","Ранжирование по удельному аннуитету (руб/пог.м). Зелёным — лучший вариант.")
wm.column_dimensions["A"].width=2; wm.column_dimensions["B"].width=42; wm.column_dimensions["C"].width=11
for cc in MACH: wm.column_dimensions[cc].width=15
MHR=4
wm[f"B{MHR}"]="Показатель"; wm[f"B{MHR}"].font=FHd; wm[f"B{MHR}"].fill=Fh; wm[f"B{MHR}"].border=BD; wm[f"B{MHR}"].alignment=C
wm[f"C{MHR}"]="Ед."; wm[f"C{MHR}"].font=FHd; wm[f"C{MHR}"].fill=Fh; wm[f"C{MHR}"].border=BD; wm[f"C{MHR}"].alignment=C
for i,cc in enumerate(MACH):
    c=wm[f"{cc}{MHR}"]; c.value=f"='{S_IN}'!{cc}{irow['name']}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
metrics=[
 ("Стоимость (в рублях)","price_rub",prodref,M,"тыс.руб"),
 ("Часовая производительность","pm_h",prodref,M2,"п.м/час"),
 ("Средний годовой объём бурения","avgdm",annref,M,"тыс.пог.м"),
 ("NPV стоимости владения","n_tot",annref,M,"тыс.руб"),
 ("Себестоимость (без дисконта)","sebest",annref,M2,"руб/пог.м"),
 ("Инвестиционный аннуитет","u_inv",annref,M2,"руб/пог.м"),
 ("Операционный аннуитет","u_op",annref,M2,"руб/пог.м"),
 ("ОБЩИЙ АННУИТЕТ","u_tot",annref,M2,"руб/пог.м"),
 ("Аннуитет на м³ (взрывание)","v_tot",annref,M2,"руб/м³"),
]
mrow={}; m=MHR+1
for label,key,reff,fmt,unit in metrics:
    mrow[key]=m
    b=wm[f"B{m}"]; b.value=label; b.border=BD; b.alignment=L; b.font=(FB if key=="u_tot" else FN)
    wm[f"C{m}"]=unit; wm[f"C{m}"].font=FU; wm[f"C{m}"].border=BD; wm[f"C{m}"].alignment=C
    for cc in MACH:
        c=wm[f"{cc}{m}"]; c.value="="+reff(key,cc); c.number_format=fmt; c.border=BD; c.alignment=Rr
        c.font=(FB if key=="u_tot" else FN)
        if key=="u_tot": c.fill=Fr
    m+=1
section(wm,m,"РАНЖИРОВАНИЕ (критерий: минимум аннуитета руб/пог.м)","B","H"); m+=1
tot_rng=f"D{mrow['u_tot']}:{LC}{mrow['u_tot']}"
wm[f"B{m}"]="Место в рейтинге"; wm[f"B{m}"].font=FB; wm[f"B{m}"].border=BD
for cc in MACH:
    c=wm[f"{cc}{m}"]; c.value=f"=RANK({cc}{mrow['u_tot']},{tot_rng},1)"; c.font=FB; c.border=BD; c.alignment=C
m+=1
wm[f"B{m}"]="Отклонение от лучшего"; wm[f"B{m}"].font=FN; wm[f"B{m}"].border=BD
for cc in MACH:
    c=wm[f"{cc}{m}"]; c.value=f"={cc}{mrow['u_tot']}/MIN({tot_rng})-1"; c.number_format=P1; c.border=BD; c.alignment=C; c.font=FN
m+=2
names_rng=f"D{MHR}:{LC}{MHR}"
wm[f"B{m}"]="РЕКОМЕНДУЕМЫЙ ВАРИАНТ:"; wm[f"B{m}"].font=Font(bold=True,size=12,color=DARK)
wm.merge_cells(f"C{m}:{LC}{m}")
rec=wm[f"C{m}"]; rec.value=f'=INDEX({names_rng},MATCH(MIN({tot_rng}),{tot_rng},0))&"  ("&TEXT(MIN({tot_rng}),"0.00")&" руб/пог.м)"'
rec.font=Font(bold=True,size=12,color="006100"); rec.alignment=C
for col in ["C","D","E","F","G","H"]: wm[f"{col}{m}"].fill=Fb; wm[f"{col}{m}"].border=BD
m+=2
section(wm,m,"СТРУКТУРА УДЕЛЬНОГО АННУИТЕТА (руб/пог.м)","B","H"); m+=1
wm[f"B{m}"]="Статья"; wm[f"B{m}"].font=FHd; wm[f"B{m}"].fill=Fh; wm[f"B{m}"].border=BD
for cc in MACH:
    c=wm[f"{cc}{m}"]; c.value=f"='{S_IN}'!{cc}{irow['name']}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
stop=m+1
struct=[("Инвестиционный","u_inv"),("Топливо/энергия","u_fuel"),("ТОиР + капремонты","u_maint"),
        ("Долота","u_bit"),("Штанги/адаптеры","u_rod"),("Расходники","u_cons"),
        ("Персонал","u_pers"),("Транспортный налог","u_ttax"),("Налог (щит)","u_tax")]
m=stop
for label,key in struct:
    b=wm[f"B{m}"]; b.value=label; b.font=FN; b.border=BD; b.alignment=L
    for cc in MACH:
        c=wm[f"{cc}{m}"]; c.value="="+annref(key,cc); c.number_format=M2; c.border=BD; c.alignment=Rr; c.font=FN
    m+=1
catref=Reference(wm,min_col=4,min_row=stop-1,max_col=LCN,max_row=stop-1)
chart1=BarChart(); chart1.type="col"; chart1.grouping="stacked"; chart1.overlap=100
chart1.title="Структура аннуитета, руб/пог.м"; chart1.height=9; chart1.width=20; chart1.y_axis.title="руб/пог.м"
for k in range(len(struct)):
    ri=stop+k; vals=Reference(wm,min_col=4,min_row=ri,max_col=LCN,max_row=ri)
    ser=Series(vals,title_from_data=False); ser.tx=SeriesLabel(strRef=StrRef(f"'{S_CMP}'!$B${ri}")); chart1.series.append(ser)
chart1.set_categories(catref); wm.add_chart(chart1,"K4")
chart2=BarChart(); chart2.type="col"; chart2.title="Общий аннуитет, руб/пог.м"; chart2.height=8; chart2.width=11; chart2.y_axis.title="руб/пог.м"
d2=Reference(wm,min_col=4,min_row=mrow['u_tot'],max_col=LCN,max_row=mrow['u_tot'])
c2=Reference(wm,min_col=4,min_row=MHR,max_col=LCN,max_row=MHR)
chart2.add_data(d2,from_rows=True); chart2.set_categories(c2); chart2.legend=None
chart2.dataLabels=DataLabelList(); chart2.dataLabels.showVal=True; chart2.dataLabels.numFmt="0.0"
wm.add_chart(chart2,"K23")
chart3=BarChart(); chart3.type="col"; chart3.title="Стоимость станка, тыс.руб"; chart3.height=8; chart3.width=11
d3=Reference(wm,min_col=4,min_row=mrow['price_rub'],max_col=LCN,max_row=mrow['price_rub'])
chart3.add_data(d3,from_rows=True); chart3.set_categories(c2); chart3.legend=None
wm.add_chart(chart3,"T23")
print("Сравнение:",len(mrow))

# ================================================================= #
# ДАШБОРД
# ================================================================= #
wd=wb.create_sheet(S_DASH); wd.sheet_view.showGridLines=False
title(wd,"ИТОГОВЫЙ ДАШБОРД — СРАВНЕНИЕ БУРОВЫХ СТАНКОВ","Ключевые показатели и графики. Данные обновляются автоматически.",last="N")
for col,w in (("A",2),("B",22),("C",22),("D",22),("E",22),("F",6),("G",13),("H",13),("I",15),("J",13)):
    wd.column_dimensions[col].width=w
UT=f"'{S_ANN}'!D{ar['u_tot']}:{LC}{ar['u_tot']}"
NMv=f"'{S_ANN}'!D{ar['n_tot']}:{LC}{ar['n_tot']}"
NR=f"'{S_IN}'!D{irow['name']}:{LC}{irow['name']}"
def tile(anchor_col,row,caption,value_formula,fmt):
    c0=anchor_col
    hc=wd[f"{c0}{row}"]; hc.value=caption; hc.font=Font(size=9,bold=True,color="FFFFFF"); hc.fill=Ftile; hc.alignment=C
    vc=wd[f"{c0}{row+1}"]; vc.value="="+value_formula; vc.font=Font(size=14,bold=True,color=DARK); vc.alignment=C
    vc.number_format=fmt; vc.fill=Fb
    for rr in (row,row+1): wd[f"{c0}{rr}"].border=BD
tr=4
section(wd,tr,"КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ","B","J"); tr+=1
tile("B",tr,"Лучший вариант",f'INDEX({NR},MATCH(MIN({UT}),{UT},0))',"General")
tile("C",tr,"Мин. аннуитет, руб/пог.м",f"MIN({UT})",M2)
tile("D",tr,"Мин. NPV владения, тыс.руб",f"MIN({NMv})",M)
tile("E",tr,"Разброс аннуитета, руб/пог.м",f"MAX({UT})-MIN({UT})",M2)
tr+=3
section(wd,tr,"РЕЙТИНГ ПО УДЕЛЬНОМУ АННУИТЕТУ (руб/пог.м)","B","J"); tr+=1
for cc,txt in (("B","Место"),("C","Буровой станок"),("G","руб/пог.м"),("I","руб/м³")):
    wd[f"{cc}{tr}"]=txt; wd[f"{cc}{tr}"].font=FHd; wd[f"{cc}{tr}"].fill=Fh; wd[f"{cc}{tr}"].border=BD; wd[f"{cc}{tr}"].alignment=C
wd.merge_cells(f"C{tr}:F{tr}"); wd.merge_cells(f"G{tr}:H{tr}")
VT=f"'{S_ANN}'!$D${ar['v_tot']}:${LC}${ar['v_tot']}"
for k in range(NM):
    rr=tr+1+k
    wd[f"B{rr}"]=k+1; wd[f"B{rr}"].font=FB; wd[f"B{rr}"].border=BD; wd[f"B{rr}"].alignment=C
    wd.merge_cells(f"C{rr}:F{rr}")
    nm=wd[f"C{rr}"]; nm.value=f'=INDEX({NR},MATCH(SMALL({UT},{k+1}),{UT},0))'; nm.border=BD; nm.alignment=L
    nm.font=(Font(bold=True,color="006100") if k==0 else FN)
    wd.merge_cells(f"G{rr}:H{rr}")
    va=wd[f"G{rr}"]; va.value=f"=SMALL({UT},{k+1})"; va.number_format=M2; va.border=BD; va.alignment=C
    va.font=(FB if k==0 else FN)
    if k==0: nm.fill=Fb; va.fill=Fb
    vt=wd[f"I{rr}"]; vt.value=f'=INDEX({VT},MATCH(SMALL({UT},{k+1}),{UT},0))'; vt.number_format=M2; vt.border=BD; vt.alignment=C; vt.font=FN
gc=BarChart(); gc.type="col"; gc.title="Общий аннуитет, руб/пог.м"; gc.height=8; gc.width=13; gc.y_axis.title="руб/пог.м"
gd=Reference(wm,min_col=4,min_row=mrow['u_tot'],max_col=LCN,max_row=mrow['u_tot'])
gcat=Reference(wm,min_col=4,min_row=MHR,max_col=LCN,max_row=MHR)
gc.add_data(gd,from_rows=True); gc.set_categories(gcat); gc.legend=None
gc.dataLabels=DataLabelList(); gc.dataLabels.showVal=True; gc.dataLabels.numFmt="0.0"
wd.add_chart(gc,"L4")
lc=LineChart(); lc.title="Траектория аннуитета по годам, руб/пог.м"; lc.height=8; lc.width=16
lc.x_axis.title="Год эксплуатации"; lc.y_axis.title="руб/пог.м"
per_cat=Reference(wf,min_col=5,min_row=PERROW,max_col=14,max_row=PERROW)
for i in range(NM):
    rr=cf[i]['ann_y']; vals=Reference(wf,min_col=5,min_row=rr,max_col=14,max_row=rr)
    ser=Series(vals,title_from_data=False); ser.tx=SeriesLabel(strRef=StrRef(f"'{S_IN}'!{MACH[i]}${irow['name']}")); lc.series.append(ser)
lc.set_categories(per_cat)
wd.add_chart(lc,"L20")
print("Дашборд готов")

# ================================================================= #
# ЧУВСТВИТЕЛЬНОСТЬ
# ================================================================= #
ws=wb.create_sheet(S_SENS); ws.sheet_view.showGridLines=False
title(ws,"АНАЛИЗ ЧУВСТВИТЕЛЬНОСТИ","Торнадо драйверов, влияние КТГ и ставки дисконтирования на аннуитет.")
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=36
for cc in "CDEFGH": ws.column_dimensions[cc].width=15
ws["B4"]="Анализируемый станок (1–5):"; ws["B4"].font=FB
sel=ws["C4"]; sel.value=1; sel.fill=Fi; sel.border=BD; sel.alignment=C; sel.font=FB
SEL="$C$4"
dv2=DataValidation(type="whole",operator="between",formula1="1",formula2=str(NM)); ws.add_data_validation(dv2); dv2.add(sel)
ws["D4"]=f"=INDEX('{S_IN}'!D{irow['name']}:{LC}{irow['name']},{SEL})"; ws["D4"].font=Font(bold=True,color=ACC); ws["D4"].alignment=L
ws.merge_cells("D4:H4")
def selann(key): return f"INDEX('{S_ANN}'!D{ar[key]}:{LC}{ar[key]},{SEL})"
section(ws,6,"БАЗОВЫЕ ВЕЛИЧИНЫ ВЫБРАННОГО ВАРИАНТА","B","H")
base=[
 ("b_tot","Общий аннуитет","руб/пог.м",selann("u_tot"),M2),
 ("b_inv","— инвестиционный","руб/пог.м",selann("u_inv"),M2),
 ("b_fuel","— топливо/энергия","руб/пог.м",selann("u_fuel"),M2),
 ("b_maint","— ТОиР + капремонты","руб/пог.м",selann("u_maint"),M2),
 ("b_pers","— персонал","руб/пог.м",selann("u_pers"),M2),
 ("b_dm","Средний годовой объём бурения","тыс.пог.м",selann("avgdm"),M),
 ("b_price","Стоимость (руб)","тыс.руб",f"INDEX('{S_PROD}'!D{prd['price_rub']}:{LC}{prd['price_rub']},{SEL})",M),
 ("b_fix","Постоянные (инв+ТОиР+перс+налоги)","руб/пог.м",
   f"{selann('u_inv')}+{selann('u_maint')}+{selann('u_pers')}+{selann('u_ttax')}+{selann('u_tax')}",M2),
 ("b_var","Переменные (топл+инструм+расходн)","руб/пог.м",
   f"{selann('u_fuel')}+{selann('u_bit')}+{selann('u_rod')}+{selann('u_cons')}",M2),
]
brow={}; b=7
for key,label,unit,fml,fmt in base:
    brow[key]=b
    ws[f"B{b}"]=label; ws[f"B{b}"].font=FN; ws[f"B{b}"].border=BD; ws[f"B{b}"].alignment=L
    ws[f"C{b}"]=unit; ws[f"C{b}"].font=FU; ws[f"C{b}"].border=BD; ws[f"C{b}"].alignment=C
    c=ws[f"D{b}"]; c.value="="+fml; c.number_format=fmt; c.border=BD; c.alignment=Rr; c.font=FN
    b+=1
def br(key): return f"$D${brow[key]}"
TAX=pc('tax'); DSF=pc('dsf')
tr=b+1; section(ws,tr,"ТОРНАДО: ВЛИЯНИЕ ДРАЙВЕРОВ НА ОБЩИЙ АННУИТЕТ (руб/пог.м)","B","H"); tr+=1
for i,h in enumerate(["Драйвер","Диапазон","При снижении","При росте","Размах"]):
    c=ws.cell(row=tr,column=2+i,value=h); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
tr+=1
drivers=[("Расход / цена ДТ",0.20,f"{br('b_fuel')}*(1-{TAX})"),
         ("Затраты на ТОиР/ремонты",0.25,f"{br('b_maint')}*(1-{TAX})"),
         ("Цена приобретения",0.15,f"{br('b_inv')}*(1-{TAX}*{DSF})"),
         ("Расходы на персонал",0.15,f"{br('b_pers')}*(1-{TAX})")]
ttop=tr
for label,pct,net in drivers:
    ws.cell(row=tr,column=2,value=label).font=FN; ws.cell(row=tr,column=2).border=BD; ws.cell(row=tr,column=2).alignment=L
    ws.cell(row=tr,column=3,value=f"±{int(pct*100)}%").alignment=C; ws.cell(row=tr,column=3).border=BD; ws.cell(row=tr,column=3).font=FN
    ws.cell(row=tr,column=4,value=f"={br('b_tot')}-{pct}*({net})")
    ws.cell(row=tr,column=5,value=f"={br('b_tot')}+{pct}*({net})")
    ws.cell(row=tr,column=6,value=f"=E{tr}-D{tr}")
    for col in (4,5,6):
        cc=ws.cell(row=tr,column=col); cc.number_format=M2; cc.border=BD; cc.alignment=Rr; cc.font=FN
    tr+=1
tbot=tr-1
ws.cell(row=ttop-1,column=9,value="низкое").font=FU; ws.cell(row=ttop-1,column=10,value="высокое").font=FU
for i in range(ttop,tbot+1):
    ws.cell(row=i,column=9,value=f"=D{i}-{br('b_tot')}").number_format=M2
    ws.cell(row=i,column=10,value=f"=E{i}-{br('b_tot')}").number_format=M2
tch=BarChart(); tch.type="bar"; tch.grouping="stacked"; tch.overlap=100
tch.title="Торнадо: отклонение аннуитета, руб/пог.м"; tch.height=6.5; tch.width=15
dl=Reference(ws,min_col=9,min_row=ttop,max_col=9,max_row=tbot); dh=Reference(ws,min_col=10,min_row=ttop,max_col=10,max_row=tbot)
cats=Reference(ws,min_col=2,min_row=ttop,max_col=2,max_row=tbot)
tch.add_data(dl); tch.add_data(dh); tch.set_categories(cats)
tch.series[0].graphicalProperties.solidFill="70AD47"; tch.series[1].graphicalProperties.solidFill="C55A11"; tch.legend=None
ws.add_chart(tch,"J6")
ur=tbot+3; section(ws,ur,"ВЛИЯНИЕ ПРОИЗВОДИТЕЛЬНОСТИ (КТГ / УТИЛИЗАЦИИ)","B","H"); ur+=1
ws.cell(row=ur,column=2,value="Изменение производительности").font=FHd
ws.cell(row=ur,column=2).fill=Fh; ws.cell(row=ur,column=2).border=BD; ws.cell(row=ur,column=2).alignment=C
levels=[-0.15,-0.10,0.0,0.10,0.15]
for j,d in enumerate(levels):
    c=ws.cell(row=ur,column=3+j,value=f"{'+' if d>0 else ''}{int(d*100)}%"); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
ur+=1
ws.cell(row=ur,column=2,value="Общий аннуитет, руб/пог.м").font=FB; ws.cell(row=ur,column=2).border=BD; ws.cell(row=ur,column=2).alignment=L
for j,d in enumerate(levels):
    c=ws.cell(row=ur,column=3+j,value=f"={br('b_fix')}/(1+{d})+{br('b_var')}"); c.number_format=M2; c.border=BD; c.alignment=Rr; c.font=FB
    if abs(d)<1e-9: c.fill=Fr
ur+=2
section(ws,ur,"ВЛИЯНИЕ СТАВКИ ДИСКОНТИРОВАНИЯ","B","H"); ur+=1
ws.cell(row=ur,column=2,value="Ставка дисконтирования").font=FHd
ws.cell(row=ur,column=2).fill=Fh; ws.cell(row=ur,column=2).border=BD; ws.cell(row=ur,column=2).alignment=C
rates=[0.08,0.10,0.12,0.14,0.16]
for j,rt in enumerate(rates):
    c=ws.cell(row=ur,column=3+j,value=rt); c.number_format=P1; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
ur+=1
ws.cell(row=ur,column=2,value="Общий аннуитет, руб/пог.м (оценка)").font=FB; ws.cell(row=ur,column=2).border=BD; ws.cell(row=ur,column=2).alignment=L
for j,rt in enumerate(rates):
    col=get_column_letter(3+j)
    pvaf=f"(1-(1+{col}{ur-1})^-{pc('horizon')})/{col}{ur-1}"
    resid_sel=f"INDEX('{S_IN}'!D{irow['resid']}:{LC}{irow['resid']},{SEL})"
    inv_ann=f"{br('b_price')}*(1-{resid_sel}/(1+{col}{ur-1})^{pc('horizon')})/({pvaf})/{br('b_dm')}"
    c=ws.cell(row=ur,column=3+j,value=f"={inv_ann}+{br('b_tot')}-{br('b_inv')}")
    c.number_format=M2; c.border=BD; c.alignment=Rr; c.font=FB
    if abs(rt-0.12)<1e-9: c.fill=Fr
ur+=2
ws.cell(row=ur,column=2,value="Оценка по ставке: инвестиционный аннуитет пересчитан точно, операционный принят "
        "постоянным (аннуитет равномерного потока ≈ инвариантен к ставке).").font=FU
ws.merge_cells(f"B{ur}:H{ur}"); ws[f"B{ur}"].alignment=L
print("Чувствительность готова")

# ================================================================= #
# АНАЛИТИКА
# ================================================================= #
wan=wb.create_sheet(S_ANL); wan.sheet_view.showGridLines=False
title(wan,"РАСШИРЕННАЯ АНАЛИТИКА РЕЗУЛЬТАТОВ","Многокритериальная оценка, разложение затрат, диапазон риска, эффективность.",last="K")
wan.column_dimensions["A"].width=2; wan.column_dimensions["B"].width=42; wan.column_dimensions["C"].width=10
for cc in MACH: wan.column_dimensions[cc].width=15
NHR=4
wan[f"B{NHR}"]="Показатель"; wan[f"B{NHR}"].font=FHd; wan[f"B{NHR}"].fill=Fh; wan[f"B{NHR}"].border=BD; wan[f"B{NHR}"].alignment=C
wan[f"C{NHR}"]="Ед."; wan[f"C{NHR}"].font=FHd; wan[f"C{NHR}"].fill=Fh; wan[f"C{NHR}"].border=BD; wan[f"C{NHR}"].alignment=C
for i,cc in enumerate(MACH):
    c=wan[f"{cc}{NHR}"]; c.value=f"='{S_IN}'!{cc}{irow['name']}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
an={}; arow=NHR+1
def aline(key,label,unit,fni,fmt=M2,res=False):
    global arow; an[key]=arow
    b=wan[f"B{arow}"]; b.value=label; b.border=BD; b.alignment=L; b.font=(FRz if res else FN)
    wan[f"C{arow}"]=unit; wan[f"C{arow}"].font=FU; wan[f"C{arow}"].border=BD; wan[f"C{arow}"].alignment=C
    for i,cc in enumerate(MACH):
        c=wan[f"{cc}{arow}"]; c.value="="+fni(i); c.number_format=fmt; c.border=BD; c.alignment=Rr; c.font=(FRz if res else FN)
        if res: c.fill=Fr
    arow+=1
def asec2(t):
    global arow; section(wan,arow,t,"B","H"); arow+=1
def A(key,i): return annref(key,MACH[i])
def ktg_avg(i): return f"AVERAGE('{S_PYR}'!E{pyr_ktg[i]}:N{pyr_ktg[i]})"

asec2("МНОГОКРИТЕРИАЛЬНАЯ ОЦЕНКА (levelized cost + драйверы)")
aline("price","Цена (CAPEX)","тыс.руб",lambda i:prodref('price_rub',MACH[i]),M)
aline("dm","Средний годовой объём бурения","тыс.пог.м",lambda i:A('avgdm',i),M)
aline("ktg","КТГ средний","коэф.",lambda i:ktg_avg(i),P1)
aline("pm","Аннуитет","руб/пог.м",lambda i:A('u_tot',i),M2,res=True)
aline("v","Аннуитет","руб/м³",lambda i:A('v_tot',i),M2)
aline("npv","NPV стоимости владения","тыс.руб",lambda i:A('n_tot',i),M)
aline("s_cap","Доля CAPEX","%",lambda i:f"{A('u_inv',i)}/{A('u_tot',i)}",P1)
aline("s_fuel","Доля топлива/энергии","%",lambda i:f"{A('u_fuel',i)}/{A('u_tot',i)}",P1)
aline("s_maint","Доля ТОиР","%",lambda i:f"{A('u_maint',i)}/{A('u_tot',i)}",P1)
pmrng=f"D{an['pm']}:{LC}{an['pm']}"
aline("rank","Место в рейтинге","",lambda i:f"RANK({MACH[i]}{an['pm']},{pmrng},1)",M,res=True)

asec2("РАЗЛОЖЕНИЕ ПРЕВЫШЕНИЯ АННУИТЕТА НАД ЛУЧШИМ (руб/пог.м)")
def dvs(key): return lambda i:f"{A(key,i)}-MIN(D{ar[key]}:{LC}{ar[key]})"
aline("d_inv","Δ инвестиции",  "руб/пог.м",dvs('u_inv'),M2)
aline("d_fuel","Δ топливо/энергия","руб/пог.м",dvs('u_fuel'),M2)
aline("d_maint","Δ ТОиР",       "руб/пог.м",dvs('u_maint'),M2)
aline("d_pers","Δ персонал",   "руб/пог.м",dvs('u_pers'),M2)
aline("d_tax","Δ налог (щит)", "руб/пог.м",dvs('u_tax'),M2)
aline("d_tot","ИТОГО превышение над лучшим","руб/пог.м",
      lambda i:f"{MACH[i]}{an['pm']}-MIN({pmrng})",M2,res=True)

asec2("ДИАПАЗОН РИСКА АННУИТЕТА (руб/пог.м)")
def fixed(i): return f"({A('u_inv',i)}+{A('u_maint',i)}+{A('u_pers',i)}+{A('u_ttax',i)}+{A('u_tax',i)})"
def var(i): return f"({A('u_fuel',i)}+{A('u_bit',i)}+{A('u_rod',i)}+{A('u_cons',i)})"
aline("opt","Оптимистичный (КТГ +10%, ДТ −20%)","руб/пог.м",
      lambda i:f"{fixed(i)}/1.1+{var(i)}-0.2*{A('u_fuel',i)}",M2)
aline("base","Базовый","руб/пог.м",lambda i:A('u_tot',i),M2,res=True)
aline("pess","Пессимистичный (КТГ −10%, ДТ +20%)","руб/пог.м",
      lambda i:f"{fixed(i)}/0.9+{var(i)}+0.2*{A('u_fuel',i)}",M2)
aline("band","Ширина диапазона (риск)","руб/пог.м",
      lambda i:f"{MACH[i]}{an['pess']}-{MACH[i]}{an['opt']}",M2)

asec2("ЭФФЕКТИВНОСТЬ И ЦЕННОСТЬ ГОТОВНОСТИ")
aline("val_ktg","Эффект +1 п.п. КТГ на аннуитет","руб/пог.м",
      lambda i:f"-{fixed(i)}*0.01/{ktg_avg(i)}",M2)

arow+=1
names_rng=f"D{NHR}:{LC}{NHR}"
wan[f"B{arow}"]="ВЫВОД:"; wan[f"B{arow}"].font=Font(bold=True,size=12,color=DARK)
wan.merge_cells(f"C{arow}:{LC}{arow}")
concl=wan[f"C{arow}"]
concl.value=(f'="Лучший по удельному TCO: "&INDEX({names_rng},MATCH(MIN({pmrng}),{pmrng},0))'
             f'&" ("&TEXT(MIN({pmrng}),"0.00")&" руб/пог.м, "&TEXT(MIN(D{an["v"]}:{LC}{an["v"]}),"0.00")&" руб/м³). "'
             f'&"Отрыв от худшего: "&TEXT(MAX({pmrng})-MIN({pmrng}),"0.00")&" руб/пог.м."')
concl.font=Font(bold=True,size=11,color="006100"); concl.fill=Fb; concl.alignment=L
for cc in "CDEFGH": wan[f"{cc}{arow}"].border=BD
arow+=2
wan[f"B{arow}"]="Метод (лучшие практики):"; wan[f"B{arow}"].font=FB; arow+=1
for t in ["Levelized cost — приведённые затраты на единицу продукции (руб/пог.м, руб/м³) уравнивают станки с разной ценой и производительностью.",
          "Многокритериальная оценка — не только цена, но и структура затрат, КТГ, доля топлива.",
          "Разложение над лучшим — показывает, за счёт каких статей вариант дороже (управляемость).",
          "Диапазон риска — устойчивость выбора к колебаниям КТГ и цены топлива.",
          "Ценность готовности — сколько руб/пог.м даёт рост КТГ на 1 п.п. (обоснование сервисного контракта)."]:
    wan.merge_cells(f"B{arow}:K{arow}"); wan[f"B{arow}"]="•  "+t; wan[f"B{arow}"].font=FU; wan[f"B{arow}"].alignment=L; arow+=1
sc=ScatterChart(); sc.title="Аннуитет vs производительность"; sc.height=8; sc.width=12
sc.x_axis.title="Годовой объём бурения, тыс.пог.м"; sc.y_axis.title="Аннуитет, руб/пог.м"; sc.x_axis.delete=False; sc.y_axis.delete=False
xref=Reference(wan,min_col=4,min_row=an['dm'],max_col=LCN,max_row=an['dm'])
yref=Reference(wan,min_col=4,min_row=an['pm'],max_col=LCN,max_row=an['pm'])
ser=Series(yref,xref,title="варианты"); ser.marker.symbol="circle"; ser.marker.size=8; ser.graphicalProperties.line.noFill=True
sc.series.append(ser); sc.legend=None
wan.add_chart(sc,"K4")
rc=BarChart(); rc.type="col"; rc.title="Диапазон риска аннуитета, руб/пог.м"; rc.height=8; rc.width=13; rc.y_axis.title="руб/пог.м"
cats=Reference(wan,min_col=4,min_row=NHR,max_col=LCN,max_row=NHR)
for key,lbl in [("opt","оптимистичный"),("base","базовый"),("pess","пессимистичный")]:
    ref=Reference(wan,min_col=4,min_row=an[key],max_col=LCN,max_row=an[key])
    s=Series(ref,title=lbl); rc.series.append(s)
rc.set_categories(cats)
wan.add_chart(rc,"K22")
print("Аналитика готова")

# ================================================================= #
# СПРАВОЧНИК
# ================================================================= #
wr=wb.create_sheet(S_REF); wr.sheet_view.showGridLines=False
title(wr,"СПРАВОЧНИК ОРИЕНТИРОВОЧНЫХ ЗНАЧЕНИЙ","Диапазоны для проверки исходных данных. Требуют уточнения по данным поставщика.",last="I")
wr.column_dimensions["A"].width=2; wr.column_dimensions["B"].width=38
for cc in "CDEFGHI": wr.column_dimensions[cc].width=18
heads=["Показатель","Класс 250–270 мм","Диапазон","Комментарий"]
for i,h in enumerate(heads):
    c=wr.cell(row=4,column=2+i,value=h); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
refdata=[
 ("Ставка дисконтирования (WACC)","12%","8–16%","Согласуется с эскалацией"),
 ("Налог на прибыль","25%","20–25%","Действующая ставка РФ"),
 ("КТГ (техготовность)","0,85–0,93","0,80–0,95","Зависит от бренда и сервиса"),
 ("КИО (использование)","0,72–0,75","0,65–0,80","Доля рабочего времени в КФВ"),
 ("Часовая производительность","16–20","14–36 п.м/час","Мощность вращателя, крепость породы, диаметр"),
 ("Мощность вращателя","136–173","80–200 кВт","Ключевой драйвер скорости бурения"),
 ("Удельный расход ДТ","39–94","40–120 кг/ч","Двигатель и режим"),
 ("Выход с 1 пог.м","19–49","15–55 м³/пог.м","Сетка БВР, диаметр скважины, порода"),
 ("Срок амортизации (НУ)","61 мес","5–7 лет","Линейная амортизация"),
 ("Остаточная стоимость","0–10%","0–20%","Реализация в конце срока"),
 ("Долота (удельный расход)","1–3","0,5–5 шт/тыс.пог.м","Порода, диаметр, тип долота"),
]
rr=5
for row in refdata:
    for i,v in enumerate(row):
        c=wr.cell(row=rr,column=2+i,value=v); c.font=FN; c.border=BD; c.alignment=(L if i in(0,3) else C)
        if rr%2==0: c.fill=Fg
    rr+=1
rr+=1
wr.cell(row=rr,column=2,value="Методические примечания:").font=FB; rr+=1
for n in ["Часовая производительность считается по физике бурения (методика листа «W час» исходного файла): "
          "скорость бурения зависит от мощности вращателя, крепости породы и диаметра заряда.",
          "ТОиР задаётся по годам (ТО, текущие и капитальные ремонты, сервис) — «рваный» график капремонтов "
          "учитывается дисконтированием.",
          "Цена станков в исходном файле не заполнена — в модель подставлена ОЦЕНКА; замените жёлтые ячейки данными ТКП.",
          "Значения таблицы — ориентир порядка величин, а не норматив."]:
    wr.cell(row=rr,column=2,value="•  "+n).font=FN; wr.cell(row=rr,column=2).alignment=L; wr.merge_cells(f"B{rr}:I{rr}"); rr+=1
print("Справочник готов")

# ================================================================= #
# МЕТОДИКА (первым)
# ================================================================= #
wme=wb.create_sheet(S_MET,0); wme.sheet_view.showGridLines=False
title(wme,"МЕТОДИКА: АННУИТЕТ ЗАТРАТ (РУБ/ПОГ.М) НА ВЛАДЕНИЕ БУРОВЫМИ СТАНКАМИ","Пост-налоговый погодовой дисконтированный аннуитет жизненного цикла.")
wme.column_dimensions["A"].width=2; wme.column_dimensions["B"].width=4; wme.column_dimensions["C"].width=122
def mh(row,t):
    c=wme.cell(row=row,column=3,value=t); c.font=FSec; c.fill=Fm; c.alignment=L; wme.cell(row=row,column=2).fill=Fm
def mt(row,t,h=None):
    c=wme.cell(row=row,column=3,value=t); c.font=FN; c.alignment=Alignment("left","top",wrap_text=True)
    if h: wme.row_dimensions[row].height=h
blocks=[
 (4,"h","1. НАЗНАЧЕНИЕ"),
 (5,"t","Экономическое сравнение буровых станков одного класса по удельному аннуитету затрат (руб/пог.м пробуренной "
       "скважины, а также руб/м³ взорванной горной массы) за полный срок эксплуатации. Метрика корректно сопоставляет "
       "станки с разной ценой, производительностью и структурой затрат. Лучший вариант — минимум руб/пог.м.",50),
 (6,"h","2. ПРОИЗВОДИТЕЛЬНОСТЬ (методика бурения из файла «W час»)"),
 (7,"t","Теоретическая скорость бурения = (0,1 × мощность вращателя) / (2,73^(0,017 × крепость породы) × диаметр_заряда²); "
       "диаметр заряда = диаметр долота × коэффициент разбура. Время цикла скважины = бурение + продувка + наращивание + "
       "подъём става + разбор + переезд. Скважин/смену = смена / время цикла. Часовая производительность (п.м/час) = "
       "скважин/смену × средняя глубина / смена. КИО = КТГ × (1 − простои_смены/(60 × смена)); эффективное время = КФВ × КИО. "
       "Годовой объём бурения (тыс.пог.м) = часовая × эффективное время; объём взрывания (тыс.м³) = бурение × выход м³/пог.м.",70),
 (8,"h","3. ПОГОДОВОЙ ДЕНЕЖНЫЙ ПОТОК (10 лет)"),
 (9,"t","Для каждого года считаются: топливо/энергия (эфф.время × расход × цена), ТОиР и сервис (задаются ПО ГОДАМ на "
       "листах «Поставщик N» — ТО, текущие и капитальные ремонты; годы капремонтов дают всплеск затрат), буровой инструмент "
       "(долота, штанги), расходники, персонал, транспортный налог. Добавляются амортизация (НУ) и налог на прибыль (щит). "
       "Операционный поток = база + налог − амортизация. Инвестиции — в год 0, остаточная стоимость — в конце срока.",65),
 (10,"h","4. АННУИТЕТ (ЯДРО)"),
 (11,"t","Потоки дисконтируются и преобразуются в эквивалентный годовой платёж: Аннуитет = NPV затрат / Σ(дисконт-факторов). "
        "Инвестиционный аннуитет = (Цена − дисконт. остаточная стоимость) / Σ(дисконт-факторов). Операционный аннуитет — "
        "приведённые эксплуатационные затраты после налога. Общий = инвестиционный + операционный. Удельный аннуитет "
        "(руб/пог.м) = общий аннуитет / средний годовой объём бурения (пог.м); руб/м³ = руб/пог.м / выход м³ с 1 пог.м.",65),
 (12,"h","5. ПОРЯДОК РАБОТЫ"),
 (13,"t","«Параметры» → «Ввод данных» (цена, физика бурения: мощность вращателя, диаметр, скорость СПО; персонал) → листы "
        "«Поставщик 1…5»: по каждой машине задаются КТГ, простои и затраты ТОиР по годам, расход ДТ, буровой инструмент — "
        "КТГ и ИТОГО ТОиР считаются автоматически. Сводка «Данные по годам» тянет итоги сама → авторасчёт "
        "«Производительность», «Денежный поток», «Расчёт аннуитета» → «Дашборд», «Сравнение», «Аналитика», "
        "«Чувствительность». Правки — в листах «Поставщик N» и «Ввод данных».",70),
 (14,"h","6. ДОПУЩЕНИЯ"),
 (15,"t","• Станки сравниваются в одном классе и одной задаче; разница в производительности учтена через руб/пог.м и руб/м³.",None),
 (16,"t","• Данные примера — «в расчёте» из корпоративного файла 000_БСЛ-2027 (карьер «Сухой Лог»), горизонт 10 лет.",None),
 (17,"t","• Цена станков в файле поставщика не указана — подставлена ОЦЕНКА, требует уточнения (жёлтые ячейки «Ввод данных»).",None),
 (18,"t","• Ставка дисконтирования и эскалации должны быть согласованы (номинальные либо реальные).",30),
]
for it in blocks:
    row,kind=it[0],it[1]
    if kind=="h": mh(row,it[2])
    else: mt(row,it[2],*( (it[3],) if len(it)>3 and it[3] else () ))
lg=20
wme.cell(row=lg,column=3,value="Обозначения:").font=FB
wme.cell(row=lg+1,column=2).fill=Fi; wme.cell(row=lg+1,column=2).border=BD
wme.cell(row=lg+1,column=3,value="жёлтые ячейки — ввод данных").font=FN
wme.cell(row=lg+2,column=2).fill=Fr; wme.cell(row=lg+2,column=2).border=BD
wme.cell(row=lg+2,column=3,value="зелёные ячейки — итоговые показатели / лучший вариант").font=FN

# порядок листов и сохранение
order=[S_MET,S_DASH,S_PAR,S_IN]+TPL_NAMES+[S_PYR,S_PROD,S_CF,S_ANN,S_CMP,S_ANL,S_SENS,S_REF]
wb._sheets.sort(key=lambda s:order.index(s.title))
wb.active=1     # Дашборд
wb.calculation.fullCalcOnLoad=True
OUT="Модель_TCO_буровых_станков_аннуитет.xlsx"
wb.save(OUT)
try:
    from inject_cache import inject
    inject(OUT, OUT)
    print("СОХРАНЕНО (с кэшем формул):",OUT)
except Exception as e:
    print("СОХРАНЕНО (без кэша, ошибка инжектора):",OUT,"|",e)
