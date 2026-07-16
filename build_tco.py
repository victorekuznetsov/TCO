# -*- coding: utf-8 -*-
"""
Генератор Excel-методики оценки совокупной стоимости владения (TCO / LCC)
горных машин и оборудования одного класса.

Автор методики: экспертная модель Life Cycle Cost (LCC) с дисконтированием.
Формируемый файл: Методика_TCO_горного_оборудования.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# ------------------------------------------------------------------ #
#  Названия листов
# ------------------------------------------------------------------ #
S_METHOD  = "Методика"
S_PARAM   = "Параметры"
S_INPUT   = "Ввод данных"
S_CALC    = "Расчёт TCO"
S_COMPARE = "Сравнение"
S_SENS    = "Чувствительность"
S_REF     = "Справочник"

MODEL_COLS = ["C", "D", "E", "F"]          # столбцы 4-х вариантов оборудования
N_MODELS   = 4

# ------------------------------------------------------------------ #
#  Палитра / стили
# ------------------------------------------------------------------ #
CLR_DARK   = "1F3864"   # тёмно-синий (заголовки)
CLR_MID    = "2E5496"   # синий (секции)
CLR_ACCENT = "C55A11"   # оранжевый (акценты)
CLR_INPUT  = "FFF2CC"   # светло-жёлтый (ячейки ввода)
CLR_RESULT = "E2EFDA"   # светло-зелёный (итоги)
CLR_GREY   = "F2F2F2"   # серый фон
CLR_HEADR  = "D6DCE5"   # шапка таблиц
CLR_BEST   = "C6EFCE"   # лучший вариант

FILL_DARK   = PatternFill("solid", fgColor=CLR_DARK)
FILL_MID    = PatternFill("solid", fgColor=CLR_MID)
FILL_INPUT  = PatternFill("solid", fgColor=CLR_INPUT)
FILL_RESULT = PatternFill("solid", fgColor=CLR_RESULT)
FILL_GREY   = PatternFill("solid", fgColor=CLR_GREY)
FILL_HEADR  = PatternFill("solid", fgColor=CLR_HEADR)
FILL_ACCENT = PatternFill("solid", fgColor="FCE4D6")

F_TITLE   = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
F_SUB     = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
F_SECTION = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_HEAD    = Font(name="Calibri", size=10, bold=True, color=CLR_DARK)
F_BOLD    = Font(name="Calibri", size=10, bold=True)
F_NORM    = Font(name="Calibri", size=10)
F_RESULT  = Font(name="Calibri", size=10, bold=True, color="1F3864")
F_UNIT    = Font(name="Calibri", size=9, italic=True, color="808080")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

AL_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_R = Alignment(horizontal="right",  vertical="center")

# форматы чисел
FMT_MONEY = '#,##0'
FMT_MON2  = '#,##0.00'
FMT_PCT   = '0.0%'
FMT_PCT2  = '0.00%'
FMT_NUM1  = '#,##0.0'
FMT_NUM2  = '#,##0.00'
FMT_INT   = '#,##0'

wb = Workbook()

# ================================================================== #
#  Вспомогательные функции
# ================================================================== #
def style_title(ws, text, sub, span="B1:H2"):
    c0 = span.split(":")[0]
    ws[c0] = text
    ws.merge_cells(span.split(":")[0] + ":" + span.split(":")[1].replace("2", "1"))
    ws[c0].font = F_TITLE
    ws[c0].fill = FILL_DARK
    ws[c0].alignment = AL_L
    # подзаголовок
    row2span = span.replace("1", "2")
    parts = row2span.split(":")
    ws[parts[0]] = sub
    ws.merge_cells(parts[0] + ":" + parts[1])
    ws[parts[0]].font = F_SUB
    ws[parts[0]].fill = FILL_DARK
    ws[parts[0]].alignment = AL_L
    for r in (1, 2):
        for col in range(2, 9):
            ws.cell(row=r, column=col).fill = FILL_DARK


def section(ws, row, text, span_cols=("B", "H")):
    ws.merge_cells(f"{span_cols[0]}{row}:{span_cols[1]}{row}")
    c = ws[f"{span_cols[0]}{row}"]
    c.value = text
    c.font = F_SECTION
    c.fill = FILL_MID
    c.alignment = AL_L
    for col in range(ord(span_cols[0]) - 64, ord(span_cols[1]) - 64 + 1):
        ws.cell(row=row, column=col).fill = FILL_MID


# финансовые формулы (генераторы строк) --------------------------- #
def GAF(g, r, N):
    """Коэффициент растущего аннуитета (дисконт. сумма растущего потока)."""
    return (f"IF(ABS({r}-{g})<1E-9,{N}/(1+{r}),"
            f"(1-((1+{g})/(1+{r}))^{N})/({r}-{g}))")

def NGF(g, N):
    """Номинальная сумма растущего потока за N лет."""
    return f"IF({g}=0,{N},((1+{g})^{N}-1)/{g})"

def PVAF(r, n):
    """Коэффициент приведения аннуитета."""
    return f"IF({r}=0,{n},(1-(1+{r})^-{n})/{r})"

def ANNUITY(principal, rate, term):
    """Годовой аннуитетный платёж по кредиту."""
    return (f"IF({rate}=0,{principal}/{term},"
            f"{principal}*{rate}/(1-(1+{rate})^-{term}))")


# ================================================================== #
#  ЛИСТ 1. ПАРАМЕТРЫ (глобальные допущения)
# ================================================================== #
wp = wb.active
wp.title = S_PARAM
wp.sheet_view.showGridLines = False
style_title(wp, "ГЛОБАЛЬНЫЕ ПАРАМЕТРЫ И МАКРОДОПУЩЕНИЯ",
            "Единые для всех сравниваемых вариантов. Заполняйте жёлтые ячейки.")

wp.column_dimensions["A"].width = 2
wp.column_dimensions["B"].width = 48
wp.column_dimensions["C"].width = 14
wp.column_dimensions["D"].width = 16
wp.column_dimensions["E"].width = 52

# заголовок таблицы
hdr_r = 4
for col, txt in (("B", "Параметр"), ("C", "Значение"), ("D", "Ед. изм."), ("E", "Комментарий")):
    c = wp[f"{col}{hdr_r}"]
    c.value = txt; c.font = F_HEAD; c.fill = FILL_HEADR; c.alignment = AL_C; c.border = BORDER

# (row, key, label, value, unit, fmt, comment)
param_defs = [
    ("cur",       "Валюта расчёта",                              "USD",  "—",       None,     "Условная денежная единица (все суммы в ней)"),
    ("disc",      "Ставка дисконтирования (WACC)",               0.12,   "%/год",   FMT_PCT,  "Стоимость капитала для приведения к текущей стоимости"),
    ("inf",       "Общая инфляция",                              0.04,   "%/год",   FMT_PCT,  "Рост постоянных затрат (страхование, налоги)"),
    ("esc_fuel",  "Эскалация цены топлива/энергии",              0.05,   "%/год",   FMT_PCT,  "Прогнозный рост цен на дизтопливо/электроэнергию"),
    ("esc_lab",   "Эскалация оплаты труда",                      0.06,   "%/год",   FMT_PCT,  "Прогнозный рост ФОТ операторов"),
    ("esc_part",  "Эскалация запчастей и ремонта",               0.05,   "%/год",   FMT_PCT,  "Рост цен на ЗИП, шины/ГЭТ, услуги ТОиР"),
    ("fuel_p",    "Цена дизельного топлива",                     0.90,   "/л",      FMT_MON2, "Средняя цена на площадке эксплуатации"),
    ("elec_p",    "Цена электроэнергии",                         0.09,   "/кВт·ч",  FMT_MON2, "Для электрических/гибридных машин"),
    ("wage",      "Оплата труда оператора (полная)",             35.0,   "/ч",      FMT_MON2, "С учётом налогов, СИЗ, обучения, накладных"),
    ("prod_val",  "Ценность продукции (выручка), опц.",          0.0,    "/т",      FMT_MON2, "Для анализа влияния простоев; 0 = не учитывать"),
]
param_rows = {}
r = hdr_r + 1
for key, label, val, unit, fmt, comment in param_defs:
    param_rows[key] = r
    wp[f"B{r}"] = label; wp[f"B{r}"].font = F_NORM; wp[f"B{r}"].border = BORDER; wp[f"B{r}"].alignment = AL_L
    cell = wp[f"C{r}"]; cell.value = val; cell.fill = FILL_INPUT; cell.border = BORDER; cell.alignment = AL_C; cell.font = F_BOLD
    if fmt: cell.number_format = fmt
    wp[f"D{r}"] = unit; wp[f"D{r}"].font = F_UNIT; wp[f"D{r}"].alignment = AL_C; wp[f"D{r}"].border = BORDER
    wp[f"E{r}"] = comment; wp[f"E{r}"].font = F_UNIT; wp[f"E{r}"].alignment = AL_L; wp[f"E{r}"].border = BORDER
    r += 1

def par(key):
    return f"'{S_PARAM}'!C{param_rows[key]}"

# легенда
lr = r + 1
wp[f"B{lr}"] = "Легенда:"; wp[f"B{lr}"].font = F_BOLD
wp[f"C{lr}"].fill = FILL_INPUT; wp[f"C{lr}"].border = BORDER
wp[f"D{lr}"] = "— ячейки для ввода данных пользователем"; wp[f"D{lr}"].font = F_NORM; wp[f"D{lr}"].alignment = AL_L
wp.merge_cells(f"D{lr}:E{lr}")
wp[f"C{lr+1}"].fill = FILL_RESULT; wp[f"C{lr+1}"].border = BORDER
wp[f"D{lr+1}"] = "— расчётные итоговые показатели"; wp[f"D{lr+1}"].font = F_NORM; wp[f"D{lr+1}"].alignment = AL_L
wp.merge_cells(f"D{lr+1}:E{lr+1}")

print("Параметры done, param_rows:", param_rows)


# ================================================================== #
#  ЛИСТ 2. ВВОД ДАННЫХ (до 4-х вариантов оборудования)
# ================================================================== #
wi = wb.create_sheet(S_INPUT)
wi.sheet_view.showGridLines = False
style_title(wi, "ВВОД ИСХОДНЫХ ДАННЫХ ПО ВАРИАНТАМ ОБОРУДОВАНИЯ",
            "Сравнение до 4-х моделей одного класса. Пример: гидравлические экскаваторы ~250 т.")

wi.column_dimensions["A"].width = 2
wi.column_dimensions["B"].width = 46
for cc in MODEL_COLS:
    wi.column_dimensions[cc].width = 15
wi.column_dimensions["G"].width = 13
wi.column_dimensions["H"].width = 40

# шапка
HR = 4
wi[f"B{HR}"] = "Показатель"; wi[f"B{HR}"].font = F_HEAD; wi[f"B{HR}"].fill = FILL_HEADR; wi[f"B{HR}"].border = BORDER; wi[f"B{HR}"].alignment = AL_C
for i, cc in enumerate(MODEL_COLS):
    c = wi[f"{cc}{HR}"]; c.value = f"Вариант {i+1}"; c.font = F_HEAD; c.fill = FILL_HEADR; c.border = BORDER; c.alignment = AL_C
for cc, t in (("G", "Ед. изм."), ("H", "Пояснение")):
    c = wi[f"{cc}{HR}"]; c.value = t; c.font = F_HEAD; c.fill = FILL_HEADR; c.border = BORDER; c.alignment = AL_C

# описание слоёв ввода:  ("SEC", title) | ("ROW", key, label, unit, fmt, [v1..v4], comment, is_text)
LAYOUT = [
    ("SEC", "A. ИДЕНТИФИКАЦИЯ"),
    ("ROW", "name",  "Наименование / модель",       "—", None,  ["Экскаватор A", "Экскаватор B", "Экскаватор C", "Экскаватор D"], "Название варианта", True),
    ("ROW", "maker", "Производитель",               "—", None,  ["Поставщик 1", "Поставщик 2", "Поставщик 3", "Поставщик 4"], "", True),
    ("ROW", "klass", "Класс / тип",                 "—", None,  ["Гидравл. экскаватор"]*4, "Один класс для корректного сравнения", True),
    ("ROW", "drive", "Тип привода",                 "—", None,  ["Дизель"]*4, "Дизель / Электро / Гибрид", True),

    ("SEC", "B. ТЕХНИЧЕСКИЕ ХАРАКТЕРИСТИКИ"),
    ("ROW", "power", "Номинальная мощность",         "кВт", FMT_INT, [1500, 1400, 1600, 1520], "Мощность двигателя/привода", False),
    ("ROW", "mass",  "Эксплуатационная масса",       "т",   FMT_NUM1,[250, 248, 255, 251], "", False),
    ("ROW", "cap",   "Ёмкость ковша / грузоподъём.", "м³/т",FMT_NUM1,[15, 15, 16, 15], "Ковш (экскаватор) или г/п (самосвал)", False),

    ("SEC", "C. РЕЖИМ ЭКСПЛУАТАЦИИ"),
    ("ROW", "life",  "Срок службы (горизонт анализа)","лет", FMT_INT, [10, 10, 10, 10], "Расчётный срок владения", False),
    ("ROW", "hours", "Наработка в год",               "мото-ч/год", FMT_INT, [5500, 5500, 5500, 5500], "Годовой фонд рабочего времени машины", False),
    ("ROW", "avail", "Коэф. технической готовности",  "%", FMT_PCT, [0.85, 0.85, 0.85, 0.85], "Справочно (надёжность/ремонтопригодность)", False),
    ("ROW", "util",  "Коэф. использования",           "%", FMT_PCT, [0.85, 0.85, 0.85, 0.85], "Доля рабочего времени в готовом времени", False),
    ("ROW", "prod",  "Эксплуатац. производительность","т/мото-ч", FMT_INT, [2500, 2450, 2550, 2500], "Фактич. производит. за мото-час работы", False),

    ("SEC", "D. КАПИТАЛЬНЫЕ ЗАТРАТЫ (CAPEX)"),
    ("ROW", "price", "Цена приобретения",             "", FMT_MONEY, [4500000, 4200000, 5000000, 4700000], "Стоимость машины (без доставки)", False),
    ("ROW", "deliv", "Доставка, монтаж, ПНР",         "", FMT_MONEY, [150000, 150000, 160000, 155000], "Логистика, сборка, пуско-наладка", False),
    ("ROW", "zip",   "Стартовый комплект ЗИП",        "", FMT_MONEY, [120000, 110000, 140000, 125000], "Первичный запас запчастей", False),
    ("ROW", "resid", "Остаточная стоимость",          "% от цены", FMT_PCT, [0.14, 0.11, 0.16, 0.13], "Стоимость реализации в конце срока", False),
    ("ROW", "loan",  "Доля заёмного финансирования",  "%", FMT_PCT, [0.60, 0.60, 0.60, 0.60], "Кредит/лизинг; 0 = покупка за свои", False),
    ("ROW", "lrate", "Ставка кредита / лизинга",      "%/год", FMT_PCT, [0.08, 0.08, 0.08, 0.08], "Годовая ставка финансирования", False),
    ("ROW", "lterm", "Срок финансирования",           "лет", FMT_INT, [6, 6, 6, 6], "Период погашения кредита", False),

    ("SEC", "E. ЭНЕРГОНОСИТЕЛИ"),
    ("ROW", "fuel",  "Удельный расход топлива",        "л/мото-ч", FMT_NUM1, [300, 325, 275, 290], "Ключевой драйвер OPEX", False),
    ("ROW", "kwh",   "Удельный расход электроэнергии", "кВт·ч/мото-ч", FMT_NUM1, [0, 0, 0, 0], "Для электро/гибрид; иначе 0", False),
    ("ROW", "lube",  "Смазочные материалы",            "% от топлива", FMT_PCT, [0.12, 0.12, 0.12, 0.12], "Масла, смазки, охл. жидкости", False),

    ("SEC", "F. РАСХОДНЫЕ ЧАСТИ / ИЗНОС"),
    ("ROW", "tset",  "Стоимость комплекта шин/гус./ГЭТ","", FMT_MONEY, [350000, 340000, 360000, 350000], "Шины (самосвал) или ходовая (экскаватор)", False),
    ("ROW", "tlife", "Ресурс комплекта",               "мото-ч", FMT_INT, [15000, 14000, 16000, 15000], "Наработка до полной замены", False),

    ("SEC", "G. ТЕХОБСЛУЖИВАНИЕ И РЕМОНТ (ТОиР)"),
    ("ROW", "maint", "Удельные затраты на ТО (плановое)","/мото-ч", FMT_MON2, [40, 46, 34, 38], "Плановое ТО, фильтры, работы", False),
    ("ROW", "ohcost","Стоимость капитального ремонта",   "", FMT_MONEY, [700000, 650000, 800000, 720000], "Стоимость одного капремонта/rebuild", False),
    ("ROW", "ohint", "Периодичность капремонта",         "мото-ч", FMT_INT, [20000, 18000, 24000, 21000], "Наработка между капремонтами", False),

    ("SEC", "H. ПЕРСОНАЛ"),
    ("ROW", "crew",  "Операторов на машину (сменность)", "чел.·смен", FMT_NUM2, [1.05, 1.05, 1.05, 1.05], "Коэф. с учётом подмены; ЗП из 'Параметры'", False),

    ("SEC", "I. ПРОЧИЕ ГОДОВЫЕ ЗАТРАТЫ"),
    ("ROW", "ins",   "Страхование",                     "% от цены/год", FMT_PCT2, [0.010, 0.010, 0.010, 0.010], "Годовая ставка страхования", False),
    ("ROW", "tax",   "Налоги и прочее",                 "% от цены/год", FMT_PCT2, [0.005, 0.005, 0.005, 0.005], "Имущественные налоги, сборы", False),
]

input_rows = {}
r = HR + 2
for item in LAYOUT:
    if item[0] == "SEC":
        section(wi, r, item[1], ("B", "H"))
        r += 1
        continue
    _, key, label, unit, fmt, vals, comment, is_text = item
    input_rows[key] = r
    wi[f"B{r}"] = label; wi[f"B{r}"].font = F_NORM; wi[f"B{r}"].border = BORDER; wi[f"B{r}"].alignment = AL_L
    for i, cc in enumerate(MODEL_COLS):
        c = wi[f"{cc}{r}"]
        c.value = vals[i]
        c.fill = FILL_INPUT; c.border = BORDER
        c.alignment = AL_L if is_text else AL_C
        c.font = F_NORM
        if fmt: c.number_format = fmt
    wi[f"G{r}"] = unit; wi[f"G{r}"].font = F_UNIT; wi[f"G{r}"].alignment = AL_C; wi[f"G{r}"].border = BORDER
    wi[f"H{r}"] = comment; wi[f"H{r}"].font = F_UNIT; wi[f"H{r}"].alignment = AL_L; wi[f"H{r}"].border = BORDER
    r += 1

def inp(key, col):
    return f"'{S_INPUT}'!{col}{input_rows[key]}"

print("Ввод done, input_rows keys:", len(input_rows))


# ================================================================== #
#  ЛИСТ 3. РАСЧЁТ TCO (по каждому варианту)
# ================================================================== #
wc = wb.create_sheet(S_CALC)
wc.sheet_view.showGridLines = False
style_title(wc, "РАСЧЁТ СОВОКУПНОЙ СТОИМОСТИ ВЛАДЕНИЯ (TCO / LCC)",
            "Модель дисконтированного жизненного цикла. Значения рассчитываются автоматически.")

wc.column_dimensions["A"].width = 2
wc.column_dimensions["B"].width = 50
for cc in MODEL_COLS:
    wc.column_dimensions[cc].width = 16
wc.column_dimensions["G"].width = 14

CHR = 4
wc[f"B{CHR}"] = "Статья"; wc[f"B{CHR}"].font = F_HEAD; wc[f"B{CHR}"].fill = FILL_HEADR; wc[f"B{CHR}"].border = BORDER; wc[f"B{CHR}"].alignment = AL_C
for cc in MODEL_COLS:
    c = wc[f"{cc}{CHR}"]; c.value = f"='{S_INPUT}'!{cc}{input_rows['name']}"
    c.font = F_HEAD; c.fill = FILL_HEADR; c.border = BORDER; c.alignment = AL_C
wc[f"G{CHR}"] = "Ед. изм."; wc[f"G{CHR}"].font = F_HEAD; wc[f"G{CHR}"].fill = FILL_HEADR; wc[f"G{CHR}"].border = BORDER; wc[f"G{CHR}"].alignment = AL_C

calc_rows = {}
cr = CHR + 1

def add_sec(title):
    global cr
    section(wc, cr, title, ("B", "G"))
    cr += 1

def add_calc(key, label, unit, formula_fn, fmt=FMT_MONEY, result=False, bold=False):
    """formula_fn(col) -> строка формулы (без '=')."""
    global cr
    calc_rows[key] = cr
    b = wc[f"B{cr}"]; b.value = label; b.border = BORDER; b.alignment = AL_L
    b.font = F_RESULT if result else (F_BOLD if bold else F_NORM)
    for cc in MODEL_COLS:
        c = wc[f"{cc}{cr}"]
        c.value = "=" + formula_fn(cc)
        c.number_format = fmt; c.border = BORDER; c.alignment = AL_R
        c.font = F_RESULT if result else F_NORM
        if result: c.fill = FILL_RESULT
    wc[f"G{cr}"] = unit; wc[f"G{cr}"].font = F_UNIT; wc[f"G{cr}"].alignment = AL_C; wc[f"G{cr}"].border = BORDER
    cr += 1

def cc_ref(key, col):
    return f"{col}{calc_rows[key]}"

cur = par("cur")  # not used numerically

# --- Производственные показатели ---
add_sec("ПРОИЗВОДСТВЕННЫЕ ПОКАЗАТЕЛИ")
add_calc("tot_hours", "Наработка за срок службы", "мото-ч",
         lambda c: f"{inp('hours',c)}*{inp('life',c)}", FMT_INT)
add_calc("ann_prod", "Годовой объём производства", "т/год",
         lambda c: f"{inp('hours',c)}*{inp('prod',c)}", FMT_INT)
add_calc("tot_prod", "Полный объём производства за срок", "т",
         lambda c: f"{cc_ref('ann_prod',c)}*{inp('life',c)}", FMT_INT)

# --- Годовые эксплуатационные затраты (год 1) ---
add_sec("ГОДОВЫЕ ЭКСПЛУАТАЦИОННЫЕ ЗАТРАТЫ (в ценах 1-го года)")
add_calc("c_fuel", "Топливо", "",
         lambda c: f"{inp('hours',c)}*{inp('fuel',c)}*{par('fuel_p')}")
add_calc("c_elec", "Электроэнергия", "",
         lambda c: f"{inp('hours',c)}*{inp('kwh',c)}*{par('elec_p')}")
add_calc("c_lube", "Смазочные материалы", "",
         lambda c: f"{cc_ref('c_fuel',c)}*{inp('lube',c)}")
add_calc("c_tire", "Шины / гусеницы / ГЭТ", "",
         lambda c: f"{inp('tset',c)}/{inp('tlife',c)}*{inp('hours',c)}")
add_calc("c_maint", "ТО плановое", "",
         lambda c: f"{inp('maint',c)}*{inp('hours',c)}")
add_calc("c_over", "Капитальные ремонты (аннуитизировано)", "",
         lambda c: f"{inp('ohcost',c)}/{inp('ohint',c)}*{inp('hours',c)}")
add_calc("c_lab", "Оплата труда операторов", "",
         lambda c: f"{inp('hours',c)}*{inp('crew',c)}*{par('wage')}")
add_calc("c_ins", "Страхование", "",
         lambda c: f"{inp('price',c)}*{inp('ins',c)}")
add_calc("c_tax", "Налоги и прочее", "",
         lambda c: f"{inp('price',c)}*{inp('tax',c)}")
add_calc("c_opex", "ИТОГО годовые эксплуатационные (год 1)", "/год",
         lambda c: "+".join(cc_ref(k, c) for k in
                            ["c_fuel","c_elec","c_lube","c_tire","c_maint","c_over","c_lab","c_ins","c_tax"]),
         FMT_MONEY, bold=True)

# --- Группировка для эскалации ---
add_sec("ГРУППЫ ЗАТРАТ ДЛЯ ЭСКАЛАЦИИ (год 1)")
add_calc("g_energy", "Энергия (топливо+электро)", "",
         lambda c: f"{cc_ref('c_fuel',c)}+{cc_ref('c_elec',c)}")
add_calc("g_labor", "Труд", "",
         lambda c: f"{cc_ref('c_lab',c)}")
add_calc("g_parts", "Запчасти и ТОиР (смазка+ГЭТ+ТО+ремонт)", "",
         lambda c: f"{cc_ref('c_lube',c)}+{cc_ref('c_tire',c)}+{cc_ref('c_maint',c)}+{cc_ref('c_over',c)}")
add_calc("g_fixed", "Постоянные (страхование+налоги)", "",
         lambda c: f"{cc_ref('c_ins',c)}+{cc_ref('c_tax',c)}")

# --- Капитальные затраты и финансирование ---
add_sec("КАПИТАЛЬНЫЕ ЗАТРАТЫ И ФИНАНСИРОВАНИЕ")
add_calc("capex", "Первоначальные инвестиции (CAPEX)", "",
         lambda c: f"{inp('price',c)}+{inp('deliv',c)}+{inp('zip',c)}", bold=True)
add_calc("own", "Собственные средства (год 0)", "",
         lambda c: f"{inp('price',c)}*(1-{inp('loan',c)})+{inp('deliv',c)}+{inp('zip',c)}")
add_calc("princ", "Заёмные средства (тело кредита)", "",
         lambda c: f"{inp('price',c)}*{inp('loan',c)}")
add_calc("pay", "Годовой платёж по кредиту", "/год",
         lambda c: ANNUITY(cc_ref('princ',c), inp('lrate',c), inp('lterm',c)))
add_calc("pay_tot", "Всего выплат по кредиту (номинал)", "",
         lambda c: f"{cc_ref('pay',c)}*{inp('lterm',c)}")
add_calc("interest", "Переплата (проценты)", "",
         lambda c: f"{cc_ref('pay_tot',c)}-{cc_ref('princ',c)}")
add_calc("resid_v", "Остаточная стоимость (год N)", "",
         lambda c: f"{inp('price',c)}*{inp('resid',c)}")

# --- Финансовые коэффициенты ---
add_sec("ФИНАНСОВЫЕ КОЭФФИЦИЕНТЫ")
add_calc("gaf_e", "Коэф. приведения энергии", "",
         lambda c: GAF(par('esc_fuel'), par('disc'), inp('life',c)), FMT_NUM2)
add_calc("gaf_l", "Коэф. приведения труда", "",
         lambda c: GAF(par('esc_lab'), par('disc'), inp('life',c)), FMT_NUM2)
add_calc("gaf_p", "Коэф. приведения запчастей/ТОиР", "",
         lambda c: GAF(par('esc_part'), par('disc'), inp('life',c)), FMT_NUM2)
add_calc("gaf_f", "Коэф. приведения постоянных", "",
         lambda c: GAF(par('inf'), par('disc'), inp('life',c)), FMT_NUM2)
add_calc("pvaf_n", "Коэф. аннуитета проекта (r, N)", "",
         lambda c: PVAF(par('disc'), inp('life',c)), FMT_NUM2)
add_calc("pvaf_l", "Коэф. аннуитета кредита (r, срок)", "",
         lambda c: PVAF(par('disc'), inp('lterm',c)), FMT_NUM2)
add_calc("df_n", "Дисконт остаточной стоимости", "",
         lambda c: f"1/(1+{par('disc')})^{inp('life',c)}", FMT_NUM2)

# --- NPV затрат ---
add_sec("ДИСКОНТИРОВАННЫЕ ЗАТРАТЫ (NPV, текущая стоимость)")
add_calc("npv_e", "NPV энергии", "",
         lambda c: f"{cc_ref('g_energy',c)}*{cc_ref('gaf_e',c)}")
add_calc("npv_l", "NPV труда", "",
         lambda c: f"{cc_ref('g_labor',c)}*{cc_ref('gaf_l',c)}")
add_calc("npv_p", "NPV запчастей и ТОиР", "",
         lambda c: f"{cc_ref('g_parts',c)}*{cc_ref('gaf_p',c)}")
add_calc("npv_f", "NPV постоянных затрат", "",
         lambda c: f"{cc_ref('g_fixed',c)}*{cc_ref('gaf_f',c)}")
add_calc("npv_own", "NPV собственных средств (год 0)", "",
         lambda c: f"{cc_ref('own',c)}")
add_calc("npv_fin", "NPV платежей по кредиту", "",
         lambda c: f"{cc_ref('pay',c)}*{cc_ref('pvaf_l',c)}")
add_calc("npv_res", "NPV остаточной стоимости (вычет)", "",
         lambda c: f"{cc_ref('resid_v',c)}*{cc_ref('df_n',c)}")
add_calc("tco_npv", "TCO — совокупная стоимость владения (NPV)", "",
         lambda c: "+".join(cc_ref(k, c) for k in
                            ["npv_e","npv_l","npv_p","npv_f","npv_own","npv_fin"]) + f"-{cc_ref('npv_res',c)}",
         FMT_MONEY, result=True)

# --- Номинальные затраты за срок ---
add_sec("СОВОКУПНЫЕ НОМИНАЛЬНЫЕ ЗАТРАТЫ ЗА СРОК СЛУЖБЫ (с эскалацией)")
add_calc("nom_e", "Энергия (номинал)", "",
         lambda c: f"{cc_ref('g_energy',c)}*{NGF(par('esc_fuel'), inp('life',c))}")
add_calc("nom_l", "Труд (номинал)", "",
         lambda c: f"{cc_ref('g_labor',c)}*{NGF(par('esc_lab'), inp('life',c))}")
add_calc("nom_p", "Запчасти и ТОиР (номинал)", "",
         lambda c: f"{cc_ref('g_parts',c)}*{NGF(par('esc_part'), inp('life',c))}")
add_calc("nom_f", "Постоянные (номинал)", "",
         lambda c: f"{cc_ref('g_fixed',c)}*{NGF(par('inf'), inp('life',c))}")
add_calc("nom_cap", "Капзатраты (свои + выплаты по кредиту)", "",
         lambda c: f"{cc_ref('own',c)}+{cc_ref('pay_tot',c)}")
add_calc("tco_nom", "TCO — совокупные затраты (номинал)", "",
         lambda c: "+".join(cc_ref(k, c) for k in
                            ["nom_e","nom_l","nom_p","nom_f","nom_cap"]) + f"-{cc_ref('resid_v',c)}",
         FMT_MONEY, bold=True)

# --- Ключевые показатели ---
add_sec("КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ ЭФФЕКТИВНОСТИ")
add_calc("m_tco", "TCO (NPV)", "",
         lambda c: f"{cc_ref('tco_npv',c)}", FMT_MONEY, result=True)
add_calc("m_eac", "EAC — эквивалентная годовая стоимость", "/год",
         lambda c: f"{cc_ref('tco_npv',c)}/{cc_ref('pvaf_n',c)}", FMT_MONEY, result=True)
add_calc("m_hour", "Удельный TCO на мото-час", "/мото-ч",
         lambda c: f"{cc_ref('tco_npv',c)}/{cc_ref('tot_hours',c)}", FMT_MON2, result=True)
add_calc("m_eachr", "Приведённые затраты на мото-час (EAC)", "/мото-ч",
         lambda c: f"{cc_ref('m_eac',c)}/{inp('hours',c)}", FMT_MON2, result=True)
add_calc("m_tonne", "Удельный TCO на тонну продукции", "/т",
         lambda c: f"{cc_ref('m_eac',c)}/{cc_ref('ann_prod',c)}", FMT_MON2, result=True)
add_calc("m_capex_share", "Доля CAPEX в TCO", "%",
         lambda c: f"({cc_ref('npv_own',c)}+{cc_ref('npv_fin',c)}-{cc_ref('npv_res',c)})/{cc_ref('tco_npv',c)}",
         FMT_PCT, result=True)
add_calc("m_opex_share", "Доля эксплуатации в TCO", "%",
         lambda c: f"({cc_ref('npv_e',c)}+{cc_ref('npv_l',c)}+{cc_ref('npv_p',c)}+{cc_ref('npv_f',c)})/{cc_ref('tco_npv',c)}",
         FMT_PCT, result=True)

def calc(key, col):
    return f"'{S_CALC}'!{col}{calc_rows[key]}"

print("Расчёт done, calc_rows:", len(calc_rows), "last row:", cr)


# ================================================================== #
#  ЛИСТ 4. СРАВНЕНИЕ (дашборд)
# ================================================================== #
wm = wb.create_sheet(S_COMPARE)
wm.sheet_view.showGridLines = False
style_title(wm, "СРАВНИТЕЛЬНЫЙ АНАЛИЗ ВАРИАНТОВ ОБОРУДОВАНИЯ",
            "Итоговый дашборд. Зелёным выделен лучший вариант по удельному TCO на тонну.")

wm.column_dimensions["A"].width = 2
wm.column_dimensions["B"].width = 44
for cc in MODEL_COLS:
    wm.column_dimensions[cc].width = 16
wm.column_dimensions["G"].width = 13

# шапка
MHR = 4
wm[f"B{MHR}"] = "Показатель"; wm[f"B{MHR}"].font = F_HEAD; wm[f"B{MHR}"].fill = FILL_HEADR; wm[f"B{MHR}"].border = BORDER; wm[f"B{MHR}"].alignment = AL_C
for cc in MODEL_COLS:
    c = wm[f"{cc}{MHR}"]; c.value = f"='{S_INPUT}'!{cc}{input_rows['name']}"
    c.font = F_HEAD; c.fill = FILL_HEADR; c.border = BORDER; c.alignment = AL_C
wm[f"G{MHR}"] = "Ед."; wm[f"G{MHR}"].font = F_HEAD; wm[f"G{MHR}"].fill = FILL_HEADR; wm[f"G{MHR}"].border = BORDER; wm[f"G{MHR}"].alignment = AL_C

# метрики
metrics = [
    ("Цена приобретения",                  "price", "inp",  FMT_MONEY, ""),
    ("Первоначальные инвестиции (CAPEX)",  "capex", "calc", FMT_MONEY, ""),
    ("Годовые эксплуатац. затраты (год 1)","c_opex","calc", FMT_MONEY, "/год"),
    ("Годовой объём производства",         "ann_prod","calc",FMT_INT,  "т/год"),
    ("Полный объём за срок службы",        "tot_prod","calc",FMT_INT,  "т"),
    ("TCO — совокупная стоимость (NPV)",   "tco_npv","calc",FMT_MONEY, ""),
    ("TCO — совокупные затраты (номинал)", "tco_nom","calc",FMT_MONEY, ""),
    ("EAC — эквивалентная годовая стоим.", "m_eac", "calc", FMT_MONEY, "/год"),
    ("Удельный TCO на мото-час",           "m_hour","calc", FMT_MON2,  "/мото-ч"),
    ("Удельный TCO на тонну",              "m_tonne","calc",FMT_MON2,  "/т"),
    ("Доля CAPEX в TCO",                   "m_capex_share","calc",FMT_PCT,"%"),
    ("Доля эксплуатации в TCO",            "m_opex_share","calc",FMT_PCT,"%"),
]
mrow = MHR + 1
metric_rows = {}
for label, key, src, fmt, unit in metrics:
    metric_rows[key] = mrow
    b = wm[f"B{mrow}"]; b.value = label; b.font = F_NORM; b.border = BORDER; b.alignment = AL_L
    for cc in MODEL_COLS:
        c = wm[f"{cc}{mrow}"]
        c.value = "=" + (inp(key, cc) if src == "inp" else calc(key, cc))
        c.number_format = fmt; c.border = BORDER; c.alignment = AL_R; c.font = F_NORM
    wm[f"G{mrow}"] = unit; wm[f"G{mrow}"].font = F_UNIT; wm[f"G{mrow}"].alignment = AL_C; wm[f"G{mrow}"].border = BORDER
    mrow += 1

# итоговая оценка / рейтинг
section(wm, mrow, "ИТОГОВАЯ ОЦЕНКА (критерий: минимум удельного TCO на тонну)", ("B", "G"))
mrow += 1
rank_row = mrow
tonne_rng = f"{MODEL_COLS[0]}{metric_rows['m_tonne']}:{MODEL_COLS[-1]}{metric_rows['m_tonne']}"
wm[f"B{rank_row}"] = "Место в рейтинге"; wm[f"B{rank_row}"].font = F_BOLD; wm[f"B{rank_row}"].border = BORDER; wm[f"B{rank_row}"].alignment = AL_L
for cc in MODEL_COLS:
    c = wm[f"{cc}{rank_row}"]
    c.value = f"=RANK({cc}{metric_rows['m_tonne']},{tonne_rng},1)"
    c.font = F_BOLD; c.border = BORDER; c.alignment = AL_C
mrow += 1
dev_row = mrow
wm[f"B{dev_row}"] = "Отклонение от лучшего, %"; wm[f"B{dev_row}"].font = F_NORM; wm[f"B{dev_row}"].border = BORDER; wm[f"B{dev_row}"].alignment = AL_L
for cc in MODEL_COLS:
    c = wm[f"{cc}{dev_row}"]
    c.value = f"={cc}{metric_rows['m_tonne']}/MIN({tonne_rng})-1"
    c.number_format = FMT_PCT; c.border = BORDER; c.alignment = AL_C; c.font = F_NORM
mrow += 2

# рекомендация
best_r = mrow
names_rng = f"{MODEL_COLS[0]}{MHR}:{MODEL_COLS[-1]}{MHR}"
wm[f"B{best_r}"] = "РЕКОМЕНДУЕМЫЙ ВАРИАНТ:"; wm[f"B{best_r}"].font = Font(bold=True, size=12, color=CLR_DARK)
wm[f"B{best_r}"].alignment = AL_L
wm.merge_cells(f"C{best_r}:F{best_r}")
rec = wm[f"C{best_r}"]
rec.value = (f'=INDEX({names_rng},MATCH(MIN({tonne_rng}),{tonne_rng},0))'
             f'&"  ("&TEXT(MIN({tonne_rng}),"0.00")&" /т)"')
rec.font = Font(bold=True, size=12, color="006100"); rec.fill = PatternFill("solid", fgColor=CLR_BEST)
rec.alignment = AL_C
for cc in "CDEF":
    wm[f"{cc}{best_r}"].fill = PatternFill("solid", fgColor=CLR_BEST)
    wm[f"{cc}{best_r}"].border = BORDER
mrow += 2

# --- Структура затрат (NPV) для диаграммы ---
struct_hdr = mrow
section(wm, struct_hdr, "СТРУКТУРА ЗАТРАТ TCO (NPV) ПО КАТЕГОРИЯМ", ("B", "G"))
mrow += 1
# шапка
wm[f"B{mrow}"] = "Категория"; wm[f"B{mrow}"].font = F_HEAD; wm[f"B{mrow}"].fill = FILL_HEADR; wm[f"B{mrow}"].border = BORDER; wm[f"B{mrow}"].alignment = AL_C
for cc in MODEL_COLS:
    c = wm[f"{cc}{mrow}"]; c.value = f"='{S_INPUT}'!{cc}{input_rows['name']}"
    c.font = F_HEAD; c.fill = FILL_HEADR; c.border = BORDER; c.alignment = AL_C
struct_top = mrow + 1
struct_cats = [
    ("Капитальные (за вычетом остат. ст-ти)", lambda c: f"{calc('npv_own',c)}+{calc('npv_fin',c)}-{calc('npv_res',c)}"),
    ("Энергия / топливо",                     lambda c: f"{calc('npv_e',c)}"),
    ("Оплата труда",                          lambda c: f"{calc('npv_l',c)}"),
    ("Запчасти и ТОиР",                        lambda c: f"{calc('npv_p',c)}"),
    ("Постоянные (страхование, налоги)",      lambda c: f"{calc('npv_f',c)}"),
]
mrow = struct_top
for label, fn in struct_cats:
    b = wm[f"B{mrow}"]; b.value = label; b.font = F_NORM; b.border = BORDER; b.alignment = AL_L
    for cc in MODEL_COLS:
        c = wm[f"{cc}{mrow}"]; c.value = "=" + fn(cc)
        c.number_format = FMT_MONEY; c.border = BORDER; c.alignment = AL_R; c.font = F_NORM
    mrow += 1
struct_bot = mrow - 1
# строка ИТОГО
tot_struct_row = mrow
b = wm[f"B{mrow}"]; b.value = "ИТОГО TCO (NPV)"; b.font = F_BOLD; b.border = BORDER
for cc in MODEL_COLS:
    c = wm[f"{cc}{mrow}"]; c.value = f"=SUM({cc}{struct_top}:{cc}{struct_bot})"
    c.number_format = FMT_MONEY; c.border = BORDER; c.alignment = AL_R; c.font = F_BOLD; c.fill = FILL_RESULT
mrow += 2

# ---- ДИАГРАММА 1: структура затрат (stacked col), серии = категории ----
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef

cats_ref = Reference(wm, min_col=3, min_row=struct_top - 1, max_col=6, max_row=struct_top - 1)
chart1 = BarChart()
chart1.type = "col"; chart1.grouping = "stacked"; chart1.overlap = 100
chart1.title = "Структура TCO (NPV) по вариантам"
chart1.height = 9; chart1.width = 18
chart1.y_axis.title = "Ден. ед. (NPV)"
for i in range(len(struct_cats)):
    row_i = struct_top + i
    vals = Reference(wm, min_col=3, min_row=row_i, max_col=6, max_row=row_i)
    ser = Series(vals, title_from_data=False)
    ser.tx = SeriesLabel(strRef=StrRef(f"'{S_COMPARE}'!$B${row_i}"))
    chart1.series.append(ser)
chart1.set_categories(cats_ref)
wm.add_chart(chart1, "I4")

# ---- ДИАГРАММА 2: удельный TCO на тонну ----
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Удельный TCO на тонну продукции"
chart2.height = 8; chart2.width = 9
chart2.y_axis.title = "/т"
d2 = Reference(wm, min_col=3, min_row=metric_rows['m_tonne'], max_col=6, max_row=metric_rows['m_tonne'])
c2 = Reference(wm, min_col=3, min_row=MHR, max_col=6, max_row=MHR)
chart2.add_data(d2, from_rows=True, titles_from_data=False)
chart2.set_categories(c2)
chart2.legend = None
chart2.dataLabels = DataLabelList(); chart2.dataLabels.showVal = True; chart2.dataLabels.numFmt = '0.00'
wm.add_chart(chart2, "I23")

# ---- ДИАГРАММА 3: TCO (NPV) итог ----
chart3 = BarChart()
chart3.type = "col"
chart3.title = "Совокупная стоимость владения TCO (NPV)"
chart3.height = 8; chart3.width = 9
chart3.y_axis.title = "Ден. ед."
d3 = Reference(wm, min_col=3, min_row=metric_rows['tco_npv'], max_col=6, max_row=metric_rows['tco_npv'])
chart3.add_data(d3, from_rows=True, titles_from_data=False)
chart3.set_categories(c2)
chart3.legend = None
wm.add_chart(chart3, "R23")

print("Сравнение done; metric_rows:", metric_rows, "struct:", struct_top, struct_bot)


# ================================================================== #
#  ЛИСТ 5. ЧУВСТВИТЕЛЬНОСТЬ
# ================================================================== #
ws = wb.create_sheet(S_SENS)
ws.sheet_view.showGridLines = False
style_title(ws, "АНАЛИЗ ЧУВСТВИТЕЛЬНОСТИ",
            "Торнадо-анализ ключевых драйверов и чувствительность к утилизации.")

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 40
for cc in "CDEFG":
    ws.column_dimensions[cc].width = 16

# выбор варианта
ws["B4"] = "Анализируемый вариант (1–4):"; ws["B4"].font = F_BOLD; ws["B4"].alignment = AL_L
sel = ws["C4"]; sel.value = 1; sel.fill = FILL_INPUT; sel.border = BORDER; sel.alignment = AL_C; sel.font = F_BOLD
SEL = "$C$4"
dv = DataValidation(type="whole", operator="between", formula1="1", formula2="4", allow_blank=False)
ws.add_data_validation(dv); dv.add(sel)
name_row = input_rows['name']
ws["D4"] = f"=INDEX('{S_INPUT}'!{MODEL_COLS[0]}{name_row}:{MODEL_COLS[-1]}{name_row},{SEL})"
ws["D4"].font = Font(bold=True, color=CLR_ACCENT); ws["D4"].alignment = AL_L
ws.merge_cells("D4:F4")

def selcalc(key):
    return f"INDEX('{S_CALC}'!{MODEL_COLS[0]}{calc_rows[key]}:{MODEL_COLS[-1]}{calc_rows[key]},{SEL})"

# базовые величины (вспомогательные ячейки) ------------------------ #
section(ws, 6, "БАЗОВЫЕ ПОКАЗАТЕЛИ ВЫБРАННОГО ВАРИАНТА", ("B", "G"))
base_defs = [
    ("b_tco",   "TCO (NPV)",                     lambda: selcalc('tco_npv'), FMT_MONEY),
    ("b_eac",   "EAC — эквивалентная годовая",    lambda: selcalc('m_eac'),   FMT_MONEY),
    ("b_tonne", "Удельный TCO на тонну",          lambda: selcalc('m_tonne'), FMT_MON2),
    ("b_pvaf",  "Коэф. аннуитета (r, N)",         lambda: selcalc('pvaf_n'),  FMT_NUM2),
    ("b_prod",  "Годовой объём производства, т",  lambda: selcalc('ann_prod'),FMT_INT),
    ("b_e",     "NPV энергии",                    lambda: selcalc('npv_e'),   FMT_MONEY),
    ("b_l",     "NPV труда",                      lambda: selcalc('npv_l'),   FMT_MONEY),
    ("b_p",     "NPV запчастей и ТОиР",           lambda: selcalc('npv_p'),   FMT_MONEY),
    ("b_cap",   "NPV капитальных (нетто)",        lambda: f"{selcalc('npv_own')}+{selcalc('npv_fin')}-{selcalc('npv_res')}+{selcalc('npv_f')}", FMT_MONEY),
    ("b_var",   "NPV переменных (энергия+труд+ТОиР)", lambda: f"{selcalc('npv_e')}+{selcalc('npv_l')}+{selcalc('npv_p')}", FMT_MONEY),
]
base_rows = {}
br = 7
for key, label, fn, fmt in base_defs:
    base_rows[key] = br
    ws[f"B{br}"] = label; ws[f"B{br}"].font = F_NORM; ws[f"B{br}"].border = BORDER; ws[f"B{br}"].alignment = AL_L
    c = ws[f"C{br}"]; c.value = "=" + fn(); c.number_format = fmt; c.border = BORDER; c.alignment = AL_R; c.font = F_NORM
    br += 1

def bref(key):
    return f"$C${base_rows[key]}"

# --- Торнадо ---
tr = br + 1
section(ws, tr, "ТОРНАДО-АНАЛИЗ: ВЛИЯНИЕ ДРАЙВЕРОВ НА TCO (NPV)", ("B", "G"))
tr += 1
hdr = ["Драйвер", "Диапазон", "TCO при снижении", "TCO при росте", "Размах", "Влияние, %"]
for i, h in enumerate(hdr):
    c = ws.cell(row=tr, column=2 + i); c.value = h; c.font = F_HEAD; c.fill = FILL_HEADR; c.border = BORDER; c.alignment = AL_C
tr += 1
# (название, ±доля, ячейка базовой категории)
drivers = [
    ("Цена топлива / энергии", 0.20, "b_e"),
    ("Оплата труда",           0.15, "b_l"),
    ("Затраты на ТОиР",        0.25, "b_p"),
    ("Цена приобретения",      0.15, "b_cap"),
]
tornado_top = tr
for label, pct, cat in drivers:
    ws.cell(row=tr, column=2, value=label).font = F_NORM
    ws.cell(row=tr, column=2).border = BORDER; ws.cell(row=tr, column=2).alignment = AL_L
    ws.cell(row=tr, column=3, value=f"±{int(pct*100)}%").alignment = AL_C
    ws.cell(row=tr, column=3).border = BORDER; ws.cell(row=tr, column=3).font = F_NORM
    low = ws.cell(row=tr, column=4, value=f"={bref('b_tco')}-{pct}*{bref(cat)}")
    high = ws.cell(row=tr, column=5, value=f"={bref('b_tco')}+{pct}*{bref(cat)}")
    swing = ws.cell(row=tr, column=6, value=f"=E{tr}-D{tr}")
    infl = ws.cell(row=tr, column=7, value=f"=F{tr}/{bref('b_tco')}")
    for col, fmt in ((4, FMT_MONEY), (5, FMT_MONEY), (6, FMT_MONEY), (7, FMT_PCT)):
        cell = ws.cell(row=tr, column=col); cell.number_format = fmt; cell.border = BORDER; cell.alignment = AL_R; cell.font = F_NORM
    tr += 1
tornado_bot = tr - 1

# отклонения для диаграммы-торнадо (скрытые вспомогательные столбцы I,J)
ws.cell(row=tornado_top - 1, column=9, value="низкое, откл.").font = F_UNIT
ws.cell(row=tornado_top - 1, column=10, value="высокое, откл.").font = F_UNIT
for i in range(tornado_top, tornado_bot + 1):
    ws.cell(row=i, column=9, value=f"=D{i}-{bref('b_tco')}").number_format = FMT_MONEY
    ws.cell(row=i, column=10, value=f"=E{i}-{bref('b_tco')}").number_format = FMT_MONEY

# диаграмма-торнадо (гориз. бары, stacked от нуля)
tchart = BarChart()
tchart.type = "bar"; tchart.grouping = "stacked"; tchart.overlap = 100
tchart.title = "Торнадо: отклонение TCO (NPV) от базового"
tchart.height = 7; tchart.width = 16
dlow = Reference(ws, min_col=9, min_row=tornado_top, max_col=9, max_row=tornado_bot)
dhigh = Reference(ws, min_col=10, min_row=tornado_top, max_col=10, max_row=tornado_bot)
cats = Reference(ws, min_col=2, min_row=tornado_top, max_col=2, max_row=tornado_bot)
tchart.add_data(dlow, titles_from_data=False)
tchart.add_data(dhigh, titles_from_data=False)
tchart.set_categories(cats)
tchart.series[0].graphicalProperties.solidFill = "70AD47"
tchart.series[1].graphicalProperties.solidFill = "C55A11"
tchart.legend = None
ws.add_chart(tchart, "I8")

# --- Чувствительность к утилизации ---
ur = tornado_bot + 3
section(ws, ur, "ЧУВСТВИТЕЛЬНОСТЬ К УТИЛИЗАЦИИ (НАРАБОТКЕ)", ("B", "G"))
ur += 1
ws.cell(row=ur, column=2, value="Изменение наработки").font = F_HEAD
ws.cell(row=ur, column=2).fill = FILL_HEADR; ws.cell(row=ur, column=2).border = BORDER; ws.cell(row=ur, column=2).alignment = AL_C
util_levels = [-0.15, -0.10, 0.0, 0.10, 0.15]
for j, d in enumerate(util_levels):
    c = ws.cell(row=ur, column=3 + j, value=f"{'+' if d>0 else ''}{int(d*100)}%")
    c.font = F_HEAD; c.fill = FILL_HEADR; c.border = BORDER; c.alignment = AL_C
ur += 1
# TCO(NPV) при изменении: fixed_part = b_tco - b_var; new = fixed + b_var*(1+d)
ws.cell(row=ur, column=2, value="TCO (NPV)").font = F_NORM
ws.cell(row=ur, column=2).border = BORDER; ws.cell(row=ur, column=2).alignment = AL_L
for j, d in enumerate(util_levels):
    c = ws.cell(row=ur, column=3 + j,
                value=f"=({bref('b_tco')}-{bref('b_var')})+{bref('b_var')}*(1+{d})")
    c.number_format = FMT_MONEY; c.border = BORDER; c.alignment = AL_R; c.font = F_NORM
tco_util_row = ur
ur += 1
ws.cell(row=ur, column=2, value="Удельный TCO на тонну").font = F_BOLD
ws.cell(row=ur, column=2).border = BORDER; ws.cell(row=ur, column=2).alignment = AL_L
for j, d in enumerate(util_levels):
    col_letter = get_column_letter(3 + j)
    # EAC = TCO/pvaf; prod = b_prod*(1+d)
    c = ws.cell(row=ur, column=3 + j,
                value=f"=({col_letter}{tco_util_row}/{bref('b_pvaf')})/({bref('b_prod')}*(1+{d}))")
    c.number_format = FMT_MON2; c.border = BORDER; c.alignment = AL_R; c.font = F_BOLD
    if abs(d) < 1e-9:
        c.fill = FILL_RESULT
ur += 2
ws.cell(row=ur, column=2, value="Примечание: рост наработки распределяет постоянные затраты на больший объём — "
        "удельный TCO на тонну снижается (эффект масштаба использования).").font = F_UNIT
ws.merge_cells(f"B{ur}:G{ur}")
ws[f"B{ur}"].alignment = AL_L

print("Чувствительность done; base_rows:", base_rows)


# ================================================================== #
#  ЛИСТ 6. СПРАВОЧНИК (ориентировочные значения по классам)
# ================================================================== #
wr = wb.create_sheet(S_REF)
wr.sheet_view.showGridLines = False
style_title(wr, "СПРАВОЧНИК ОРИЕНТИРОВОЧНЫХ ЗНАЧЕНИЙ",
            "Диапазоны для проверки исходных данных. Требуют уточнения по данным поставщика/эксплуатации.",
            span="B1:J2")

wr.column_dimensions["A"].width = 2
wr.column_dimensions["B"].width = 30
for cc in "CDEFGHIJ":
    wr.column_dimensions[cc].width = 15

ref_headers = ["Класс оборудования", "Мощность, кВт", "Расход топлива, л/мото-ч",
               "Ресурс ходовой/шин, мото-ч", "Плановое ТО, /мото-ч",
               "Периодич. капрем., мото-ч", "Ресурс до списания, мото-ч",
               "Коэф. готовности", "Остат. ст-ть, %"]
hr = 4
for i, h in enumerate(ref_headers):
    c = wr.cell(row=hr, column=2 + i, value=h); c.font = F_HEAD; c.fill = FILL_HEADR; c.border = BORDER; c.alignment = AL_C
ref_data = [
    ["Карьерный самосвал (90–100 т)",  "750–900",   "45–75",   "5 000–7 000 (шины)", "12–22", "18 000–24 000", "60 000–90 000", "80–88%", "10–18%"],
    ["Карьерный самосвал (220 т+)",    "1 800–2 700","180–280", "6 000–9 000 (шины)", "22–40", "18 000–24 000", "60 000–100 000","82–90%", "12–20%"],
    ["Гидравлический экскаватор ~250 т","1 300–1 700","260–340", "12 000–18 000 (ход.)","30–50", "18 000–24 000", "60 000–90 000", "80–88%", "10–16%"],
    ["Электрический канатный экскав.",  "2 000–4 500","(электро)","40 000–60 000 (ход.)","25–45", "40 000–60 000", "120 000–200 000","85–92%","8–15%"],
    ["Колёсный погрузчик (10–15 м³)",  "500–900",   "60–110",  "4 000–6 000 (шины)", "18–32", "16 000–20 000", "50 000–70 000", "80–88%", "12–20%"],
    ["Бульдозер (крупный, 60–100 т)",  "400–700",   "40–70",   "8 000–12 000 (ход.)","20–35", "15 000–20 000", "45 000–70 000", "78–86%", "10–18%"],
    ["Буровой станок (шарошечный)",     "400–900",   "50–90",   "—",                  "20–40", "18 000–24 000", "60 000–90 000", "80–88%", "10–18%"],
]
rr = hr + 1
for row in ref_data:
    for i, v in enumerate(row):
        c = wr.cell(row=rr, column=2 + i, value=v)
        c.font = F_NORM; c.border = BORDER
        c.alignment = AL_L if i == 0 else AL_C
        if rr % 2 == 0:
            c.fill = FILL_GREY
    rr += 1

rr += 1
notes = [
    "Смазочные материалы: обычно 8–15% от стоимости топлива.",
    "Оплата труда: закладывается «полная» ставка (налоги, СИЗ, обучение, накладные, подмена).",
    "Эскалация: цена топлива и ФОТ, как правило, растут быстрее общей инфляции.",
    "Ставка дисконтирования = WACC компании (обычно 8–15% в реальном или номинальном выражении — согласуйте с эскалацией).",
    "Капремонты (rebuild) двигателя/гидравлики/трансмиссии учитываются как аннуитизированные затраты по наработке.",
    "Значения таблицы — ориентир для проверки порядка величин, а не норматив.",
]
wr.cell(row=rr, column=2, value="Методические примечания:").font = F_BOLD
rr += 1
for n in notes:
    wr.cell(row=rr, column=2, value="•  " + n).font = F_NORM
    wr.cell(row=rr, column=2).alignment = AL_L
    wr.merge_cells(f"B{rr}:J{rr}")
    rr += 1

print("Справочник done")


# ================================================================== #
#  ЛИСТ 0. МЕТОДИКА (описание, помещается первым)
# ================================================================== #
wme = wb.create_sheet(S_METHOD, 0)
wme.sheet_view.showGridLines = False
style_title(wme, "МЕТОДИКА ОЦЕНКИ СТОИМОСТИ ВЛАДЕНИЯ ГОРНЫХ МАШИН (TCO / LCC)",
            "Сравнение вариантов оборудования одного класса на основе полного жизненного цикла.",
            span="B1:J2")
wme.column_dimensions["A"].width = 2
wme.column_dimensions["B"].width = 4
wme.column_dimensions["C"].width = 120

def mtext(row, text, font=F_NORM, fill=None, height=None):
    c = wme.cell(row=row, column=3, value=text)
    c.font = font; c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if fill: c.fill = fill
    if height: wme.row_dimensions[row].height = height

def mhead(row, text):
    c = wme.cell(row=row, column=3, value=text)
    c.font = F_SECTION; c.fill = FILL_MID; c.alignment = AL_L
    wme.cell(row=row, column=2).fill = FILL_MID

blocks = [
    (4,  "mhead", "1. НАЗНАЧЕНИЕ"),
    (5,  "t", "Модель предназначена для экономического сравнения нескольких моделей горного оборудования одного класса "
              "(экскаваторы, самосвалы, погрузчики, бульдозеры, буровые станки) по критерию совокупной стоимости владения "
              "за полный жизненный цикл (Total Cost of Ownership / Life Cycle Cost). Результат — обоснованный выбор варианта "
              "с минимальными приведёнными затратами на единицу продукции.", 60),
    (6,  "mhead", "2. СОСТАВ ЗАТРАТ ЖИЗНЕННОГО ЦИКЛА"),
    (7,  "t", "• Капитальные (CAPEX): цена приобретения, доставка/монтаж/ПНР, стартовый ЗИП, финансирование (кредит/лизинг), "
              "за вычетом остаточной стоимости.", 30),
    (8,  "t", "• Энергоносители: дизельное топливо или электроэнергия (ключевой драйвер OPEX).", None),
    (9,  "t", "• Смазочные материалы и технические жидкости.", None),
    (10, "t", "• Расходные части и износ: шины (самосвалы, погрузчики) или ходовая часть/ГЭТ (экскаваторы, бульдозеры).", 30),
    (11, "t", "• ТОиР: плановое техническое обслуживание и капитальные ремонты (rebuild) основных узлов.", None),
    (12, "t", "• Оплата труда операторов (с учётом сменности и полной стоимости персонала).", None),
    (13, "t", "• Постоянные затраты: страхование, налоги, сборы.", None),
    (14, "mhead", "3. МЕТОДОЛОГИЯ РАСЧЁТА"),
    (15, "t", "Все потоки затрат приводятся к текущей стоимости (NPV) по ставке дисконтирования (WACC). Затраты каждой "
              "категории растут с индивидуальной эскалацией (топливо, труд, запчасти, инфляция) и приводятся через "
              "коэффициент растущего аннуитета GAF(g, r, N) = (1 − ((1+g)/(1+r))^N) / (r − g).", 60),
    (16, "t", "Капитальные ремонты аннуитизируются по наработке: годовые затраты = стоимость ремонта / межремонтная "
              "наработка × годовая наработка.", 30),
    (17, "t", "Финансирование: собственные средства учитываются в год 0; платежи по кредиту (аннуитет) дисконтируются "
              "за срок финансирования. Переплата по процентам входит в TCO.", 30),
    (18, "t", "Итоговые метрики: TCO (NPV); EAC — эквивалентная годовая стоимость = TCO / коэф. аннуитета(r, N); "
              "удельный TCO на мото-час; удельный TCO на тонну продукции (levelized cost) = EAC / годовой объём.", 45),
    (19, "mhead", "4. ПОРЯДОК РАБОТЫ"),
    (20, "t", "Шаг 1. Лист «Параметры» — задайте макродопущения (ставка дисконтирования, цены топлива и труда, эскалации).", None),
    (21, "t", "Шаг 2. Лист «Ввод данных» — заполните жёлтые ячейки по каждому из 1–4 вариантов (один класс техники).", None),
    (22, "t", "Шаг 3. Лист «Расчёт TCO» — автоматический расчёт (просмотр детализации, менять не требуется).", None),
    (23, "t", "Шаг 4. Лист «Сравнение» — итоговый дашборд, рейтинг, рекомендация и диаграммы.", None),
    (24, "t", "Шаг 5. Лист «Чувствительность» — торнадо-анализ драйверов и влияние утилизации.", None),
    (25, "t", "Лист «Справочник» — ориентировочные диапазоны для проверки исходных данных.", None),
    (26, "mhead", "5. ВАЖНЫЕ ДОПУЩЕНИЯ И ОГРАНИЧЕНИЯ"),
    (27, "t", "• Сравниваемые машины должны быть одного класса и решать одну производственную задачу (сопоставимая "
              "производительность). Разницу в производительности модель учитывает через удельный TCO на тонну.", 30),
    (28, "t", "• Затраты на ремонт равномерно распределены по наработке (без явного «старения»). При необходимости "
              "заложите рост через удельные затраты ТОиР.", 30),
    (29, "t", "• Ставка дисконтирования и эскалации должны быть согласованы (номинальные либо реальные).", None),
    (30, "t", "• Ориентировочные значения справочника не заменяют данные поставщика и статистику эксплуатации.", None),
]
for item in blocks:
    row, kind = item[0], item[1]
    if kind == "mhead":
        mhead(row, item[2])
    else:
        text = item[2]
        h = item[3] if len(item) > 3 else None
        mtext(row, text, F_NORM, height=h)

# легенда цветов
lgr = 32
wme.cell(row=lgr, column=3, value="Обозначения:").font = F_BOLD
wme.cell(row=lgr+1, column=2).fill = FILL_INPUT; wme.cell(row=lgr+1, column=2).border = BORDER
wme.cell(row=lgr+1, column=3, value="жёлтые ячейки — ввод данных пользователем").font = F_NORM
wme.cell(row=lgr+2, column=2).fill = FILL_RESULT; wme.cell(row=lgr+2, column=2).border = BORDER
wme.cell(row=lgr+2, column=3, value="зелёные ячейки — итоговые показатели / лучший вариант").font = F_NORM

print("Методика done")

# ================================================================== #
#  Порядок листов и сохранение
# ================================================================== #
order = [S_METHOD, S_PARAM, S_INPUT, S_CALC, S_COMPARE, S_SENS, S_REF]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active = 0

wb.calculation.fullCalcOnLoad = True   # принудительный пересчёт при открытии

OUT = "Методика_TCO_горного_оборудования.xlsx"
wb.save(OUT)
print("СОХРАНЕНО:", OUT)





