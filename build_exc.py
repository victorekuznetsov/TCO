# -*- coding: utf-8 -*-
"""
Расширенная расчётная модель TCO горных экскаваторов (методика аннуитета руб/м3).
Полный погодовой дисконтированный жизненный цикл (пост-налоговый), воспроизводящий
и развивающий логику корпоративной модели сравнения экскаваторов.

Возможности:
  - погодовой движок (10 лет): производство, топливо, ТОиР, капремонты по графику,
    расходники, персонал, амортизация, налог на прибыль (щит), эскалация цен,
    остаточная стоимость, снижение КТГ с возрастом;
  - удельный аннуитет руб/м3 и руб/т (операционный + инвестиционный, разбивка по статьям);
  - исполнительный дашборд, траектория аннуитета по годам;
  - сравнение с рейтингом, торнадо и чувствительность к КТГ и ставке дисконтирования;
  - справочник ориентировочных значений.

Файл: Модель_TCO_экскаваторов_аннуитет.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef
from openpyxl.worksheet.datavalidation import DataValidation

S_MET="Методика"; S_DASH="Дашборд"; S_PAR="Параметры"; S_IN="Ввод данных"
S_PYR="Данные по годам"; S_TPL="Шаблон поставщика"
S_PROD="Производительность"; S_CF="Денежный поток"; S_ANN="Расчёт аннуитета"
S_CMP="Сравнение"; S_ANL="Аналитика"; S_SENS="Чувствительность"; S_REF="Справочник"

EXC=["D","E","F","G","H","I"]; NEXC=6
H=10                                   # горизонт, лет (периоды 0..10)
PCOLS=[get_column_letter(4+p) for p in range(H+1)]   # D..N (период 0..10)

# ---- палитра ----
DARK="1F3864"; MID="2E5496"; ACC="C55A11"; INP="FFF2CC"; RES="E2EFDA"; HEAD="D6DCE5"
GREY="F2F2F2"; BEST="C6EFCE"; TILE="1F3864"
Fd=PatternFill("solid",fgColor=DARK); Fm=PatternFill("solid",fgColor=MID); Fi=PatternFill("solid",fgColor=INP)
Fr=PatternFill("solid",fgColor=RES); Fh=PatternFill("solid",fgColor=HEAD); Fg=PatternFill("solid",fgColor=GREY)
Fb=PatternFill("solid",fgColor=BEST); Ftile=PatternFill("solid",fgColor=TILE); Fa=PatternFill("solid",fgColor="FCE4D6")
FT=Font(size=15,bold=True,color="FFFFFF"); FSub=Font(size=10,italic=True,color="FFFFFF")
FSec=Font(size=11,bold=True,color="FFFFFF"); FHd=Font(size=9,bold=True,color=DARK)
FB=Font(size=10,bold=True); FN=Font(size=10); FRz=Font(size=10,bold=True,color=DARK); FU=Font(size=8,italic=True,color="808080")
thin=Side(style="thin",color="BFBFBF"); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
L=Alignment("left",vertical="center",wrap_text=True); C=Alignment("center",vertical="center",wrap_text=True); Rr=Alignment("right",vertical="center")
M="#,##0"; M1="#,##0.0"; M2="#,##0.00"; P1="0.0%"; NUM="#,##0.000"

wb=Workbook()

def title(ws,t,s,last="K"):
    ws.merge_cells(f"B1:{last}1"); ws["B1"]=t; ws["B1"].font=FT; ws["B1"].alignment=L
    ws.merge_cells(f"B2:{last}2"); ws["B2"]=s; ws["B2"].font=FSub; ws["B2"].alignment=L
    for r in (1,2):
        for col in range(2,ord(last[-1])-64+1 if len(last)==1 else 40): ws.cell(row=r,column=col).fill=Fd
def section(ws,row,text,c0="B",c1="K"):
    ws.merge_cells(f"{c0}{row}:{c1}{row}"); cc=ws[f"{c0}{row}"]; cc.value=text; cc.font=FSec; cc.alignment=L
    a0=ws[f"{c0}1"].column; a1=ws[f"{c1}1"].column
    for col in range(a0,a1+1): ws.cell(row=row,column=col).fill=Fm
def PVAF(r,n): return f"IF({r}=0,{n},(1-(1+{r})^-{n})/{r})"

# ================================================================= #
# ПАРАМЕТРЫ
# ================================================================= #
wp=wb.active; wp.title=S_PAR; wp.sheet_view.showGridLines=False
title(wp,"ГЛОБАЛЬНЫЕ ПАРАМЕТРЫ МОДЕЛИ","Единые допущения для всех вариантов. Жёлтые ячейки — ввод.",last="F")
wp.column_dimensions["A"].width=2; wp.column_dimensions["B"].width=46
wp.column_dimensions["C"].width=13; wp.column_dimensions["D"].width=12; wp.column_dimensions["E"].width=52
def phdr(row):
    for col,t in (("B","Параметр"),("C","Значение"),("D","Ед."),("E","Комментарий")):
        c=wp[f"{col}{row}"]; c.value=t; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
prow={}
def padd(row,key,label,val,unit,fmt,comment):
    prow[key]=row
    wp[f"B{row}"]=label; wp[f"B{row}"].font=FN; wp[f"B{row}"].border=BD; wp[f"B{row}"].alignment=L
    c=wp[f"C{row}"]; c.value=val; c.fill=Fi; c.border=BD; c.alignment=C; c.font=FB
    if fmt: c.number_format=fmt
    wp[f"D{row}"]=unit; wp[f"D{row}"].font=FU; wp[f"D{row}"].alignment=C; wp[f"D{row}"].border=BD
    wp[f"E{row}"]=comment; wp[f"E{row}"].font=FU; wp[f"E{row}"].alignment=L; wp[f"E{row}"].border=BD
def pcalc(row,key,label,formula,unit,fmt,comment):
    prow[key]=row
    wp[f"B{row}"]=label; wp[f"B{row}"].font=FN; wp[f"B{row}"].border=BD; wp[f"B{row}"].alignment=L
    c=wp[f"C{row}"]; c.value="="+formula; c.fill=Fr; c.border=BD; c.alignment=C; c.font=FRz
    if fmt: c.number_format=fmt
    wp[f"D{row}"]=unit; wp[f"D{row}"].font=FU; wp[f"D{row}"].alignment=C; wp[f"D{row}"].border=BD
    wp[f"E{row}"]=comment; wp[f"E{row}"].font=FU; wp[f"E{row}"].alignment=L; wp[f"E{row}"].border=BD
def pc(key): return f"'{S_PAR}'!C{prow[key]}"

section(wp,4,"Финансовые допущения","B","E"); phdr(5)
padd(6,"disc","Ставка дисконтирования (WACC)",0.12,"%/год",P1,"Стоимость капитала для приведения к текущей стоимости")
padd(7,"tax","Налог на прибыль",0.25,"%",P1,"Формирует налоговый щит на затраты и амортизацию")
padd(8,"horizon","Срок эксплуатации (горизонт)",10,"лет",M,"Фиксирован = 10 лет в погодовом движке")
padd(9,"dep_years","Срок амортизации (налоговый учёт)",5.17,"лет",M2,"Линейная; 62 мес ≈ 5,17 года")
padd(10,"fuel_p","Цена дизельного топлива",54.83,"руб/кг","#,##0.00","Средняя цена на площадке эксплуатации")

section(wp,12,"Эскалация цен (реальный/номинальный рост)","B","E"); phdr(13)
padd(14,"esc_fuel","Эскалация цены топлива",0.0,"%/год",P1,"Рост цены ДТ (0 = данные уже с учётом инфляции)")
padd(15,"esc_parts","Эскалация расходников и персонала",0.0,"%/год",P1,"ТОиР задаётся по годам напрямую (см. «Данные по годам»)")
padd(16,"esc_lab","Эскалация оплаты труда",0.0,"%/год",P1,"Рост ФОТ экипажей")

section(wp,18,"Курсы валют","B","E"); phdr(19)
padd(20,"cny","Курс юаня (CNY)",11.2068,"руб",NUM,"Для оборудования из Китая")
padd(21,"usd","Курс доллара (USD)",76.1258,"руб",NUM,"")
padd(22,"eur","Курс евро (EUR)",86.8976,"руб",NUM,"")

section(wp,24,"Параметры производительности (карьер)","B","E"); phdr(25)
padd(26,"kfv","Календарный фонд времени",8760,"час/год",M,"24 ч × 365 дн")
padd(27,"dens","Плотность породы",2.787,"т/м³",NUM,"В целике; используется и для пересчёта руб/т")
padd(28,"loosen","Коэффициент разрыхления",1.5,"коэф.",M2,"")
padd(29,"fill","Коэффициент наполнения ковша",0.9,"коэф.",M2,"")
padd(30,"loss","Коэффициент потерь при экскавации",0.9,"коэф.",M2,"")
padd(31,"cap","Грузоподъёмность самосвала",136,"т",M,"Эталонный самосвал для расчёта")
padd(32,"pos","Время постановки самосвала",32,"сек",M,"Время оборота а/с под погрузку")
padd(33,"shift","Длительность смены",12,"час",M,"Для пересчёта простоев")

section(wp,35,"Расчётные финансовые коэффициенты","B","E"); phdr(36)
pcalc(37,"S","Коэф. аннуитета Σдиск.факторов (r, N)",PVAF(pc('disc'),pc('horizon')),"коэф.",NUM,"Сумма дисконт-факторов за срок службы")
_d=pc('dep_years'); _r=pc('disc')
_dsf=f"(1/{_d})*((1-(1+{_r})^-INT({_d}))/{_r}+({_d}-INT({_d}))*(1+{_r})^-(INT({_d})+1))"
pcalc(38,"dsf","Коэф. приведённой амортизации",_dsf,"коэф.",NUM,"Приведённая амортизация на 1 руб. цены (налог. щит)")

lg=40
wp[f"B{lg}"]="Легенда:"; wp[f"B{lg}"].font=FB
wp[f"C{lg}"].fill=Fi; wp[f"C{lg}"].border=BD; wp[f"D{lg}"]="— ввод данных"; wp[f"D{lg}"].font=FN; wp.merge_cells(f"D{lg}:E{lg}")
wp[f"C{lg+1}"].fill=Fr; wp[f"C{lg+1}"].border=BD; wp[f"D{lg+1}"]="— расчётные показатели"; wp[f"D{lg+1}"].font=FN; wp.merge_cells(f"D{lg+1}:E{lg+1}")
print("Параметры:",len(prow))

# ================================================================= #
# ВВОД ДАННЫХ
# ================================================================= #
wi=wb.create_sheet(S_IN); wi.sheet_view.showGridLines=False
title(wi,"ВВОД ИСХОДНЫХ ДАННЫХ ПО ЭКСКАВАТОРАМ","Сравнение до 6 машин одного класса. Пример: экскаваторы класса 2000 (~12 м³).")
wi.column_dimensions["A"].width=2; wi.column_dimensions["B"].width=44; wi.column_dimensions["C"].width=12
for cc in EXC: wi.column_dimensions[cc].width=15
wi.column_dimensions["K"].width=34
wi.freeze_panes="D5"
HR=4
wi[f"B{HR}"]="Показатель"; wi[f"B{HR}"].font=FHd; wi[f"B{HR}"].fill=Fh; wi[f"B{HR}"].border=BD; wi[f"B{HR}"].alignment=C
wi[f"C{HR}"]="Ед."; wi[f"C{HR}"].font=FHd; wi[f"C{HR}"].fill=Fh; wi[f"C{HR}"].border=BD; wi[f"C{HR}"].alignment=C
for i,cc in enumerate(EXC):
    c=wi[f"{cc}{HR}"]; c.value=f"Экскаватор {i+1}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
wi[f"K{HR}"]="Пояснение"; wi[f"K{HR}"].font=FHd; wi[f"K{HR}"].fill=Fh; wi[f"K{HR}"].border=BD; wi[f"K{HR}"].alignment=C

names=["Komatsu PC2000-8\nMining Solutions","Komatsu PC2000-11R\nИСТК","Zoomlion ZE2000G\nОрикс",
       "Sany SY2000\nАСТ","Shantui SE2000LCW","TZCO TZ2000\nДзикай"]
supplier=["Mining Solutions","ИСТК","Орикс","АСТ","Строймпорттехника","Дзикай"]
engines=["Komatsu SAA12V140E","Komatsu SAA12V140E-7","Cummins QST30","Cummins QSK38","Weichai 12M33","Weichai 12M33"]
price_val=[3800,3401.64,20129,19621.62,16742.05,15400]
price_cur=["USD","USD","CNY","CNY","CNY","CNY"]
install=[0,0,0,0,0,0]; resid=[0.0]*6
bucket=[10,10,10,10,9.6,12]; cycle=[37.3]*6; downtime=[266.1]*6
fuel_kgh=[88.46,85.81,90.23,92.04,102,120]
bucket_m3=[0.87]*6; personnel=[10015]*6

LAY=[
 ("SEC","A. ИДЕНТИФИКАЦИЯ"),
 ("R","name","Наименование / модель","—",names,"Название варианта",True,None),
 ("R","supplier","Поставщик","—",supplier,"",True,None),
 ("R","engine","Двигатель","—",engines,"",True,None),
 ("SEC","B. СТОИМОСТЬ И ОСТАТОЧНАЯ ЦЕНА"),
 ("R","price_val","Цена за единицу","тыс.вал.",price_val,"В валюте контракта",False,M),
 ("R","price_cur","Валюта цены","—",price_cur,"CNY / USD / EUR / RUB",True,None),
 ("R","install","Монтаж, доставка, ПНР","тыс.руб",install,"Дополнительно к цене",False,M),
 ("R","resid","Остаточная стоимость","% от цены",resid,"Стоимость реализации в конце срока (0 = не учитывать)",False,P1),
 ("SEC","C. ПРОИЗВОДИТЕЛЬНОСТЬ"),
 ("R","bucket","Геометрический объём ковша","м³",bucket,"Участвует в расчёте производительности",False,M1),
 ("R","cycle","Время цикла экскавации","сек",cycle,"Черпание–поворот–разгрузка–поворот",False,M1),
 ("R","downtime","Ежесменные простои","мин/см",downtime,"ОТМ, ожидание а/с, прочие организационные",False,M1),
 ("SEC","D. ПРОЧИЕ ГОДОВЫЕ ЗАТРАТЫ"),
 ("R","personnel","Расходы на персонал (экипаж)","тыс.руб/год",personnel,"ФОТ с отчислениями на 1 машину",False,M),
 ("NOTE","КТГ, расход ДТ, ТОиР и расходники на ковш задаются ПО ГОДАМ на листе «Данные по годам». "
         "Пустой бланк для сбора данных с поставщика — на листе «Шаблон поставщика»."),
]
irow={}; r=HR+2
for it in LAY:
    if it[0]=="SEC": section(wi,r,it[1],"B","K"); r+=1; continue
    if it[0]=="NOTE":
        wi.merge_cells(f"B{r}:K{r}"); nc=wi[f"B{r}"]; nc.value="ℹ  "+it[1]; nc.font=Font(size=9,bold=True,color=ACC); nc.alignment=L
        nc.fill=Fa; r+=1; continue
    _,key,label,unit,vals,comment,istext,fmt=it
    irow[key]=r
    wi[f"B{r}"]=label; wi[f"B{r}"].font=FN; wi[f"B{r}"].border=BD; wi[f"B{r}"].alignment=L
    wi[f"C{r}"]=unit; wi[f"C{r}"].font=FU; wi[f"C{r}"].border=BD; wi[f"C{r}"].alignment=C
    for i,cc in enumerate(EXC):
        c=wi[f"{cc}{r}"]; c.value=vals[i]; c.fill=Fi; c.border=BD; c.font=FN
        c.alignment=L if istext else C
        if fmt: c.number_format=fmt
    wi[f"K{r}"]=comment; wi[f"K{r}"].font=FU; wi[f"K{r}"].border=BD; wi[f"K{r}"].alignment=L
    r+=1
dv=DataValidation(type="list",formula1='"CNY,USD,EUR,RUB"',allow_blank=False); wi.add_data_validation(dv)
dv.add(f"D{irow['price_cur']}:I{irow['price_cur']}")
def inp(key,col): return f"'{S_IN}'!{col}{irow[key]}"
print("Ввод:",len(irow))

# ================================================================= #
# ЛИСТЫ-ШАБЛОНЫ ПОСТАВЩИКОВ (по одному на машину) + СВОДКА «Данные по годам»
# ================================================================= #
# Реальные данные "в расчёте" из исходной модели (АО «Полюс»), 10 лет
KTG_Y={
 0:[0.924063,0.9,0.86,0.85,0.86,0.812148,0.786563,0.842639,0.737381,0.826705],
 1:[0.924063,0.797738,0.824534,0.763944,0.762815,0.812148,0.786563,0.842639,0.737381,0.826705],
 2:[0.854652,0.728327,0.755123,0.694534,0.693404,0.742737,0.717152,0.773228,0.66797,0.757295],
 3:[0.854652,0.728327,0.755123,0.694534,0.693404,0.742737,0.717152,0.773228,0.66797,0.757295],
 4:[0.854652,0.728327,0.755123,0.694534,0.693404,0.742737,0.717152,0.773228,0.66797,0.757295],
 5:[0.854652,0.728327,0.755123,0.694534,0.693404,0.742737,0.717152,0.773228,0.66797,0.757295]}
TOIR_Y={  # тыс. валюты / год; включает ТО, текущие и капитальные ремонты, сервис
 0:[100.666,141.1002,633.3182,1430.6176,211.666,673.7523,141.1002,1390.1835,784.7523,100.666],
 1:[100.666,141.1002,633.3182,1430.6176,211.666,673.7523,141.1002,1390.1835,784.7523,100.666],
 2:[160.7578,178.0231,1021.9348,178.0231,291.7228,1039.2001,160.7578,178.0231,1021.9348,308.9881],
 3:[337.7887,530.6868,1544.312,531.869,658.6154,1741.348,442.6356,786.9538,1587.0,467.2187],
 4:[160.7578,178.0231,1021.9348,178.0231,291.7228,1039.2001,160.7578,178.0231,1021.9348,308.9881],
 5:[170.997,177.397,170.997,177.397,310.197,177.397,2540.277,177.397,170.997,2718.997]}
TOIR_CUR=["USD","USD","USD","USD","USD","USD"]
FUEL_Y={i:[fuel_kgh[i]]*H for i in range(NEXC)}
CONS_Y={i:[bucket_m3[i]]*H for i in range(NEXC)}
KFV_LIT=8760
TPL_NAMES=[f"Поставщик {i+1}" for i in range(NEXC)]

# --- 6 листов-шаблонов поставщиков: модели ТО и Ремонтов + полные вводимые списки ---
# Сводка: КТГ=11, ИТОГО ТОиР=16, ДТ=17, расходники=18, валюта=D5
TO_INT=[500,1000,2000,4000,5000]; TO_DUR=[10.2,9.4,12.4,7.8,5.4]; TO_PERS=[2,2,2,2,2]
NARA=7000.0
_to_dt=sum(NARA/TO_INT[k]*TO_DUR[k] for k in range(5))
_to_ch=sum(NARA/TO_INT[k]*TO_DUR[k]*TO_PERS[k] for k in range(5))
RATE=50.0
_to_serv=_to_ch*RATE/1000
# полный каталог запчастей ТО: (наименование, цена за ед, кол-во на ТО, № ТО)
TO_PARTS=[
 ("Масляный фильтр",114.83,4,1),("Топливный фильтр",141.33,2,1),("Масло ДВС",6.87,130,1),
 ("Фильтр управляющего контура",126,1,1),("Сливной фильтр",30.67,2,1),("Сливной фильтр насоса",201,3,1),
 ("Смазка",12.24,200,1),("Топливный фильтр (ТО-2)",80.17,2,2),("Фильтр сапуна топл.бака",19.17,1,2),
 ("Фильтр сапуна гидробака",19.17,2,2),("Гидравлический фильтр",344.33,3,2),("О-кольцо",24.5,3,2),
 ("Воздушный фильтр",401.33,2,2),("О-кольцо (2)",9.83,2,2),("Масло редуктора поворота",6.89,60,2),
 ("Масло механизма отбора мощн.",6.89,36,2),("Фильтр гидромотора поворота",102,2,2),
 ("Прокладка клапанной крышки",4.17,12,3),("Фильтр салонный кондиц.",10.83,2,3),("Фильтр наружный кондиц.",51.33,2,3),
 ("Масло конечной передачи",6.89,170,3),("Фильтр охлажд. жидкости",134.83,2,4),("Ремень генератора",28,1,4),
 ("Ремень генератора (2)",157.5,1,4),("Ремень кондиционера",61,2,4),("Охлаждающая жидкость",2.36,190,4),
 ("Гидравлическое масло",6.43,1300,5)]
_cat_sum=sum(NARA/TO_INT[t-1]*q*p for _,p,q,t in TO_PARTS)/1000   # тыс/год
# список узлов/видов ремонта: (наименование, нормо-часы)
REM_ITEMS=[
 ("Звёздочка (ходовая)",24),("Редуктор хода",24),("Натяжное колесо",24),("Гусеничная лента",36),
 ("Гидромотор хода",12),("Поддерживающий каток",12),("Опорный каток",36),
 ("Г.цилиндр стрелы",12),("Г.цилиндр стрелы (2)",12),("Г.цилиндр рукояти",8),("Г.цилиндр ковша",12),
 ("Г.распределитель",36),("Г.распределитель (2)",36),("Г.насос управления",24),("Г.насос",36),
 ("Г.мотор вентилятора",8),("Г.мотор поворота",24),("Поворотный круг",144),("Редуктор поворота",24),
 ("ДВС (капремонт)",72),("Механизм отбора мощности",48)]

# позиции строк
CAT0=38; CATN=CAT0+len(TO_PARTS)-1; CORR=CATN+1                    # каталог ТО
REMSEC=CORR+2; REMHDR=REMSEC+1; REM0=REMHDR+1; REMN=REM0+len(REM_ITEMS)-1; REMPLUG=REMN+1; NOTEROW=REMPLUG+2
tpl={}
for i in range(NEXC):
    ws=wb.create_sheet(TPL_NAMES[i]); ws.sheet_view.showGridLines=False
    title(ws,f"ШАБЛОН ДАННЫХ ПОСТАВЩИКА — ВАРИАНТ {i+1}",
          "Модели ТО и Ремонтов с полными списками. Зелёное — авто; сводка тянется в «Данные по годам».",last="N")
    ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=38; ws.column_dimensions["C"].width=11; ws.column_dimensions["D"].width=11
    for p in range(1,H+1): ws.column_dimensions[PCOLS[p]].width=10
    ws.freeze_panes="E7"
    ws["B4"]="Экскаватор:"; ws["B4"].font=FB
    ws.merge_cells("E4:N4"); ws["E4"]=f"='{S_IN}'!{EXC[i]}{irow['name']}"; ws["E4"].font=Font(bold=True,color=DARK); ws["E4"].alignment=L
    ws["B5"]="Валюта затрат ТОиР:"; ws["B5"].font=FN
    cu=ws["D5"]; cu.value=TOIR_CUR[i]; cu.fill=Fi; cu.border=BD; cu.alignment=C; cu.font=FB
    dvt=DataValidation(type="list",formula1='"CNY,USD,EUR,RUB"'); ws.add_data_validation(dvt); dvt.add("D5")
    ws["B6"]="Показатель"; ws["C6"]="Ед."
    for cc in ("B","C","D"):
        ws[f"{cc}6"].font=FHd; ws[f"{cc}6"].fill=Fh; ws[f"{cc}6"].border=BD; ws[f"{cc}6"].alignment=C
    for p in range(1,H+1):
        c=ws[f"{PCOLS[p]}6"]; c.value=f"год {p}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
    def yline(r,label,unit,fmt,green=False,formula=None,values=None,inp=False):
        b=ws[f"B{r}"]; b.value=label; b.border=BD; b.alignment=L; b.font=(FRz if green else FN)
        ws[f"C{r}"]=unit; ws[f"C{r}"].font=FU; ws[f"C{r}"].border=BD; ws[f"C{r}"].alignment=C
        for p in range(1,H+1):
            col=PCOLS[p]; c=ws[f"{col}{r}"]; c.border=BD; c.alignment=C; c.number_format=fmt; c.font=(FRz if green else FN)
            if formula: c.value="="+formula(col); c.fill=Fr
            else:
                if values is not None: c.value=values[p-1]
                if inp: c.fill=Fi
    def sline(r,label,unit,fmt,value=None,formula=None,green=False):
        b=ws[f"B{r}"]; b.value=label; b.border=BD; b.alignment=L; b.font=(FRz if green else FN)
        ws[f"C{r}"]=unit; ws[f"C{r}"].font=FU; ws[f"C{r}"].border=BD; ws[f"C{r}"].alignment=C
        c=ws[f"E{r}"]; c.border=BD; c.alignment=C; c.number_format=fmt; c.font=(FRz if green else FN)
        if formula: c.value="="+formula; c.fill=Fr
        else:
            if value is not None: c.value=value
            c.fill=(Fr if green else Fi)

    # ---- СВОДКА ----
    section(ws,7,"СВОДКА ПО ГОДАМ (зелёное — авто из моделей ниже; тянется в расчёт)","B","N")
    yline(8,"Простои на ТО","ч/год",M,green=True,formula=lambda col:"$G$27")
    yline(9,"Простои в ремонтах","ч/год",M,green=True,formula=lambda col:f"{col}32")
    yline(10,"Ежесменный осмотр","ч/год",M,values=[0]*H,inp=True)
    yline(11,"КТГ (авто) = (8760 − простои)/8760","коэф.",P1,green=True,
          formula=lambda col:f"({pc('kfv')}-{col}8-{col}9-{col}10)/{pc('kfv')}")
    yline(12,"ТО (запчасти + смазочные)","тыс/год",M,green=True,formula=lambda col:"$E$28")
    yline(13,"Текущие ремонты","тыс/год",M,green=True,formula=lambda col:f"{col}33")
    yline(14,"Капитальные ремонты (ППР)","тыс/год",M,green=True,formula=lambda col:f"{col}34")
    yline(15,"Сервис (ТО + ремонты)","тыс/год",M,green=True,formula=lambda col:f"$E$30+{col}35")
    yline(16,"ИТОГО ТОиР (авто) → «Данные по годам»","тыс/год",M,green=True,
          formula=lambda col:f"{col}12+{col}13+{col}14+{col}15")
    yline(17,"Удельный расход ДТ","кг/час",M1,values=FUEL_Y[i],inp=True)
    yline(18,"Расходники на ковш","руб/м³",M2,values=CONS_Y[i],inp=True)

    # ---- МОДЕЛЬ ТО ----
    section(ws,19,"РАСЧЁТ ТО (по интервалам обслуживания)","B","N")
    sline(20,"Годовая наработка","м/час",M,value=NARA)
    for k,h in enumerate(["Вид ТО","интервал","длит.простоя","персонал","ТО/год","простой,ч","ч·час"]):
        c=ws.cell(row=21,column=2+k,value=h); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
    for k in range(5):
        r=22+k
        ws.cell(row=r,column=2,value=f"ТО-{k+1}").font=FN; ws.cell(row=r,column=2).border=BD; ws.cell(row=r,column=2).alignment=L
        ws.cell(row=r,column=3,value=TO_INT[k]); ws.cell(row=r,column=4,value=TO_DUR[k]); ws.cell(row=r,column=5,value=TO_PERS[k])
        for cc in (3,4,5): ws.cell(row=r,column=cc).fill=Fi; ws.cell(row=r,column=cc).border=BD; ws.cell(row=r,column=cc).alignment=C; ws.cell(row=r,column=cc).number_format=M1
        ws.cell(row=r,column=6,value=f"=$E$20/C{r}"); ws.cell(row=r,column=7,value=f"=F{r}*D{r}"); ws.cell(row=r,column=8,value=f"=G{r}*E{r}")
        for cc in (6,7,8): ws.cell(row=r,column=cc).fill=Fr; ws.cell(row=r,column=cc).border=BD; ws.cell(row=r,column=cc).alignment=C; ws.cell(row=r,column=cc).number_format=M1; ws.cell(row=r,column=cc).font=FRz
    ws.cell(row=27,column=2,value="Итого простои / ч·часы ТО").font=FRz; ws.cell(row=27,column=2).border=BD
    ws.cell(row=27,column=7,value="=SUM(G22:G26)"); ws.cell(row=27,column=8,value="=SUM(H22:H26)")
    for cc in (7,8): ws.cell(row=27,column=cc).fill=Fr; ws.cell(row=27,column=cc).border=BD; ws.cell(row=27,column=cc).number_format=M1; ws.cell(row=27,column=cc).font=FRz
    _base=min(TOIR_Y[i]); _parts=round(_base-_to_serv,3)
    sline(28,"Затраты на запчасти ТО (из каталога)","тыс/год",M,formula=f"SUM(F{CAT0}:F{CORR})",green=True)
    sline(29,"Ставка сервиса","вал/ч·час",M,value=RATE)
    sline(30,"Сервис ТО (авто) = ч·часы × ставка","тыс/год",M,formula="$H$27*$E$29/1000",green=True)

    # ---- МОДЕЛЬ РЕМОНТОВ ----
    section(ws,31,"РАСЧЁТ РЕМОНТОВ (график по годам)","B","N")
    rem_dt=[round(max(0,KFV_LIT*(1-KTG_Y[i][p])-_to_dt),2) for p in range(H)]
    rem_cap=[round(TOIR_Y[i][p]-_base,3) for p in range(H)]
    yline(32,"Простои в ремонтах","ч/год",M,values=rem_dt,inp=True)
    yline(33,"Текущие ремонты","тыс/год",M,values=[0]*H,inp=True)
    yline(34,"Капитальные ремонты (ППР) — из списка","тыс/год",M,green=True,
          formula=lambda col:f"SUM({col}{REM0}:{col}{REMPLUG})")
    yline(35,"Сервис ремонтов (трудозатраты)","тыс/год",M,values=[0]*H,inp=True)

    # ---- ПОЛНЫЙ КАТАЛОГ ЗАПЧАСТЕЙ ТО ----
    section(ws,36,"КАТАЛОГ ЗАПЧАСТЕЙ ТО (заполняет подрядчик)","B","N")
    for k,h in enumerate(["Наименование","цена за ед","кол-во/ТО","№ ТО","стоимость,тыс/год"]):
        c=ws.cell(row=37,column=2+k,value=h); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
    for k,(nm,price,qty,tt) in enumerate(TO_PARTS):
        r=CAT0+k
        ws.cell(row=r,column=2,value=nm).font=FN; ws.cell(row=r,column=2).border=BD; ws.cell(row=r,column=2).alignment=L
        ws.cell(row=r,column=3,value=price); ws.cell(row=r,column=4,value=qty); ws.cell(row=r,column=5,value=tt)
        for cc in (3,4,5): ws.cell(row=r,column=cc).fill=Fi; ws.cell(row=r,column=cc).border=BD; ws.cell(row=r,column=cc).alignment=C; ws.cell(row=r,column=cc).number_format=M2
        f=ws.cell(row=r,column=6,value=f"=($E$20/INDEX($C$22:$C$26,E{r}))*D{r}*C{r}/1000")
        f.fill=Fr; f.border=BD; f.alignment=C; f.number_format=M2; f.font=FRz
    ws.cell(row=CORR,column=2,value="Корректировка (поглощения/прочее)").font=FN; ws.cell(row=CORR,column=2).border=BD; ws.cell(row=CORR,column=2).alignment=L
    cval=ws.cell(row=CORR,column=6,value=round(_parts-_cat_sum,3)); cval.fill=Fi; cval.border=BD; cval.alignment=C; cval.number_format=M2

    # ---- ПОЛНЫЙ СПИСОК УЗЛОВ / КАПРЕМОНТОВ ----
    section(ws,REMSEC,"СПИСОК УЗЛОВ И КАПРЕМОНТОВ (заполняет подрядчик, стоимость по годам)","B","N")
    ws.cell(row=REMHDR,column=2,value="Узел / вид ремонта").font=FHd; ws.cell(row=REMHDR,column=2).fill=Fh; ws.cell(row=REMHDR,column=2).border=BD; ws.cell(row=REMHDR,column=2).alignment=C
    ws.cell(row=REMHDR,column=3,value="нормо-час").font=FHd; ws.cell(row=REMHDR,column=3).fill=Fh; ws.cell(row=REMHDR,column=3).border=BD; ws.cell(row=REMHDR,column=3).alignment=C
    for p in range(1,H+1):
        c=ws.cell(row=REMHDR,column=4+p,value=f"год {p}"); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
    for k,(nm,nh) in enumerate(REM_ITEMS):
        r=REM0+k
        ws.cell(row=r,column=2,value=nm).font=FN; ws.cell(row=r,column=2).border=BD; ws.cell(row=r,column=2).alignment=L
        ws.cell(row=r,column=3,value=nh); ws.cell(row=r,column=3).fill=Fi; ws.cell(row=r,column=3).border=BD; ws.cell(row=r,column=3).alignment=C; ws.cell(row=r,column=3).number_format=M
        for p in range(1,H+1):
            c=ws.cell(row=r,column=4+p); c.fill=Fi; c.border=BD; c.alignment=C; c.number_format=M
    # строка-плуг (укрупнённый капремонт по годам — пример)
    ws.cell(row=REMPLUG,column=2,value="Прочие/укрупнённо (капремонт)").font=FN; ws.cell(row=REMPLUG,column=2).border=BD; ws.cell(row=REMPLUG,column=2).alignment=L
    for p in range(1,H+1):
        c=ws.cell(row=REMPLUG,column=4+p,value=rem_cap[p-1]); c.fill=Fi; c.border=BD; c.alignment=C; c.number_format=M

    tpl[i]={"sheet":TPL_NAMES[i],"ktg":11,"itogo":16,"fuel":17,"cons":18,"cur":"$D$5"}
    ws.merge_cells(f"B{NOTEROW}:N{NOTEROW}")
    nt=ws[f"B{NOTEROW}"]; nt.value=("Заполняйте: каталог запчастей ТО (цена, кол-во, № ТО) и список узлов/капремонтов (стоимость по годам). "
     "КТГ, затраты ТО и капремонты пересчитаются снизу вверх; сводка «Данные по годам» и весь расчёт обновятся автоматически. "
     "Строка «Корректировка» и «Прочие/укрупнённо» — балансирующие, замените детализацией.")
    nt.font=Font(size=9,italic=True,color="808080"); nt.alignment=L
print("Листы-шаблоны с полными списками ТО/Ремонтов:",NEXC)

# --- СВОДКА «Данные по годам» (тянет из листов-шаблонов) ---
wy=wb.create_sheet(S_PYR); wy.sheet_view.showGridLines=False
title(wy,"ПОГОДОВЫЕ ДАННЫЕ ЭКСПЛУАТАЦИИ (СВОДКА)","Автоматически из листов «Поставщик 1…6». Правьте данные в тех листах, здесь — только просмотр.",last="N")
wy.column_dimensions["A"].width=2; wy.column_dimensions["B"].width=34; wy.column_dimensions["C"].width=11; wy.column_dimensions["D"].width=9
for p in range(1,H+1): wy.column_dimensions[PCOLS[p]].width=10
wy.freeze_panes="E6"
wy.merge_cells("B4:N4")
ins=wy["B4"]; ins.value=("Эти значения ТЯНУТСЯ из листов-шаблонов «Поставщик 1…6» (КТГ и ИТОГО ТОиР — авто). "
 "Чтобы изменить данные по машине — откройте её лист-шаблон и правьте жёлтые ячейки; сводка и расчёт обновятся сами.")
ins.font=Font(size=9,italic=True,color=ACC); ins.alignment=L; ins.fill=Fa
YHR=5
wy[f"B{YHR}"]="Показатель"; wy[f"C{YHR}"]="Ед."; wy[f"D{YHR}"]="Валюта"
for cc in ("B","C","D"):
    wy[f"{cc}{YHR}"].font=FHd; wy[f"{cc}{YHR}"].fill=Fh; wy[f"{cc}{YHR}"].border=BD; wy[f"{cc}{YHR}"].alignment=C
for p in range(1,H+1):
    c=wy[f"{PCOLS[p]}{YHR}"]; c.value=f"год {p}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
pyr_ktg={}; pyr_toir={}; pyr_cur={}; pyr_fuel={}; pyr_cons={}
yr=YHR+1
def yrow_pull(store,i,label,unit,tkey,fmt,cur=False):
    global yr
    store[i]=yr
    wy[f"B{yr}"]=label; wy[f"B{yr}"].font=FN; wy[f"B{yr}"].border=BD; wy[f"B{yr}"].alignment=L
    wy[f"C{yr}"]=unit; wy[f"C{yr}"].font=FU; wy[f"C{yr}"].border=BD; wy[f"C{yr}"].alignment=C
    if cur:
        pyr_cur[i]=yr; cu=wy[f"D{yr}"]; cu.value=f"='{tpl[i]['sheet']}'!{tpl[i]['cur']}"
        cu.fill=Fr; cu.border=BD; cu.alignment=C; cu.font=FB
    for p in range(1,H+1):
        c=wy[f"{PCOLS[p]}{yr}"]; c.value=f"='{tpl[i]['sheet']}'!{PCOLS[p]}{tpl[i][tkey]}"
        c.fill=Fr; c.border=BD; c.number_format=fmt; c.alignment=C; c.font=FN
    yr+=1
for i in range(NEXC):
    wy.merge_cells(f"B{yr}:N{yr}")
    hc=wy[f"B{yr}"]; hc.value=f"=\"Вариант {i+1}:  \"&'{S_IN}'!{EXC[i]}{irow['name']}"; hc.font=FSec; hc.alignment=L
    for col in range(2,15): wy.cell(row=yr,column=col).fill=Fm
    yr+=1
    yrow_pull(pyr_ktg,i,"КТГ (техготовность)","коэф.","ktg",P1)
    yrow_pull(pyr_toir,i,"Затраты на ТОиР и сервис","тыс.вал/г","itogo",M,cur=True)
    yrow_pull(pyr_fuel,i,"Удельный расход ДТ","кг/час","fuel",M1)
    yrow_pull(pyr_cons,i,"Расходники на ковш","руб/м³","cons",M2)
_PYR_ROW={"ktg":pyr_ktg,"toir":pyr_toir,"fuel":pyr_fuel,"cons":pyr_cons}
def pyr(kind,i,p): return f"'{S_PYR}'!{PCOLS[p]}{_PYR_ROW[kind][i]}"
def pyr_cur_ref(i): return f"'{S_PYR}'!$D${pyr_cur[i]}"
def toir_rate(i):
    cur=pyr_cur_ref(i)
    return f'IF({cur}="CNY",{pc("cny")},IF({cur}="USD",{pc("usd")},IF({cur}="EUR",{pc("eur")},1)))'
print("Данные по годам (сводка) готова")

# ================================================================= #
# ПРОИЗВОДИТЕЛЬНОСТЬ (показатели 1-го года)
# ================================================================= #
wpr=wb.create_sheet(S_PROD); wpr.sheet_view.showGridLines=False
title(wpr,"РАСЧЁТ ПРОИЗВОДИТЕЛЬНОСТИ (1-й год)","Часовая и годовая производительность. По годам — на листе «Денежный поток».")
wpr.column_dimensions["A"].width=2; wpr.column_dimensions["B"].width=42; wpr.column_dimensions["C"].width=12
for cc in EXC: wpr.column_dimensions[cc].width=15
PHR=4
wpr[f"B{PHR}"]="Показатель"; wpr[f"B{PHR}"].font=FHd; wpr[f"B{PHR}"].fill=Fh; wpr[f"B{PHR}"].border=BD; wpr[f"B{PHR}"].alignment=C
wpr[f"C{PHR}"]="Ед."; wpr[f"C{PHR}"].font=FHd; wpr[f"C{PHR}"].fill=Fh; wpr[f"C{PHR}"].border=BD; wpr[f"C{PHR}"].alignment=C
for i,cc in enumerate(EXC):
    c=wpr[f"{cc}{PHR}"]; c.value=f"='{S_IN}'!{cc}{irow['name']}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
prd={}; pr=PHR+1
def padd2(key,label,unit,fn,fmt=M1,res=False,bold=False):
    global pr
    prd[key]=pr
    b=wpr[f"B{pr}"]; b.value=label; b.border=BD; b.alignment=L; b.font=(FRz if res else (FB if bold else FN))
    wpr[f"C{pr}"]=unit; wpr[f"C{pr}"].font=FU; wpr[f"C{pr}"].border=BD; wpr[f"C{pr}"].alignment=C
    for cc in EXC:
        c=wpr[f"{cc}{pr}"]; c.value="="+fn(cc); c.number_format=fmt; c.border=BD; c.alignment=Rr
        c.font=(FRz if res else FN)
        if res: c.fill=Fr
    pr+=1
def pref(key,col): return f"{col}{prd[key]}"
def cur_rate(col):
    cur=inp('price_cur',col)
    return f'IF({cur}="CNY",{pc("cny")},IF({cur}="USD",{pc("usd")},IF({cur}="EUR",{pc("eur")},1)))'
padd2("price_rub","Стоимость в рублях","тыс.руб",lambda c:f"{inp('price_val',c)}*{cur_rate(c)}+{inp('install',c)}",M,bold=True)
padd2("rock","Порода в ковше","т",lambda c:f"{inp('bucket',c)}*{pc('dens')}*({pc('fill')}/{pc('loosen')})*{pc('loss')}",M2)
padd2("buckets","Ковшей на самосвал (циклов)","шт",lambda c:f"ROUNDUP({pc('cap')}/{pref('rock',c)},0)",M)
padd2("trucks","Самосвалов за час работы","а/с/ч",lambda c:f"3600/({pref('buckets',c)}*{inp('cycle',c)}+{pc('pos')})",M2)
padd2("th","Часовая производительность","т/час",lambda c:f"{pc('cap')}*{pref('trucks',c)}",M)
padd2("m3h","Часовая производительность","м³/час",lambda c:f"{pref('th',c)}/{pc('dens')}",M1)
padd2("tech","Время техготовности (1-й год)","час/год",lambda c:f"{pc('kfv')}*{pyr('ktg',EXC.index(c),1)}",M)
padd2("eff","Эффективное время (1-й год)","час/год",lambda c:f"{pref('tech',c)}*(1-{inp('downtime',c)}/(60*{pc('shift')}))",M)
padd2("kio","КИО (1-й год)","коэф.",lambda c:f"{pref('eff',c)}/{pc('kfv')}",P1)
padd2("am3","Производительность (1-й год)","тыс.м³",lambda c:f"{pref('m3h',c)}*{pref('eff',c)}/1000",M,res=True)
def prodref(key,col): return f"'{S_PROD}'!{col}{prd[key]}"
print("Производительность:",len(prd))

# ================================================================= #
# ДЕНЕЖНЫЙ ПОТОК (погодовой движок, 6 блоков)
# ================================================================= #
wf=wb.create_sheet(S_CF); wf.sheet_view.showGridLines=False
title(wf,"ПОГОДОВОЙ ДЕНЕЖНЫЙ ПОТОК И АННУИТЕТ","Пост-налоговый жизненный цикл по годам: период 0 — инвестиции, 1–10 — эксплуатация.",last="N")
wf.column_dimensions["A"].width=2; wf.column_dimensions["B"].width=40; wf.column_dimensions["C"].width=11
for p,col in enumerate(PCOLS): wf.column_dimensions[col].width=11
wf.freeze_panes="D7"
# общие строки: период (4), год (5), дисконт-фактор (6)
wf["B4"]="Период"; wf["B4"].font=FHd; wf["B4"].alignment=L
wf["B5"]="Календарный год"; wf["B5"].font=FU; wf["B5"].alignment=L
wf["B6"]="Дисконт-фактор"; wf["B6"].font=FHd; wf["B6"].alignment=L
for p,col in enumerate(PCOLS):
    a=wf[f"{col}4"]; a.value=p; a.font=FB; a.alignment=C; a.fill=Fh; a.border=BD
    y=wf[f"{col}5"]; y.value=2026+p; y.font=FU; y.alignment=C; y.border=BD
    d=wf[f"{col}6"]; d.value=f"=1/(1+{pc('disc')})^{col}$4"; d.number_format=NUM; d.alignment=C; d.border=BD; d.font=FN
DFROW=6; PERROW=4

# строки блока (ключ, метка, ед., генератор формулы(col,exc_letter,base_row_map), только период1..10?)
CF_ROWS=[
 ("ktg","КТГ","коэф.",P1,True),
 ("eff","Эффективное время","час/год",M,True),
 ("prod","Производительность","тыс.м³",M,True),
 ("fuel","Дизтопливо","тыс.руб",M,True),
 ("maint","ТОиР и сервис (по годам)","тыс.руб",M,True),
 ("bucket","Расходники на ковш","тыс.руб",M,True),
 ("pers","Персонал","тыс.руб",M,True),
 ("cash","Итого денежные затраты","тыс.руб",M,True),
 ("dep","Амортизация (НУ)","тыс.руб",M,True),
 ("base","Налогооблагаемая база","тыс.руб",M,True),
 ("tax","Налог на прибыль (щит)","тыс.руб",M,True),
 ("ocf","Операц. денежный поток","тыс.руб",M,True),
 ("invr","Инвестиции / остаточная ст-ть","тыс.руб",M,False),
 ("cf","Чистый денежный поток","тыс.руб",M,False),
 ("dcf","Дисконтированный поток","тыс.руб",M,False),
 ("cdcf","Кумул. дисконт. поток","тыс.руб",M,False),
 ("ann_y","Аннуитет по году","руб/м³",M2,True),
]
BLOCK_H=len(CF_ROWS)+2      # +заголовок +пустая
cf={}                       # cf[exc_index][key]=row
first_block=8
for i in range(NEXC):
    bs=first_block+i*BLOCK_H
    ec=EXC[i]
    # заголовок блока
    wf.merge_cells(f"B{bs}:N{bs}")
    hc=wf[f"B{bs}"]; hc.value=f"=\"Вариант {i+1}:  \"&'{S_IN}'!{ec}{irow['name']}"; hc.font=FSec; hc.alignment=L
    for col in range(2,15): wf.cell(row=bs,column=col).fill=Fm
    rowmap={}
    for j,(key,label,unit,fmt,op) in enumerate(CF_ROWS):
        rr=bs+1+j; rowmap[key]=rr
        b=wf[f"B{rr}"]; b.value=label; b.font=(FRz if key=="ann_y" else FN); b.border=BD; b.alignment=L
        wf[f"C{rr}"]=unit; wf[f"C{rr}"].font=FU; wf[f"C{rr}"].border=BD; wf[f"C{rr}"].alignment=C
    cf[i]=rowmap
    # входные ссылки
    def I(k): return f"'{S_IN}'!{ec}{irow[k]}"
    PR=prodref('price_rub',ec); M3H=prodref('m3h',ec)
    for p,col in enumerate(PCOLS):
        per=f"{col}${PERROW}"
        def put(key,formula,fmt=M,op_only=True):
            rr=rowmap[key]; cell=wf[f"{col}{rr}"]
            if op_only and p==0:
                cell.value=None
            else:
                cell.value="="+formula
            cell.number_format=fmt; cell.border=BD; cell.alignment=Rr; cell.font=(FRz if key=="ann_y" else FN)
            if key in ("cash","ocf","cf","ann_y") and cell.value: cell.font=(FRz if key=="ann_y" else FB)
        # операционные (период 1..10)
        put("ktg", (f"{pyr('ktg',i,p)}" if p>0 else "0"), P1)
        put("eff", f"{pc('kfv')}*{col}{rowmap['ktg']}*(1-{I('downtime')}/(60*{pc('shift')}))", M)
        put("prod", f"{M3H}*{col}{rowmap['eff']}/1000", M)
        put("fuel", f"{col}{rowmap['eff']}*{pyr('fuel',i,p)}*{pc('fuel_p')}*(1+{pc('esc_fuel')})^({per}-1)/1000", M)
        put("maint", (f"{pyr('toir',i,p)}*{toir_rate(i)}" if p>0 else "0"), M)
        put("bucket", f"{pyr('cons',i,p)}*{col}{rowmap['prod']}*(1+{pc('esc_parts')})^({per}-1)", M)
        put("pers", f"{I('personnel')}*(1+{pc('esc_lab')})^({per}-1)", M)
        put("cash", f"{col}{rowmap['fuel']}+{col}{rowmap['maint']}+{col}{rowmap['bucket']}+{col}{rowmap['pers']}", M)
        put("dep", f"IF({per}<=INT({pc('dep_years')}),{PR}/{pc('dep_years')},IF({per}=INT({pc('dep_years')})+1,{PR}*({pc('dep_years')}-INT({pc('dep_years')}))/{pc('dep_years')},0))", M)
        put("base", f"{col}{rowmap['cash']}+{col}{rowmap['dep']}", M)
        put("tax", f"-{col}{rowmap['base']}*{pc('tax')}", M)
        put("ocf", f"{col}{rowmap['base']}+{col}{rowmap['tax']}-{col}{rowmap['dep']}", M)
        # инвестиции (период 0) / остаточная стоимость (период N)
        cell=wf[f"{col}{rowmap['invr']}"]
        cell.value=f"=IF({per}=0,{PR},IF({per}={pc('horizon')},-{I('resid')}*{PR},0))"
        cell.number_format=M; cell.border=BD; cell.alignment=Rr; cell.font=FN
        # чистый поток
        ocf_ref=f"{col}{rowmap['ocf']}" if p>0 else "0"
        cfc=wf[f"{col}{rowmap['cf']}"]
        cfc.value=f"={ocf_ref}+{col}{rowmap['invr']}"
        cfc.number_format=M; cfc.border=BD; cfc.alignment=Rr; cfc.font=FB
        # дисконтированный
        dcfc=wf[f"{col}{rowmap['dcf']}"]
        dcfc.value=f"={col}{rowmap['cf']}*{col}${DFROW}"
        dcfc.number_format=M; dcfc.border=BD; dcfc.alignment=Rr; dcfc.font=FN
        # кумулятивный дисконтированный
        cdc=wf[f"{col}{rowmap['cdcf']}"]
        cdc.value=f"=SUM($D{rowmap['dcf']}:{col}{rowmap['dcf']})"
        cdc.number_format=M; cdc.border=BD; cdc.alignment=Rr; cdc.font=FN
        # аннуитет по году (период1..10)
        ayc=wf[f"{col}{rowmap['ann_y']}"]
        if p==0: ayc.value=None
        else:
            ayc.value=(f"=SUM($D{rowmap['dcf']}:{col}{rowmap['dcf']})/SUM($E${DFROW}:{col}${DFROW})"
                       f"/(SUM($E{rowmap['prod']}:{col}{rowmap['prod']})/{per})")
        ayc.number_format=M2; ayc.border=BD; ayc.alignment=Rr; ayc.font=FRz
print("Денежный поток: блоков",NEXC,"высота блока",BLOCK_H)

def cfrng(i,key,p0=1,p1=H):
    c0=PCOLS[p0]; c1=PCOLS[p1]; return f"'{S_CF}'!{c0}{cf[i][key]}:{c1}{cf[i][key]}"
def cfcell(i,key,p): return f"'{S_CF}'!{PCOLS[p]}{cf[i][key]}"
DF_RNG=f"'{S_CF}'!$E${DFROW}:$N${DFROW}"

# ================================================================= #
# РАСЧЁТ АННУИТЕТА (свод из погодового движка)
# ================================================================= #
wa=wb.create_sheet(S_ANN); wa.sheet_view.showGridLines=False
title(wa,"РАСЧЁТ АННУИТЕТА (СВОД)","Приведённые NPV затрат из погодового движка → аннуитет руб/м³ и руб/т.")
wa.column_dimensions["A"].width=2; wa.column_dimensions["B"].width=44; wa.column_dimensions["C"].width=12
for cc in EXC: wa.column_dimensions[cc].width=15
AHR=4
wa[f"B{AHR}"]="Статья"; wa[f"B{AHR}"].font=FHd; wa[f"B{AHR}"].fill=Fh; wa[f"B{AHR}"].border=BD; wa[f"B{AHR}"].alignment=C
wa[f"C{AHR}"]="Ед."; wa[f"C{AHR}"].font=FHd; wa[f"C{AHR}"].fill=Fh; wa[f"C{AHR}"].border=BD; wa[f"C{AHR}"].alignment=C
for i,cc in enumerate(EXC):
    c=wa[f"{cc}{AHR}"]; c.value=f"='{S_IN}'!{cc}{irow['name']}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
ar={}; a=AHR+1
def asec(t):
    global a; section(wa,a,t,"B","I"); a+=1
def aadd(key,label,unit,fni,fmt=M,res=False,bold=False):
    global a; ar[key]=a
    b=wa[f"B{a}"]; b.value=label; b.border=BD; b.alignment=L; b.font=(FRz if res else (FB if bold else FN))
    wa[f"C{a}"]=unit; wa[f"C{a}"].font=FU; wa[f"C{a}"].border=BD; wa[f"C{a}"].alignment=C
    for i,cc in enumerate(EXC):
        c=wa[f"{cc}{a}"]; c.value="="+fni(i); c.number_format=fmt; c.border=BD; c.alignment=Rr
        c.font=(FRz if res else (FB if bold else FN))
        if res: c.fill=Fr
    a+=1
def aref(key,col): return f"{col}{ar[key]}"
def acol(i): return EXC[i]
def avgprod(i): return f"AVERAGE({cfrng(i,'prod')})"
def npv(i,key): return f"SUMPRODUCT({cfrng(i,key)},{DF_RNG})"
DF10=f"'{S_CF}'!$N${DFROW}"

asec("ПРИВЕДЁННЫЕ ЗАТРАТЫ ЗА ЖИЗНЕННЫЙ ЦИКЛ (NPV, тыс.руб)")
aadd("n_fuel","NPV дизтоплива",       "тыс.руб",lambda i:npv(i,'fuel'))
aadd("n_maint","NPV ТОиР (по годам, вкл. капремонты)","тыс.руб",lambda i:npv(i,'maint'))
aadd("n_bucket","NPV расходников на ковш","тыс.руб",lambda i:npv(i,'bucket'))
aadd("n_pers","NPV персонала",        "тыс.руб",lambda i:npv(i,'pers'))
aadd("n_tax","NPV налога (щит)",      "тыс.руб",lambda i:npv(i,'tax'))
aadd("n_inv","NPV инвестиций (за вычетом остат.)","тыс.руб",
     lambda i:f"{prodref('price_rub',acol(i))}*(1-{inp('resid',acol(i))}*{DF10})")
aadd("n_op","NPV операционных затрат","тыс.руб",
     lambda i:f"{aref('n_fuel',acol(i))}+{aref('n_maint',acol(i))}+{aref('n_bucket',acol(i))}+{aref('n_pers',acol(i))}+{aref('n_tax',acol(i))}",bold=True)
aadd("n_tot","NPV полной стоимости владения","тыс.руб",
     lambda i:f"{aref('n_inv',acol(i))}+{aref('n_op',acol(i))}",res=True)
aadd("avgprod","Средний годовой объём","тыс.м³/год",lambda i:avgprod(i))

asec("УДЕЛЬНЫЙ АННУИТЕТ (руб/м³) — ГЛАВНЫЙ КРИТЕРИЙ")
def annm3(nkey): return lambda i:f"{aref(nkey,acol(i))}/{pc('S')}/{aref('avgprod',acol(i))}"
aadd("u_inv","Инвестиционный","руб/м³",annm3('n_inv'),M2)
aadd("u_fuel","Дизтопливо","руб/м³",annm3('n_fuel'),M2)
aadd("u_maint","ТОиР + капремонты","руб/м³",annm3('n_maint'),M2)
aadd("u_bucket","Расходники на ковш","руб/м³",annm3('n_bucket'),M2)
aadd("u_pers","Персонал","руб/м³",annm3('n_pers'),M2)
aadd("u_tax","Налог на прибыль (щит)","руб/м³",annm3('n_tax'),M2)
aadd("u_op","Операционный аннуитет","руб/м³",annm3('n_op'),M2,bold=True)
aadd("u_tot","ОБЩИЙ АННУИТЕТ","руб/м³",annm3('n_tot'),M2,res=True)

asec("ДОПОЛНИТЕЛЬНЫЕ ПОКАЗАТЕЛИ")
aadd("t_tot","Удельный аннуитет на тонну","руб/т",
     lambda i:f"{aref('u_tot',acol(i))}/{pc('dens')}",M2,bold=True)
aadd("sebest","Себестоимость (без дисконта)","руб/м³",
     lambda i:f"AVERAGE({cfrng(i,'cash')})/{aref('avgprod',acol(i))}",M2)
aadd("cap_share","Доля инвестиций в аннуитете","%",
     lambda i:f"{aref('u_inv',acol(i))}/{aref('u_tot',acol(i))}",P1)
def annref(key,col): return f"'{S_ANN}'!{col}{ar[key]}"
print("Аннуитет:",len(ar))

# ================================================================= #
# СРАВНЕНИЕ
# ================================================================= #
wm=wb.create_sheet(S_CMP); wm.sheet_view.showGridLines=False
title(wm,"СРАВНИТЕЛЬНЫЙ АНАЛИЗ ЭКСКАВАТОРОВ","Ранжирование по удельному аннуитету (руб/м³). Зелёным — лучший вариант.")
wm.column_dimensions["A"].width=2; wm.column_dimensions["B"].width=40; wm.column_dimensions["C"].width=11
for cc in EXC: wm.column_dimensions[cc].width=15
MHR=4
wm[f"B{MHR}"]="Показатель"; wm[f"B{MHR}"].font=FHd; wm[f"B{MHR}"].fill=Fh; wm[f"B{MHR}"].border=BD; wm[f"B{MHR}"].alignment=C
wm[f"C{MHR}"]="Ед."; wm[f"C{MHR}"].font=FHd; wm[f"C{MHR}"].fill=Fh; wm[f"C{MHR}"].border=BD; wm[f"C{MHR}"].alignment=C
for i,cc in enumerate(EXC):
    c=wm[f"{cc}{MHR}"]; c.value=f"='{S_IN}'!{cc}{irow['name']}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
metrics=[
 ("Стоимость (в рублях)","price_rub",prodref,M,"тыс.руб"),
 ("Средний годовой объём","avgprod",annref,M,"тыс.м³"),
 ("КИО (1-й год)","kio",prodref,P1,"коэф."),
 ("NPV стоимости владения","n_tot",annref,M,"тыс.руб"),
 ("Себестоимость (без дисконта)","sebest",annref,M2,"руб/м³"),
 ("Инвестиционный аннуитет","u_inv",annref,M2,"руб/м³"),
 ("Операционный аннуитет","u_op",annref,M2,"руб/м³"),
 ("ОБЩИЙ АННУИТЕТ","u_tot",annref,M2,"руб/м³"),
 ("Аннуитет на тонну","t_tot",annref,M2,"руб/т"),
]
mrow={}; m=MHR+1
for label,key,reff,fmt,unit in metrics:
    mrow[key]=m
    b=wm[f"B{m}"]; b.value=label; b.border=BD; b.alignment=L; b.font=(FB if key=="u_tot" else FN)
    wm[f"C{m}"]=unit; wm[f"C{m}"].font=FU; wm[f"C{m}"].border=BD; wm[f"C{m}"].alignment=C
    for cc in EXC:
        c=wm[f"{cc}{m}"]; c.value="="+reff(key,cc); c.number_format=fmt; c.border=BD; c.alignment=Rr
        c.font=(FB if key=="u_tot" else FN)
        if key=="u_tot": c.fill=Fr
    m+=1
section(wm,m,"РАНЖИРОВАНИЕ (критерий: минимум аннуитета руб/м³)","B","I"); m+=1
tot_rng=f"D{mrow['u_tot']}:I{mrow['u_tot']}"
wm[f"B{m}"]="Место в рейтинге"; wm[f"B{m}"].font=FB; wm[f"B{m}"].border=BD
for cc in EXC:
    c=wm[f"{cc}{m}"]; c.value=f"=RANK({cc}{mrow['u_tot']},{tot_rng},1)"; c.font=FB; c.border=BD; c.alignment=C
rank_row=m; m+=1
wm[f"B{m}"]="Отклонение от лучшего"; wm[f"B{m}"].font=FN; wm[f"B{m}"].border=BD
for cc in EXC:
    c=wm[f"{cc}{m}"]; c.value=f"={cc}{mrow['u_tot']}/MIN({tot_rng})-1"; c.number_format=P1; c.border=BD; c.alignment=C; c.font=FN
m+=2
names_rng=f"D{MHR}:I{MHR}"
wm[f"B{m}"]="РЕКОМЕНДУЕМЫЙ ВАРИАНТ:"; wm[f"B{m}"].font=Font(bold=True,size=12,color=DARK)
wm.merge_cells(f"C{m}:I{m}")
rec=wm[f"C{m}"]; rec.value=f'=INDEX({names_rng},MATCH(MIN({tot_rng}),{tot_rng},0))&"  ("&TEXT(MIN({tot_rng}),"0.00")&" руб/м³)"'
rec.font=Font(bold=True,size=12,color="006100"); rec.alignment=C
for col in ["C","D","E","F","G","H","I"]: wm[f"{col}{m}"].fill=Fb; wm[f"{col}{m}"].border=BD
m+=2
# структура для диаграммы
section(wm,m,"СТРУКТУРА УДЕЛЬНОГО АННУИТЕТА (руб/м³)","B","I"); m+=1
wm[f"B{m}"]="Статья"; wm[f"B{m}"].font=FHd; wm[f"B{m}"].fill=Fh; wm[f"B{m}"].border=BD
for cc in EXC:
    c=wm[f"{cc}{m}"]; c.value=f"='{S_IN}'!{cc}{irow['name']}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
stop=m+1
struct=[("Инвестиционный","u_inv"),("Дизтопливо","u_fuel"),("ТОиР + капремонты","u_maint"),
        ("Расходники на ковш","u_bucket"),("Персонал","u_pers"),("Налог (щит)","u_tax")]
m=stop
for label,key in struct:
    b=wm[f"B{m}"]; b.value=label; b.font=FN; b.border=BD; b.alignment=L
    for cc in EXC:
        c=wm[f"{cc}{m}"]; c.value="="+annref(key,cc); c.number_format=M2; c.border=BD; c.alignment=Rr; c.font=FN
    m+=1
sbot=m-1
catref=Reference(wm,min_col=4,min_row=stop-1,max_col=9,max_row=stop-1)
chart1=BarChart(); chart1.type="col"; chart1.grouping="stacked"; chart1.overlap=100
chart1.title="Структура аннуитета, руб/м³"; chart1.height=9; chart1.width=20; chart1.y_axis.title="руб/м³"
for k in range(len(struct)):
    ri=stop+k; vals=Reference(wm,min_col=4,min_row=ri,max_col=9,max_row=ri)
    ser=Series(vals,title_from_data=False); ser.tx=SeriesLabel(strRef=StrRef(f"'{S_CMP}'!$B${ri}")); chart1.series.append(ser)
chart1.set_categories(catref); wm.add_chart(chart1,"K4")
chart2=BarChart(); chart2.type="col"; chart2.title="Общий аннуитет, руб/м³"; chart2.height=8; chart2.width=11; chart2.y_axis.title="руб/м³"
d2=Reference(wm,min_col=4,min_row=mrow['u_tot'],max_col=9,max_row=mrow['u_tot'])
c2=Reference(wm,min_col=4,min_row=MHR,max_col=9,max_row=MHR)
chart2.add_data(d2,from_rows=True); chart2.set_categories(c2); chart2.legend=None
chart2.dataLabels=DataLabelList(); chart2.dataLabels.showVal=True; chart2.dataLabels.numFmt="0.0"
wm.add_chart(chart2,"K23")
chart3=BarChart(); chart3.type="col"; chart3.title="Стоимость экскаватора, тыс.руб"; chart3.height=8; chart3.width=11
d3=Reference(wm,min_col=4,min_row=mrow['price_rub'],max_col=9,max_row=mrow['price_rub'])
chart3.add_data(d3,from_rows=True); chart3.set_categories(c2); chart3.legend=None
wm.add_chart(chart3,"T23")
print("Сравнение:",len(mrow))

# ================================================================= #
# ДАШБОРД (исполнительный)
# ================================================================= #
wd=wb.create_sheet(S_DASH); wd.sheet_view.showGridLines=False
title(wd,"ИТОГОВЫЙ ДАШБОРД — СРАВНЕНИЕ ЭКСКАВАТОРОВ","Ключевые показатели и графики. Данные обновляются автоматически.",last="N")
for col,w in (("A",2),("B",22),("C",22),("D",22),("E",22),("F",6),("G",13),("H",13),("I",15),("J",13)):
    wd.column_dimensions[col].width=w
# KPI-плитки
utot="'{s}'!{r}".format(s=S_ANN,r="") # not used
UT=f"'{S_ANN}'!D{ar['u_tot']}:I{ar['u_tot']}"
NM=f"'{S_ANN}'!D{ar['n_tot']}:I{ar['n_tot']}"
NR=f"'{S_IN}'!D{irow['name']}:I{irow['name']}"
def tile(anchor_col,row,caption,value_formula,fmt):
    c0=anchor_col
    # заголовок
    hc=wd[f"{c0}{row}"]; hc.value=caption; hc.font=Font(size=9,bold=True,color="FFFFFF"); hc.fill=Ftile; hc.alignment=C
    wd.merge_cells(f"{c0}{row}:{c0}{row}")
    vc=wd[f"{c0}{row+1}"]; vc.value="="+value_formula; vc.font=Font(size=14,bold=True,color=DARK); vc.alignment=C
    vc.number_format=fmt; vc.fill=Fb
    for rr in (row,row+1):
        wd[f"{c0}{rr}"].border=BD
tr=4
section(wd,tr,"КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ","B","J"); tr+=1
# плитки в столбцах B,C,D,E
tile("B",tr,"Лучший вариант",f'INDEX({NR},MATCH(MIN({UT}),{UT},0))',"General")
tile("C",tr,"Мин. аннуитет, руб/м³",f"MIN({UT})",M2)
tile("D",tr,"Мин. NPV владения, тыс.руб",f"MIN({NM})",M)
tile("E",tr,"Разброс аннуитета, руб/м³",f"MAX({UT})-MIN({UT})",M2)
tr+=3
# мини-рейтинг
section(wd,tr,"РЕЙТИНГ ПО УДЕЛЬНОМУ АННУИТЕТУ (руб/м³)","B","J"); tr+=1
wd[f"B{tr}"]="Место"; wd[f"C{tr}"]="Экскаватор"; wd[f"G{tr}"]="Аннуитет, руб/м³"; wd[f"I{tr}"]="руб/т"
for cc,txt in (("B","Место"),("C","Экскаватор"),("G","Аннуитет, руб/м³"),("I","руб/т")):
    wd[f"{cc}{tr}"].font=FHd; wd[f"{cc}{tr}"].fill=Fh; wd[f"{cc}{tr}"].border=BD; wd[f"{cc}{tr}"].alignment=C
wd.merge_cells(f"C{tr}:F{tr}"); wd.merge_cells(f"G{tr}:H{tr}")
TT=f"'{S_ANN}'!$D${ar['t_tot']}:$I${ar['t_tot']}"
for k in range(NEXC):
    rr=tr+1+k
    wd[f"B{rr}"]=k+1; wd[f"B{rr}"].font=FB; wd[f"B{rr}"].border=BD; wd[f"B{rr}"].alignment=C
    wd.merge_cells(f"C{rr}:F{rr}")
    nm=wd[f"C{rr}"]; nm.value=f'=INDEX({NR},MATCH(SMALL({UT},{k+1}),{UT},0))'; nm.border=BD; nm.alignment=L
    nm.font=(Font(bold=True,color="006100") if k==0 else FN)
    wd.merge_cells(f"G{rr}:H{rr}")
    va=wd[f"G{rr}"]; va.value=f"=SMALL({UT},{k+1})"; va.number_format=M2; va.border=BD; va.alignment=C
    va.font=(FB if k==0 else FN)
    if k==0: nm.fill=Fb; va.fill=Fb
    vt=wd[f"I{rr}"]; vt.value=f'=INDEX({TT},MATCH(SMALL({UT},{k+1}),{UT},0))'; vt.number_format=M2; vt.border=BD; vt.alignment=C; vt.font=FN
# график: общий аннуитет
gc=BarChart(); gc.type="col"; gc.title="Общий аннуитет, руб/м³"; gc.height=8; gc.width=13; gc.y_axis.title="руб/м³"
gd=Reference(wm,min_col=4,min_row=mrow['u_tot'],max_col=9,max_row=mrow['u_tot'])
gcat=Reference(wm,min_col=4,min_row=MHR,max_col=9,max_row=MHR)
gc.add_data(gd,from_rows=True); gc.set_categories(gcat); gc.legend=None
gc.dataLabels=DataLabelList(); gc.dataLabels.showVal=True; gc.dataLabels.numFmt="0.0"
wd.add_chart(gc,"L4")
# график: траектория аннуитета по годам (line)
lc=LineChart(); lc.title="Траектория аннуитета по годам, руб/м³"; lc.height=8; lc.width=16
lc.x_axis.title="Год эксплуатации"; lc.y_axis.title="руб/м³"
per_cat=Reference(wf,min_col=5,min_row=PERROW,max_col=14,max_row=PERROW)   # периоды 1..10
for i in range(NEXC):
    rr=cf[i]['ann_y']; vals=Reference(wf,min_col=5,min_row=rr,max_col=14,max_row=rr)
    ser=Series(vals,title_from_data=False); ser.tx=SeriesLabel(strRef=StrRef(f"'{S_IN}'!{EXC[i]}${irow['name']}")); lc.series.append(ser)
lc.set_categories(per_cat)
wd.add_chart(lc,"L20")
print("Дашборд готов")

# ================================================================= #
# ЧУВСТВИТЕЛЬНОСТЬ
# ================================================================= #
ws=wb.create_sheet(S_SENS); ws.sheet_view.showGridLines=False
title(ws,"АНАЛИЗ ЧУВСТВИТЕЛЬНОСТИ","Торнадо драйверов, влияние КТГ и ставки дисконтирования на аннуитет.")
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=34
for cc in "CDEFGH": ws.column_dimensions[cc].width=15
ws["B4"]="Анализируемый экскаватор (1–6):"; ws["B4"].font=FB
sel=ws["C4"]; sel.value=1; sel.fill=Fi; sel.border=BD; sel.alignment=C; sel.font=FB
SEL="$C$4"
dv2=DataValidation(type="whole",operator="between",formula1="1",formula2="6"); ws.add_data_validation(dv2); dv2.add(sel)
ws["D4"]=f"=INDEX('{S_IN}'!D{irow['name']}:I{irow['name']},{SEL})"; ws["D4"].font=Font(bold=True,color=ACC); ws["D4"].alignment=L
ws.merge_cells("D4:H4")
def selann(key): return f"INDEX('{S_ANN}'!D{ar[key]}:I{ar[key]},{SEL})"
section(ws,6,"БАЗОВЫЕ ВЕЛИЧИНЫ ВЫБРАННОГО ВАРИАНТА","B","H")
base=[
 ("b_tot","Общий аннуитет","руб/м³",selann("u_tot"),M2),
 ("b_inv","— инвестиционный","руб/м³",selann("u_inv"),M2),
 ("b_fuel","— дизтопливо","руб/м³",selann("u_fuel"),M2),
 ("b_maint","— ТОиР + капремонты","руб/м³",selann("u_maint"),M2),
 ("b_pers","— персонал","руб/м³",selann("u_pers"),M2),
 ("b_am3","Средний годовой объём","тыс.м³",selann("avgprod"),M),
 ("b_price","Стоимость (руб)","тыс.руб",f"INDEX('{S_PROD}'!D{prd['price_rub']}:I{prd['price_rub']},{SEL})",M),
 ("b_fix","Постоянные (инв+ТОиР+перс+щит)","руб/м³",
   f"{selann('u_inv')}+{selann('u_maint')}+{selann('u_pers')}+{selann('u_tax')}",M2),
 ("b_var","Переменные (топливо+ковш)","руб/м³",f"{selann('u_fuel')}+{selann('u_bucket')}",M2),
]
brow={}; b=7
for key,label,unit,fml,fmt in base:
    brow[key]=b
    ws[f"B{b}"]=label; ws[f"B{b}"].font=FN; ws[f"B{b}"].border=BD; ws[f"B{b}"].alignment=L
    ws[f"C{b}"]=unit; ws[f"C{b}"].font=FU; ws[f"C{b}"].border=BD; ws[f"C{b}"].alignment=C
    c=ws[f"D{b}"]; c.value="="+fml; c.number_format=fmt; c.border=BD; c.alignment=Rr; c.font=FN
    b+=1
def br(key): return f"$D${brow[key]}"
TAX=pc('tax'); DSF=pc('dsf'); SS=pc('S'); DISC=pc('disc')
# торнадо
tr=b+1; section(ws,tr,"ТОРНАДО: ВЛИЯНИЕ ДРАЙВЕРОВ НА ОБЩИЙ АННУИТЕТ (руб/м³)","B","H"); tr+=1
for i,h in enumerate(["Драйвер","Диапазон","При снижении","При росте","Размах"]):
    c=ws.cell(row=tr,column=2+i,value=h); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
tr+=1
drivers=[("Расход / цена ДТ",0.20,f"{br('b_fuel')}*(1-{TAX})"),
         ("Затраты на ТОиР/ремонты",0.25,f"{br('b_maint')}*(1-{TAX})"),
         ("Цена приобретения",0.15,f"{br('b_inv')}*(1-{TAX}*{DSF})"),
         ("Расходы на персонал",0.15,f"{br('b_pers')}*(1-{TAX})")]
ttop=tr
for label,pct,net in drivers:
    ws.cell(row=tr,column=2,value=label).font=FN; ws.cell(row=tr,column=2).border=BD; ws.cell(row=tr,column=2).alignment=L
    ws.cell(row=tr,column=3,value=f"±{int(pct*100)}%").alignment=C; ws.cell(row=tr,column=3).border=BD; ws.cell(row=tr,column=3).font=FN
    ws.cell(row=tr,column=4,value=f"={br('b_tot')}-{pct}*({net})")
    ws.cell(row=tr,column=5,value=f"={br('b_tot')}+{pct}*({net})")
    ws.cell(row=tr,column=6,value=f"=E{tr}-D{tr}")
    for col in (4,5,6):
        cc=ws.cell(row=tr,column=col); cc.number_format=M2; cc.border=BD; cc.alignment=Rr; cc.font=FN
    tr+=1
tbot=tr-1
ws.cell(row=ttop-1,column=9,value="низкое").font=FU; ws.cell(row=ttop-1,column=10,value="высокое").font=FU
for i in range(ttop,tbot+1):
    ws.cell(row=i,column=9,value=f"=D{i}-{br('b_tot')}").number_format=M2
    ws.cell(row=i,column=10,value=f"=E{i}-{br('b_tot')}").number_format=M2
tch=BarChart(); tch.type="bar"; tch.grouping="stacked"; tch.overlap=100
tch.title="Торнадо: отклонение аннуитета, руб/м³"; tch.height=6.5; tch.width=15
dl=Reference(ws,min_col=9,min_row=ttop,max_col=9,max_row=tbot); dh=Reference(ws,min_col=10,min_row=ttop,max_col=10,max_row=tbot)
cats=Reference(ws,min_col=2,min_row=ttop,max_col=2,max_row=tbot)
tch.add_data(dl); tch.add_data(dh); tch.set_categories(cats)
tch.series[0].graphicalProperties.solidFill="70AD47"; tch.series[1].graphicalProperties.solidFill="C55A11"; tch.legend=None
ws.add_chart(tch,"J6")
# КТГ
ur=tbot+3; section(ws,ur,"ВЛИЯНИЕ ПРОИЗВОДИТЕЛЬНОСТИ (КТГ / УТИЛИЗАЦИИ)","B","H"); ur+=1
ws.cell(row=ur,column=2,value="Изменение производительности").font=FHd
ws.cell(row=ur,column=2).fill=Fh; ws.cell(row=ur,column=2).border=BD; ws.cell(row=ur,column=2).alignment=C
levels=[-0.15,-0.10,0.0,0.10,0.15]
for j,d in enumerate(levels):
    c=ws.cell(row=ur,column=3+j,value=f"{'+' if d>0 else ''}{int(d*100)}%"); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
ur+=1
ws.cell(row=ur,column=2,value="Общий аннуитет, руб/м³").font=FB; ws.cell(row=ur,column=2).border=BD; ws.cell(row=ur,column=2).alignment=L
for j,d in enumerate(levels):
    c=ws.cell(row=ur,column=3+j,value=f"={br('b_fix')}/(1+{d})+{br('b_var')}"); c.number_format=M2; c.border=BD; c.alignment=Rr; c.font=FB
    if abs(d)<1e-9: c.fill=Fr
ur+=2
# ставка дисконтирования
section(ws,ur,"ВЛИЯНИЕ СТАВКИ ДИСКОНТИРОВАНИЯ","B","H"); ur+=1
ws.cell(row=ur,column=2,value="Ставка дисконтирования").font=FHd
ws.cell(row=ur,column=2).fill=Fh; ws.cell(row=ur,column=2).border=BD; ws.cell(row=ur,column=2).alignment=C
rates=[0.08,0.10,0.12,0.14,0.16]
for j,rt in enumerate(rates):
    c=ws.cell(row=ur,column=3+j,value=rt); c.number_format=P1; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
ur+=1
ws.cell(row=ur,column=2,value="Общий аннуитет, руб/м³ (оценка)").font=FB; ws.cell(row=ur,column=2).border=BD; ws.cell(row=ur,column=2).alignment=L
# инвестиц. аннуитет пересчитывается точно, операционный ≈ постоянный
for j,rt in enumerate(rates):
    col=get_column_letter(3+j)
    pvaf=f"(1-(1+{col}{ur-1})^-{pc('horizon')})/{col}{ur-1}"
    # инвест. аннуитет(r) = price*(1 - resid*DF10(r))/PVAF(r,N)/avgprod ; resid выбранного варианта
    resid_sel=f"INDEX('{S_IN}'!D{irow['resid']}:I{irow['resid']},{SEL})"
    inv_ann=f"{br('b_price')}*(1-{resid_sel}/(1+{col}{ur-1})^{pc('horizon')})/({pvaf})/{br('b_am3')}"
    c=ws.cell(row=ur,column=3+j,value=f"={inv_ann}+{br('b_tot')}-{br('b_inv')}")
    c.number_format=M2; c.border=BD; c.alignment=Rr; c.font=FB
    if abs(rt-0.12)<1e-9: c.fill=Fr
ur+=2
ws.cell(row=ur,column=2,value="Оценка по ставке: инвестиционный аннуитет пересчитан точно, операционный принят "
        "постоянным (аннуитет равномерного потока ≈ инвариантен к ставке).").font=FU
ws.merge_cells(f"B{ur}:H{ur}"); ws[f"B{ur}"].alignment=L
print("Чувствительность готова")

# ================================================================= #
# АНАЛИТИКА (расширенный анализ результатов, лучшие практики)
# ================================================================= #
wan=wb.create_sheet(S_ANL); wan.sheet_view.showGridLines=False
title(wan,"РАСШИРЕННАЯ АНАЛИТИКА РЕЗУЛЬТАТОВ","Многокритериальная оценка, разложение затрат, диапазон риска, эффективность.",last="K")
wan.column_dimensions["A"].width=2; wan.column_dimensions["B"].width=40; wan.column_dimensions["C"].width=10
for cc in EXC: wan.column_dimensions[cc].width=15
NHR=4
wan[f"B{NHR}"]="Показатель"; wan[f"B{NHR}"].font=FHd; wan[f"B{NHR}"].fill=Fh; wan[f"B{NHR}"].border=BD; wan[f"B{NHR}"].alignment=C
wan[f"C{NHR}"]="Ед."; wan[f"C{NHR}"].font=FHd; wan[f"C{NHR}"].fill=Fh; wan[f"C{NHR}"].border=BD; wan[f"C{NHR}"].alignment=C
for i,cc in enumerate(EXC):
    c=wan[f"{cc}{NHR}"]; c.value=f"='{S_IN}'!{cc}{irow['name']}"; c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
an={}; arow=NHR+1
def aline(key,label,unit,fni,fmt=M2,best="min",res=False):
    global arow; an[key]=arow
    b=wan[f"B{arow}"]; b.value=label; b.border=BD; b.alignment=L; b.font=(FRz if res else FN)
    wan[f"C{arow}"]=unit; wan[f"C{arow}"].font=FU; wan[f"C{arow}"].border=BD; wan[f"C{arow}"].alignment=C
    for i,cc in enumerate(EXC):
        c=wan[f"{cc}{arow}"]; c.value="="+fni(i); c.number_format=fmt; c.border=BD; c.alignment=Rr; c.font=(FRz if res else FN)
        if res: c.fill=Fr
    arow+=1
def asec2(t):
    global arow; section(wan,arow,t,"B","I"); arow+=1
def acol(i): return EXC[i]
def A(key,i): return annref(key,acol(i))
def ktg_avg(i): return f"AVERAGE('{S_PYR}'!E{pyr_ktg[i]}:N{pyr_ktg[i]})"

asec2("МНОГОКРИТЕРИАЛЬНАЯ ОЦЕНКА (levelized cost + драйверы)")
aline("price","Цена (CAPEX)","тыс.руб",lambda i:prodref('price_rub',acol(i)),M)
aline("prod","Средний годовой объём","тыс.м³",lambda i:A('avgprod',i),M)
aline("ktg","КТГ средний","коэф.",lambda i:ktg_avg(i),P1)
aline("m3","Аннуитет","руб/м³",lambda i:A('u_tot',i),M2,res=True)
aline("t","Аннуитет","руб/т",lambda i:A('t_tot',i),M2)
aline("npv","NPV стоимости владения","тыс.руб",lambda i:A('n_tot',i),M)
aline("s_cap","Доля CAPEX","%",lambda i:f"{A('u_inv',i)}/{A('u_tot',i)}",P1)
aline("s_fuel","Доля топлива","%",lambda i:f"{A('u_fuel',i)}/{A('u_tot',i)}",P1)
aline("s_maint","Доля ТОиР","%",lambda i:f"{A('u_maint',i)}/{A('u_tot',i)}",P1)
m3rng=f"D{an['m3']}:I{an['m3']}"
aline("rank","Место в рейтинге","",lambda i:f"RANK({acol(i)}{an['m3']},{m3rng},1)",M,res=True)

asec2("РАЗЛОЖЕНИЕ ПРЕВЫШЕНИЯ АННУИТЕТА НАД ЛУЧШИМ (руб/м³)")
def dvs(key): return lambda i:f"{A(key,i)}-MIN(D{ar[key]}:I{ar[key]})"
aline("d_inv","Δ инвестиции",  "руб/м³",dvs('u_inv'),M2)
aline("d_fuel","Δ топливо",    "руб/м³",dvs('u_fuel'),M2)
aline("d_maint","Δ ТОиР",       "руб/м³",dvs('u_maint'),M2)
aline("d_pers","Δ персонал",   "руб/м³",dvs('u_pers'),M2)
aline("d_tax","Δ налог (щит)", "руб/м³",dvs('u_tax'),M2)
aline("d_tot","ИТОГО превышение над лучшим","руб/м³",
      lambda i:f"{acol(i)}{an['m3']}-MIN({m3rng})",M2,res=True)

asec2("ДИАПАЗОН РИСКА АННУИТЕТА (руб/м³)")
# fixed = инвест+ТОиР+персонал+налог; var = топливо+ковш
def fixed(i): return f"({A('u_inv',i)}+{A('u_maint',i)}+{A('u_pers',i)}+{A('u_tax',i)})"
def var(i): return f"({A('u_fuel',i)}+{A('u_bucket',i)})"
aline("opt","Оптимистичный (КТГ +10%, ДТ −20%)","руб/м³",
      lambda i:f"{fixed(i)}/1.1+{var(i)}-0.2*{A('u_fuel',i)}",M2)
aline("base","Базовый","руб/м³",lambda i:A('u_tot',i),M2,res=True)
aline("pess","Пессимистичный (КТГ −10%, ДТ +20%)","руб/м³",
      lambda i:f"{fixed(i)}/0.9+{var(i)}+0.2*{A('u_fuel',i)}",M2)
aline("band","Ширина диапазона (риск)","руб/м³",
      lambda i:f"{acol(i)}{an['pess']}-{acol(i)}{an['opt']}",M2)

asec2("ЭФФЕКТИВНОСТЬ И ЦЕННОСТЬ ГОТОВНОСТИ")
aline("fuel_kgm3","Топливоёмкость","кг/м³",lambda i:f"{A('u_fuel',i)}/{pc('fuel_p')}",M2,best="min")
aline("fuel_rub","Топливо на м³","руб/м³",lambda i:A('u_fuel',i),M2)
aline("val_ktg","Эффект +1 п.п. КТГ на аннуитет","руб/м³",
      lambda i:f"-{fixed(i)}*0.01/{ktg_avg(i)}",M2)

# вывод (формульный)
arow+=1
names_rng=f"D{NHR}:I{NHR}"
wan[f"B{arow}"]="ВЫВОД:"; wan[f"B{arow}"].font=Font(bold=True,size=12,color=DARK)
wan.merge_cells(f"C{arow}:I{arow}")
concl=wan[f"C{arow}"]
concl.value=(f'="Лучший по удельному TCO: "&INDEX({names_rng},MATCH(MIN({m3rng}),{m3rng},0))'
             f'&" ("&TEXT(MIN({m3rng}),"0.00")&" руб/м³, "&TEXT(MIN(D{an["t"]}:I{an["t"]}),"0.00")&" руб/т). "'
             f'&"Отрыв от худшего: "&TEXT(MAX({m3rng})-MIN({m3rng}),"0.00")&" руб/м³."')
concl.font=Font(bold=True,size=11,color="006100"); concl.fill=Fb; concl.alignment=L
for cc in "CDEFGHI": wan[f"{cc}{arow}"].border=BD
concl_row=arow; arow+=2
# лучшие практики (пояснение метода)
wan[f"B{arow}"]="Метод (лучшие практики):"; wan[f"B{arow}"].font=FB; arow+=1
for t in ["Levelized cost — приведённые затраты на единицу продукции (руб/м³, руб/т) уравнивают машины с разной ценой и производительностью.",
          "Многокритериальная оценка — не только цена, но и структура затрат, КТГ, топливоёмкость.",
          "Разложение над лучшим — показывает, за счёт каких статей вариант дороже (управляемость).",
          "Диапазон риска — устойчивость выбора к колебаниям КТГ и цены топлива (сценарии P-).",
          "Ценность готовности — сколько руб/м³ даёт рост КТГ на 1 п.п. (обоснование сервисного контракта)."]:
    wan.merge_cells(f"B{arow}:K{arow}"); wan[f"B{arow}"]="•  "+t; wan[f"B{arow}"].font=FU; wan[f"B{arow}"].alignment=L; arow+=1

# диаграммы
# 1) scatter: производительность vs аннуитет
sc=ScatterChart(); sc.title="Аннуитет vs производительность"; sc.height=8; sc.width=12
sc.x_axis.title="Годовой объём, тыс.м³"; sc.y_axis.title="Аннуитет, руб/м³"; sc.x_axis.delete=False; sc.y_axis.delete=False
xref=Reference(wan,min_col=4,min_row=an['prod'],max_col=9,max_row=an['prod'])
yref=Reference(wan,min_col=4,min_row=an['m3'],max_col=9,max_row=an['m3'])
ser=Series(yref,xref,title="варианты"); ser.marker.symbol="circle"; ser.marker.size=8; ser.graphicalProperties.line.noFill=True
sc.series.append(ser); sc.legend=None
wan.add_chart(sc,"K4")
# 2) диапазон риска (опт/база/песс)
rc=BarChart(); rc.type="col"; rc.title="Диапазон риска аннуитета, руб/м³"; rc.height=8; rc.width=13; rc.y_axis.title="руб/м³"
cats=Reference(wan,min_col=4,min_row=NHR,max_col=9,max_row=NHR)
for key,lbl in [("opt","оптимистичный"),("base","базовый"),("pess","пессимистичный")]:
    ref=Reference(wan,min_col=4,min_row=an[key],max_col=9,max_row=an[key])
    s=Series(ref,title=lbl); rc.series.append(s)
rc.set_categories(cats)
wan.add_chart(rc,"K22")
print("Аналитика готова")

# ================================================================= #
# СПРАВОЧНИК
# ================================================================= #
wr=wb.create_sheet(S_REF); wr.sheet_view.showGridLines=False
title(wr,"СПРАВОЧНИК ОРИЕНТИРОВОЧНЫХ ЗНАЧЕНИЙ","Диапазоны для проверки исходных данных. Требуют уточнения по данным поставщика.",last="I")
wr.column_dimensions["A"].width=2; wr.column_dimensions["B"].width=34
for cc in "CDEFGHI": wr.column_dimensions[cc].width=16
heads=["Показатель","Класс 2000 (~12 м³)","Диапазон","Комментарий"]
for i,h in enumerate(heads):
    c=wr.cell(row=4,column=2+i,value=h); c.font=FHd; c.fill=Fh; c.border=BD; c.alignment=C
refdata=[
 ("Ставка дисконтирования (WACC)","12%","8–16%","Согласуется с эскалацией (реальн./номин.)"),
 ("Налог на прибыль","25%","20–25%","Действующая ставка РФ"),
 ("КТГ (техготовность)","0,74–0,84","0,70–0,90","Зависит от бренда и сервиса"),
 ("КИО (использование)","0,47–0,58","0,40–0,65","Доля рабочего времени в КФВ"),
 ("Удельный расход ДТ","85–120","70–140 кг/ч","Двигатель и режим"),
 ("Срок амортизации (НУ)","62 мес","5–7 лет","Линейная амортизация"),
 ("Остаточная стоимость","10%","8–20%","Реализация в конце срока"),
 ("Время цикла экскавации","37 сек","28–45 сек","Черпание–поворот–разгрузка"),
 ("Расходники на ковш","0,87","0,5–1,5 руб/м³","Зубья, коронки, фиксаторы"),
 ("Эскалация ДТ / ТОиР / ФОТ","4 / 5 / 6%","3–8%/год","Обычно выше общей инфляции"),
]
rr=5
for row in refdata:
    for i,v in enumerate(row):
        c=wr.cell(row=rr,column=2+i,value=v); c.font=FN; c.border=BD; c.alignment=(L if i in(0,3) else C)
        if rr%2==0: c.fill=Fg
    rr+=1
rr+=1
wr.cell(row=rr,column=2,value="Методические примечания:").font=FB; rr+=1
for n in ["Годовые затраты можно задавать усреднённо (в ТОиР) либо разнести капремонты по графику (раздел G ввода).",
          "Расход ДТ — ключевой драйвер OPEX; уточняйте по факту двигателя, а не по паспорту.",
          "Значения таблицы — ориентир порядка величин, а не норматив."]:
    wr.cell(row=rr,column=2,value="•  "+n).font=FN; wr.cell(row=rr,column=2).alignment=L; wr.merge_cells(f"B{rr}:I{rr}"); rr+=1
print("Справочник готов")

# ================================================================= #
# МЕТОДИКА (первым)
# ================================================================= #
wme=wb.create_sheet(S_MET,0); wme.sheet_view.showGridLines=False
title(wme,"МЕТОДИКА: АННУИТЕТ ЗАТРАТ (РУБ/М³) НА ВЛАДЕНИЕ ЭКСКАВАТОРАМИ","Пост-налоговый погодовой дисконтированный аннуитет жизненного цикла.")
wme.column_dimensions["A"].width=2; wme.column_dimensions["B"].width=4; wme.column_dimensions["C"].width=122
def mh(row,t):
    c=wme.cell(row=row,column=3,value=t); c.font=FSec; c.fill=Fm; c.alignment=L; wme.cell(row=row,column=2).fill=Fm
def mt(row,t,h=None):
    c=wme.cell(row=row,column=3,value=t); c.font=FN; c.alignment=Alignment("left","top",wrap_text=True)
    if h: wme.row_dimensions[row].height=h
blocks=[
 (4,"h","1. НАЗНАЧЕНИЕ"),
 (5,"t","Экономическое сравнение экскаваторов одного класса по удельному аннуитету затрат (руб/м³ вынутой горной "
       "массы) за полный срок эксплуатации. Метрика корректно сопоставляет машины с разной ценой, "
       "производительностью и структурой затрат. Лучший вариант — минимум руб/м³.",50),
 (6,"h","2. ПРОИЗВОДИТЕЛЬНОСТЬ"),
 (7,"t","Порода в ковше = ковш × плотность × (наполнение/разрыхление) × потери. Ковшей на самосвал = "
       "ОКРУГЛ.ВВЕРХ(грузоподъёмность / порода в ковше). Самосвалов/час = 3600/(ковшей × цикл + постановка а/с). "
       "т/ч = грузоподъёмность × самосвалов/час; м³/ч = т/ч / плотность. Годовой объём = м³/ч × эффективное время; "
       "эффективное время = КФВ × КТГ × (1 − простои/смена). КТГ снижается с возрастом машины.",60),
 (8,"h","3. ПОГОДОВОЙ ДЕНЕЖНЫЙ ПОТОК (10 лет)"),
 (9,"t","Для каждого года считаются: дизтопливо (эфф.время × кг/ч × цена), ТОиР и сервис (задаются ПО ГОДАМ на листе "
       "«Данные по годам» — включают плановое ТО, текущие и капитальные ремонты; годы капремонтов дают всплеск затрат), "
       "расходники на ковш, персонал. Добавляются амортизация (НУ) и налог на прибыль (щит). Операционный поток = база + "
       "налог − амортизация. Инвестиции — в год 0, остаточная стоимость — в конце срока. КТГ также задаётся по годам.",65),
 (10,"h","4. АННУИТЕТ (ЯДРО)"),
 (11,"t","Потоки дисконтируются и преобразуются в эквивалентный годовой платёж: Аннуитет = NPV затрат / Σ(дисконт-"
        "факторов). Инвестиционный аннуитет = (Цена − дисконт. остаточная стоимость) / Σ(дисконт-факторов). "
        "Операционный аннуитет — приведённые эксплуатационные затраты после налога. Общий = инвестиционный + "
        "операционный. Удельный аннуитет (руб/м³) = общий аннуитет / средний годовой объём; руб/т = руб/м³ / плотность.",65),
 (12,"h","5. ПОРЯДОК РАБОТЫ"),
 (13,"t","«Параметры» → «Ввод данных» (цена, ковш, персонал) → листы «Поставщик 1…6»: по каждой машине встроены модели "
        "ТО (интервалы обслуживания → простои, запчасти, сервис) и Ремонтов (простои, текущие и капитальные ремонты, "
        "сервис по годам) — КТГ и ИТОГО ТОиР считаются автоматически снизу вверх. Сводка «Данные по годам» тянет итоги "
        "сама → авторасчёт «Производительность», «Денежный поток», «Расчёт аннуитета» → «Дашборд», «Сравнение», "
        "«Аналитика» (многокритериальная оценка, разложение над лучшим, диапазон риска, ценность готовности), "
        "«Чувствительность». Правки — только в листах «Поставщик N».",70),
 (14,"h","6. ДОПУЩЕНИЯ"),
 (15,"t","• Машины сравниваются в одном классе и одной задаче; разница в производительности учтена через руб/м³.",None),
 (16,"t","• Капремонты можно учитывать усреднённо (в ТОиР) или разнести по графику (раздел G ввода).",None),
 (17,"t","• Ставка дисконтирования и эскалации должны быть согласованы (номинальные либо реальные).",None),
 (18,"t","• Методика воспроизводит корпоративный подход «аннуитета руб/м³» с погодовым дисконтированием, амортизацией и "
        "налоговым щитом.",30),
]
for it in blocks:
    row,kind=it[0],it[1]
    if kind=="h": mh(row,it[2])
    else: mt(row,it[2],*( (it[3],) if len(it)>3 and it[3] else () ))
lg=20
wme.cell(row=lg,column=3,value="Обозначения:").font=FB
wme.cell(row=lg+1,column=2).fill=Fi; wme.cell(row=lg+1,column=2).border=BD
wme.cell(row=lg+1,column=3,value="жёлтые ячейки — ввод данных").font=FN
wme.cell(row=lg+2,column=2).fill=Fr; wme.cell(row=lg+2,column=2).border=BD
wme.cell(row=lg+2,column=3,value="зелёные ячейки — итоговые показатели / лучший вариант").font=FN

# порядок листов и сохранение
order=[S_MET,S_DASH,S_PAR,S_IN]+TPL_NAMES+[S_PYR,S_PROD,S_CF,S_ANN,S_CMP,S_ANL,S_SENS,S_REF]
wb._sheets.sort(key=lambda s:order.index(s.title))
wb.active=1     # Дашборд
wb.calculation.fullCalcOnLoad=True
OUT="Модель_TCO_экскаваторов_аннуитет.xlsx"
wb.save(OUT)
print("СОХРАНЕНО:",OUT)
